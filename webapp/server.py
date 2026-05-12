"""
쿠팡 로켓배송 자동화 - FastAPI 서버
3순위: 1단계 실행 + WebSocket 실시간 로그 + 체크리스트.
"""
import asyncio
import json
import threading
import time
import uuid
from collections import deque
from pathlib import Path

from fastapi import FastAPI, UploadFile, File, HTTPException, WebSocket, WebSocketDisconnect, Body
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel
import uvicorn

BASE_DIR = Path(__file__).parent
PROJECT_ROOT = BASE_DIR.parent
STATIC_DIR = BASE_DIR / "static"
LOGS_DIR = BASE_DIR / "logs"
ENV_FILE = BASE_DIR / ".env"
CREDENTIALS_FILE = BASE_DIR / "credentials.json"
STAGE1_COMPLETED_FILE = BASE_DIR / ".stage1_completed"
FAVORITES_FILE = BASE_DIR / "category_favorites.json"

LOGS_DIR.mkdir(exist_ok=True)
for folder in ("templates", "fonts", "chrome_profile", "temp_downloads", "완료된_상품", "input"):
    (PROJECT_ROOT / folder).mkdir(exist_ok=True)

DEFAULTS = {
    "TARGET_SHEET_NAME": "상품등록목록",
    "EXCHANGE_FACTOR": "350",
    "ADD_LOGISTICS_COST": "4000",
    "SALE_PRICE_RATE": "1.67",
    "REC_PRICE_RATE": "1.30",
    "ROUND_UNIT": "-2",
}


def load_env() -> dict:
    if not ENV_FILE.exists():
        return {}
    data = {}
    for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        data[k.strip()] = v.strip()
    return data


def save_env(data: dict):
    lines = [f"{k}={v}" for k, v in data.items()]
    ENV_FILE.write_text("\n".join(lines) + "\n", encoding="utf-8")


def load_favorites() -> list:
    """카테고리 즐겨찾기 목록을 JSON에서 읽어 반환."""
    if not FAVORITES_FILE.exists():
        return []
    try:
        data = json.loads(FAVORITES_FILE.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else []
    except Exception:
        return []


def save_favorites(favs: list):
    FAVORITES_FILE.write_text(
        json.dumps(favs, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def safe_tag(s: str) -> str:
    """카테고리 라벨을 파일명/glob 패턴에 쓸 수 있도록 정규화.

    Windows 금지문자(/ \\ : * ? " < > |)를 _ 로 치환하고 연속된 _ 는 하나로 압축.
    'stored 파일명' 과 'glob 검색 패턴' 양쪽에 반드시 동일하게 적용해야 재매치가 성립.
    """
    import re as _re
    s = (s or "").strip()
    s = _re.sub(r'[/\\:*?"<>|]', '_', s)
    s = _re.sub(r'_+', '_', s)
    return s.strip("_")


def log_to_file(entry: dict):
    """stage1 이벤트를 날짜별 로그 파일에 append."""
    date = time.strftime("%Y-%m-%d")
    logfile = LOGS_DIR / f"{date}.log"
    try:
        ts = time.strftime("%H:%M:%S", time.localtime(entry.get("ts", time.time())))
        line = f"[{ts}] [{entry.get('type', 'log')}] {entry.get('msg', '')}\n"
        with logfile.open("a", encoding="utf-8") as f:
            f.write(line)
    except Exception:
        pass


class SettingsPayload(BaseModel):
    google_sheet_url: str = ""
    target_sheet_name: str = "상품등록목록"
    gemini_api_key: str = ""
    supply_id: str = ""
    supply_pw: str = ""
    my_brand_name: str = ""
    my_company_name: str = ""
    my_phone_number: str = ""
    exchange_factor: str = "350"
    add_logistics_cost: str = "4000"
    sale_price_rate: str = "1.67"
    rec_price_rate: str = "1.30"
    round_unit: str = "-2"


class ChecklistPayload(BaseModel):
    item1: bool
    item2: bool
    item3: bool


class Stage1Runner:
    def __init__(self):
        self.lock = threading.Lock()
        self.running = False
        self.log_buffer = deque(maxlen=5000)
        self.stop_event = threading.Event()
        self.thread = None
        self.status = "idle"  # idle | running | completed | error | stopped
        self.current_step = ""
        self.progress = {"step": 0, "total": 7, "label": ""}
        self.run_id = None  # 재연결 시 done 중복 처리 방지용 고유 ID

    def _emit(self, type_, **kwargs):
        # done 이벤트에는 현재 실행의 run_id 를 자동 주입 (클라이언트 dedup 용)
        if type_ == "done" and "run_id" not in kwargs and self.run_id:
            kwargs["run_id"] = self.run_id
        entry = {"type": type_, "ts": time.time(), **kwargs}
        with self.lock:
            self.log_buffer.append(entry)
        log_to_file(entry)

    def start(self, config):
        with self.lock:
            if self.running:
                return False
            self.running = True
            self.stop_event.clear()
            self.log_buffer.clear()
            self.status = "running"
            self.current_step = "시작 중..."
            self.progress = {"step": 0, "total": 7, "label": ""}
            self.run_id = uuid.uuid4().hex
        # 새 실행 시 완료 마커 제거 (체크리스트 초기화)
        if STAGE1_COMPLETED_FILE.exists():
            try: STAGE1_COMPLETED_FILE.unlink()
            except Exception: pass
        self.thread = threading.Thread(target=self._run, args=(config,), daemon=True)
        self.thread.start()
        return True

    def stop(self):
        self.stop_event.set()
        self._emit("log", level="warn", msg="⏹️ 중지 요청됨. 현재 단계 종료 후 중단됩니다.")

    def _run(self, config):
        try:
            from modules.stage1 import run_stage1, Stopped

            def log(msg, level="info"):
                self._emit("log", level=level, msg=str(msg))

            def progress(step, total, label):
                with self.lock:
                    self.current_step = label
                    self.progress = {"step": step, "total": total, "label": label}
                self._emit("progress", step=step, total=total, label=label)

            def should_stop():
                return self.stop_event.is_set()

            run_stage1(config, log, progress, should_stop)
            with self.lock:
                self.status = "completed"
            self._emit("done", ok=True)

        except Exception as e:
            try:
                from modules.stage1 import Stopped as _Stopped
            except Exception:
                _Stopped = None
            if _Stopped is not None and isinstance(e, _Stopped):
                with self.lock:
                    self.status = "stopped"
                self._emit("done", ok=False, error="사용자가 중지했습니다.")
            else:
                import traceback
                tb = traceback.format_exc()
                with self.lock:
                    self.status = "error"
                self._emit("log", level="error", msg=f"❌ 에러: {e}")
                self._emit("log", level="error", msg=tb)
                self._emit("done", ok=False, error=str(e))
        finally:
            with self.lock:
                self.running = False


runner = Stage1Runner()


class Stage2Runner:
    def __init__(self):
        self.lock = threading.Lock()
        self.running = False
        self.log_buffer = deque(maxlen=10000)
        self.stop_event = threading.Event()
        self.thread = None
        self.status = "idle"  # idle | running | waiting_gate | completed | error | stopped
        self.current_step = ""
        self.progress = {"step": 0, "total": 9, "label": ""}
        self.counter = {"current": 0, "limit": 50}
        self.pending_gate = None              # 현재 대기 중인 게이트 이름 (없으면 None)
        self.pending_gate_payload = None      # 게이트와 함께 전달되는 페이로드 (재연결 복원용) 20260512
        self.gate_events = {}                 # name -> threading.Event
        self.run_id = None  # 재연결 시 done 중복 처리 방지용 고유 ID

    def _emit(self, type_, **kwargs):
        if type_ == "done" and "run_id" not in kwargs and self.run_id:
            kwargs["run_id"] = self.run_id
        entry = {"type": type_, "ts": time.time(), **kwargs}
        with self.lock:
            self.log_buffer.append(entry)
        log_to_file(entry)

    def start(self, config):
        with self.lock:
            if self.running:
                return False
            self.running = True
            self.stop_event.clear()
            self.log_buffer.clear()
            self.status = "running"
            self.current_step = "시작 중..."
            self.progress = {"step": 0, "total": 9, "label": ""}
            self.counter = {"current": 0, "limit": 50}
            self.pending_gate = None
            self.gate_events = {}
            self.run_id = uuid.uuid4().hex
        self.thread = threading.Thread(target=self._run, args=(config,), daemon=True)
        self.thread.start()
        return True

    def stop(self):
        self.stop_event.set()
        # 대기 중인 게이트가 있으면 깨워서 Stopped 예외로 빠져나가게 함
        with self.lock:
            for ev in self.gate_events.values():
                ev.set()
        self._emit("log", level="warn", msg="⏹️ 중지 요청됨. 현재 단계 종료 후 중단됩니다.")

    def release_gate(self, name):
        with self.lock:
            ev = self.gate_events.get(name)
            if self.pending_gate == name:
                self.pending_gate = None
                self.pending_gate_payload = None
            if self.status == "waiting_gate":
                self.status = "running"
        if ev is not None:
            ev.set()
            return True
        return False

    def _run(self, config):
        try:
            from modules.stage2 import run_stage2, Stopped

            def log(msg, level="info"):
                self._emit("log", level=level, msg=str(msg))

            def progress(step, total, label):
                with self.lock:
                    self.current_step = label
                    self.progress = {"step": step, "total": total, "label": label}
                self._emit("progress", step=step, total=total, label=label)

            def should_stop():
                return self.stop_event.is_set()

            def set_counter(current, limit):
                with self.lock:
                    self.counter = {"current": int(current), "limit": int(limit)}
                self._emit("counter", current=int(current), limit=int(limit))

            def gate_signal(name, payload=None):
                # payload 20260512: 견적서 게이트 등에서 파일 정보를 UI 로 전달.
                # 기본 None 이라 기존 호출(`gate_signal("...")`) 도 그대로 동작.
                with self.lock:
                    self.gate_events[name] = threading.Event()
                    self.pending_gate = name
                    self.pending_gate_payload = payload
                    self.status = "waiting_gate"
                self._emit("gate", name=name, payload=payload)

            def gate_wait(name):
                with self.lock:
                    ev = self.gate_events.get(name)
                if ev is None:
                    return
                # 폴링 대기: stop 요청이 오면 Stopped 로 빠져나감
                while not ev.wait(timeout=0.5):
                    if self.stop_event.is_set():
                        from modules.stage2 import Stopped as _S
                        raise _S()

            run_stage2(config, log, progress, should_stop,
                       set_counter=set_counter,
                       gate_signal=gate_signal,
                       gate_wait=gate_wait)
            with self.lock:
                self.status = "completed"
            self._emit("done", ok=True)

        except Exception as e:
            try:
                from modules.stage2 import Stopped as _Stopped
            except Exception:
                _Stopped = None
            if _Stopped is not None and isinstance(e, _Stopped):
                with self.lock:
                    self.status = "stopped"
                self._emit("done", ok=False, error="사용자가 중지했습니다.")
            else:
                import traceback
                tb = traceback.format_exc()
                with self.lock:
                    self.status = "error"
                self._emit("log", level="error", msg=f"❌ 에러: {e}")
                self._emit("log", level="error", msg=tb)
                self._emit("done", ok=False, error=str(e))
        finally:
            with self.lock:
                self.running = False


runner2 = Stage2Runner()

# ──────────────────────────────────────────────────────────────
# 카테고리 추가 등록 Runner
# ──────────────────────────────────────────────────────────────

class CategoryRunner:
    """완료 탭 상품을 다른 카테고리 견적서로 재등록"""
    def __init__(self):
        self.lock = threading.Lock()
        self.running = False
        self.log_buffer = deque(maxlen=5000)
        self.stop_event = threading.Event()
        self.thread = None
        self.status = "idle"
        self.progress = {"step": 0, "total": 4, "label": ""}
        self.pending_gate = None
        self.gate_events = {}
        # 검색(/api/category/search)과 실행(/api/category/start) 간에 Chrome 세션을 공유한다.
        # search 호출에서 Chrome을 띄우고 모달까지 열어둔 뒤, start가 같은 driver를 재사용한다.
        self.search_driver = None
        self.search_wait = None
        self.driver_lock = threading.Lock()
        # 정상 완료 경로 플래그. True 이면 finally 에서 _shutdown_driver 를 생략하고
        # 사용자가 크롬 창에서 수동 업로드할 수 있도록 세션을 유지한다.
        self._normal_complete = False
        self.run_id = None  # 재연결 시 done 중복 처리 방지용 고유 ID

    def _emit(self, type_, **kwargs):
        if type_ == "done" and "run_id" not in kwargs and self.run_id:
            kwargs["run_id"] = self.run_id
        entry = {"type": type_, "ts": time.time(), **kwargs}
        with self.lock:
            self.log_buffer.append(entry)
        log_to_file(entry)

    def start(self, config, products, selection):
        """
        products: [{"name": "상품명", "path": "폴더경로"}, ...]
        selection: {
            "keyword": str,         # 쿠팡허브 검색어
            "category_id": str,     # 후보 중 매칭할 카테고리 ID (빈 값이면 path로 매칭)
            "path": str,            # "대분류 > 중분류 > ..." 전체 경로 (레이블/매칭용)
            "save_as_favorite": bool,
            "favorite_label": str,  # save_as_favorite=true일 때 라벨
        }
        """
        with self.lock:
            if self.running:
                return False
            self.running = True
            self.stop_event.clear()
            self.log_buffer.clear()
            self.status = "running"
            self.pending_gate = None
            self.gate_events = {}
            self._normal_complete = False  # 새 실행 시 플래그 초기화
            self.run_id = uuid.uuid4().hex
        self.thread = threading.Thread(
            target=self._run, args=(config, products, selection), daemon=True
        )
        self.thread.start()
        return True

    def stop(self):
        self.stop_event.set()
        with self.lock:
            for ev in self.gate_events.values():
                ev.set()

    def release_gate(self, name):
        with self.lock:
            ev = self.gate_events.get(name)
            if self.pending_gate == name:
                self.pending_gate = None
            if self.status == "waiting_gate":
                self.status = "running"
        if ev:
            ev.set()
            return True
        return False

    def _gate(self, name, payload=None):
        with self.lock:
            self.gate_events[name] = threading.Event()
            self.pending_gate = name
            self.status = "waiting_gate"
        if payload is not None:
            self._emit("gate", name=name, payload=payload)
        else:
            self._emit("gate", name=name)
        ev = self.gate_events[name]
        while not ev.wait(timeout=0.5):
            if self.stop_event.is_set():
                raise RuntimeError("사용자가 중지했습니다.")

    def _log(self, msg, level="info"):
        self._emit("log", level=level, msg=str(msg))

    def _prog(self, step, total, label):
        with self.lock:
            self.progress = {"step": step, "total": total, "label": label}
        self._emit("progress", step=step, total=total, label=label)

    def find_downloaded_file(self, patterns, timeout: int = 60,
                             started_at: float = None, config: dict = None):
        """여러 후보 디렉토리에서 완성된 파일을 찾아 Path 로 반환 (없으면 None).

        patterns: 글롭 또는 확장자 리스트. 예) ["*.zip", "*.xlsx"] / [".zip", ".xlsx"]
                  str 하나만 넘겨도 자동으로 리스트화.

        판정 조건:
          - 확장자가 patterns 로부터 도출된 허용 확장자에 해당 (.crdownload 제외)
          - 크기 > 0 바이트
          - mtime 가 현재 시각보다 1초 이상 과거 (= 최근 1초 내에 쓰기 없었음 → 안정)
          - started_at 지정 시 mtime >= started_at (이전 실행 잔재 배제)

        탐색 순서:
          1) PROJECT_ROOT/temp_downloads  (stage1 규약과 일관)
          2) ~/Downloads                  (chrome_profile 기본값)

        패턴은 받은 순서대로 디렉토리별로 순회 — 쿠팡허브 실 반환이 .zip 이므로
        호출부에서 .zip 을 먼저 둘 것 (["*.zip", "*.xlsx"]).
        """
        import os as _os
        import time as _time

        # patterns 정규화
        if isinstance(patterns, (list, tuple)):
            pat_list = list(patterns)
        else:
            pat_list = [patterns]
        normalized = []
        allowed_suffixes = set()
        for p in pat_list:
            if not p:
                continue
            # .xlsx 스타일 → *.xlsx
            if p.startswith("*"):
                norm = p
            elif p.startswith("."):
                norm = f"*{p}"
            else:
                norm = f"*.{p}"
            normalized.append(norm)
            idx = norm.rfind(".")
            if idx >= 0:
                allowed_suffixes.add(norm[idx:].lower())

        candidate_dirs = []
        project_root = (config or {}).get("PROJECT_ROOT")
        if project_root:
            candidate_dirs.append(Path(project_root) / "temp_downloads")
        candidate_dirs.append(Path(_os.path.expanduser("~")) / "Downloads")

        # 중복 제거(순서 유지)
        seen, unique = set(), []
        for d in candidate_dirs:
            key = str(d)
            if key not in seen:
                seen.add(key)
                unique.append(d)
        candidate_dirs = unique

        self._log(f"📂 검색 후보 위치: {[str(d) for d in candidate_dirs]}")
        self._log(f"🔎 검색 패턴: {normalized}")
        if started_at is not None:
            self._log(f"⏱️ 기준 시각: {_time.strftime('%H:%M:%S', _time.localtime(started_at))} (이전 생성 파일 제외)")

        start = _time.time()
        last_log = 0
        while _time.time() - start < timeout:
            for dir_path in candidate_dirs:
                if not dir_path.exists():
                    continue
                for pattern in normalized:
                    try:
                        matches = list(dir_path.glob(pattern))
                    except OSError:
                        continue
                    for f in matches:
                        try:
                            if f.suffix.lower() not in allowed_suffixes:
                                continue
                            if f.name.lower().endswith(".crdownload"):
                                continue
                            st = f.stat()
                        except OSError:
                            continue
                        if st.st_size <= 0:
                            continue
                        if started_at is not None and st.st_mtime < started_at:
                            continue
                        if (_time.time() - st.st_mtime) < 1.0:
                            # 아직 쓰기가 끝나지 않았을 수 있음 — 다음 라운드 대기
                            continue
                        self._log(f"📥 다운로드 파일 발견: {f.name}", "success")
                        self._log(f"   위치: {dir_path}")
                        return f

            elapsed = int(_time.time() - start)
            if elapsed >= last_log + 5:
                self._log(f"⏳ {elapsed}초 — 아직 못 찾음, 계속 검색 중...")
                last_log = elapsed
            _time.sleep(0.5)

        self._log(f"⏱️ {timeout}초 타임아웃: 어디서도 매치 파일 못 찾음", "error")
        self._log(f"   확인: ~/Downloads 또는 temp_downloads 에 수동으로 받아진 파일 있는지", "warn")
        return None

    def extract_excel_from_zip(self, zip_path: Path):
        """쿠팡허브가 내려주는 .zip 안에서 xlsx 1개를 바이트로 꺼내 반환.

        반환: (bytes, basename) 또는 (None, None) on failure.
        'bytes' 를 직접 받아 상품 폴더에 바로 쓰는 구조라, 임시폴더 cleanup 경쟁 없음.
        """
        import zipfile as _zf
        import os as _os

        self._log(f"📦 압축 파일 감지: {zip_path.name}")
        self._log(f"   압축 해제 중...")
        try:
            with _zf.ZipFile(str(zip_path), "r") as z:
                all_names = z.namelist()
                xlsx_names = [n for n in all_names if n.lower().endswith(".xlsx")]
                self._log(f"   압축 안 파일 수: {len(all_names)}")
                self._log(f"   .xlsx 파일: {len(xlsx_names)}개")
                for n in xlsx_names[:5]:
                    self._log(f"      - {n}")
                if not xlsx_names:
                    self._log(f"❌ zip 안에 .xlsx 파일이 없음", "error")
                    self._log(f"   전체 파일(처음 10개): {all_names[:10]}", "warn")
                    return None, None
                # "견적서" 또는 "quotation" 포함 우선, 없으면 첫 번째
                target = None
                for n in xlsx_names:
                    low = n.lower()
                    if ("견적서" in n) or ("quotation" in low):
                        target = n
                        break
                if target is None:
                    target = xlsx_names[0]
                self._log(f"   선택된 xlsx: {target}")
                data = z.read(target)
                basename = _os.path.basename(target) or target
                return data, basename
        except Exception as e:
            self._log(f"❌ zip 해제 실패: {type(e).__name__}: {e}", "error")
            return None, None

    def _get_real_value(self, ws, row, col):
        """병합 셀 고려해서 실제 값 가져오기 (원본 Cell 9 get_real_value 이식)."""
        cell = ws.cell(row, col)
        if cell.value is not None:
            return cell.value
        for range_ in ws.merged_cells.ranges:
            if cell.coordinate in range_:
                return ws.cell(range_.min_row, range_.min_col).value
        return None

    def fill_quotation_excel(self, quotation_path, prod_name, raw_df,
                             prod_path=None,
                             brand="상세페이지 참조",
                             company="상세페이지 참조",
                             phone="010-0000-0000"):
        """[원본 Cell 9/17 이식] 견적서 xlsx 를 '완료' 탭 데이터로 자동 작성.

        Args:
            quotation_path: 견적서 xlsx 파일의 절대 경로 (str 또는 Path)
            prod_name:      변환상품명 (raw_df 내 매칭 키)
            raw_df:         '완료' 탭 전체 DataFrame (열: 변환상품명, 옵션1, 한글 옵션명, 등)
            prod_path:      상품 폴더 절대 경로. 상세이미지 컬럼 fallback 용으로
                            "{prod_path}/상세페이지/" 하위 파일을 스캔해 채움.
            brand/company/phone: config 에서 전달된 전역값 (비어 있으면 기본값)

        Returns:
            bool: 저장까지 성공하면 True, 아니면 False.

        사용자가 수동 입력해야 하는 빈 칸은 노란색(FFFF00)으로 칠해 준다.
        """
        import openpyxl as _xl
        import time as _time
        import os as _os
        from openpyxl.styles import PatternFill

        qpath = str(quotation_path)
        if not _os.path.exists(qpath):
            self._log(f"    ❌ 견적서 파일 없음: {qpath}", "error")
            return False

        # 해당 상품의 옵션행 필터링
        try:
            name_col = raw_df['변환상품명'].astype(str).str.strip()
            src_rows = raw_df[name_col == str(prod_name).strip()]
        except Exception as e:
            self._log(f"    ❌ 완료 탭 필터 실패: {e}", "error")
            return False
        if src_rows.empty:
            self._log(f"    ⚠️ '완료' 탭에서 매칭 데이터 없음: {prod_name}", "warn")
            return False
        src_data = src_rows.to_dict('records')

        # 상세이미지 컬럼명 결정 (원본 Cell 9 로직)
        detail_col_name = "상세페이지명"
        for col in ["상세이미지파일명", "상세이미지", "상세페이지"]:
            if col in raw_df.columns:
                detail_col_name = col
                break

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # FIXED_VALUES — MY_BRAND_NAME/COMPANY/PHONE 은 config 에서 주입받음
        FIXED_VALUES = {
            "FIXED_BARCODE": "바코드 없음(쿠팡 바코드 생성 요청)",
            "FIXED_BRAND": brand or "상세페이지 참조",
            "FIXED_MANUFACTURER": company or "상세페이지 참조",
            "FIXED_PHONE": phone or "010-0000-0000",
            "FIXED_SIZE_CHART": "사이즈표.jpg",
            "FIXED_TAX": "과세",
            "FIXED_DEAL_TYPE": "기타 도소매업자",
            "FIXED_IMPORT": "수입상품",
            "FIXED_SKU_COUNT": 50,
            "FIXED_CAUTION": "해당사항없음",
            "FIXED_HANGER": "N",
            "FIXED_DIMENSION": "상세페이지참조",
            "FIXED_YEAR": 2025,
            "FIXED_SEASON": "사계절",
            "FIXED_MARK": "barcode.png",
            "FIXED_ORIGIN": "Made in China",
            "FIXED_WASH": "상세페이지 참조",
            "FIXED_WARRANTY": "제품 이상 시 공정거래위원회 고시 소비자분쟁해결 기준에 의거 보상합니다",
            "FIXED_NA": "해당사항없음",
            "FIXED_SHELF_LIFE": 0,
        }

        MAPPING_CONFIG = {
            "권장": "권장가", "소비자": "권장가", "공급가": "공급가", "판매가": "쿠팡판매가",
            "과세": "FIXED_TAX", "거래": "FIXED_DEAL_TYPE", "수입": "FIXED_IMPORT",
            "상품명": "전체옵션명", "바코드": "FIXED_BARCODE", "브랜드": "FIXED_BRAND",
            "제조사": "FIXED_MANUFACTURER", "수입자": "FIXED_MANUFACTURER",
            "검색태그": "태그", "색상": "CLEAN_COLOR_NAME", "사이즈": "사이즈", "대표이미지": "전체옵션명",
            "상세이미지": detail_col_name,
            "수량": "CALC_SET_COUNT",
            "차트": "FIXED_SIZE_CHART", "대체": "한글 옵션명",
            "SKU": "FIXED_SKU_COUNT", "행어": "FIXED_HANGER", "포장무게": "CALC_WEIGHT",
            "단품포장": "포장사이즈", "소재": "재질", "재질": "재질", "치수": "FIXED_DIMENSION",
            "출시": "FIXED_YEAR", "계절": "FIXED_SEASON",
            "취급": "FIXED_WASH", "주의": "FIXED_WASH",
            "보증": "FIXED_WARRANTY", "AS": "FIXED_PHONE",
            "사유": "FIXED_NA", "취급주의": "FIXED_NA",
            "마크": "FIXED_NA", "타입": "FIXED_NA", "인증번호": "FIXED_NA", "EMC": "FIXED_NA",
            "안전기준": "FIXED_NA", "신고번호": "FIXED_NA", "적합확인": "FIXED_NA",
            "유통기한": "FIXED_SHELF_LIFE", "소비기한": "FIXED_SHELF_LIFE",
            "라벨": "FIXED_MARK", "도안": "FIXED_MARK", "표시사항": "FIXED_MARK",
            "제조국": "FIXED_ORIGIN", "원산지": "FIXED_ORIGIN",
            "인증/허가": "FIXED_NA", "허가사항": "FIXED_NA",
        }

        # 상세이미지 파일명은 "상세_{전체옵션명}.jpg" 규칙으로 루프 내에서 생성한다.
        # (구버전의 상세페이지/상세copy이미지/ 폴더 스캔 fallback 은 제거됨 — 기존 성공 견적서에서
        # 확인된 규칙은 폴더 파일 리스트가 아니라 "상품명 기반 고정 패턴".)

        # 전처리: 무게 → CALC_WEIGHT, 색상 → CLEAN_COLOR_NAME
        for data in src_data:
            try:
                data['CALC_WEIGHT'] = int(''.join(filter(str.isdigit, str(data.get('무게', 0)))) or 0) + 10
            except Exception:
                data['CALC_WEIGHT'] = 10
            opt1 = str(data.get('옵션1', '')).strip()
            full_opt = str(data.get('한글 옵션명', '')).strip()
            data['CLEAN_COLOR_NAME'] = opt1 if opt1 else full_opt

        # 엑셀 열기
        try:
            _time.sleep(0.3)
            wb_tgt = _xl.load_workbook(qpath)
        except Exception as e:
            self._log(f"    ❌ 엑셀 열기 실패: {type(e).__name__}: {e}", "error")
            return False

        # 헤더 행 자동 탐색 ('상품명' 포함 행)
        ws_tgt = None
        header_row_idx = 5
        for sheet in wb_tgt.worksheets:
            for r in range(1, 15):
                row_vals = [str(c.value) for c in sheet[r] if c.value]
                if any("상품명" in v for v in row_vals):
                    ws_tgt = sheet
                    header_row_idx = r
                    break
            if ws_tgt:
                break
        if not ws_tgt:
            ws_tgt = wb_tgt.active
        self._log(f"    📋 헤더 행 발견: {header_row_idx}행 ({ws_tgt.title})")

        # SKU 열 자동 탐색 ('박스' + 'SKU' + '수량')
        sku_col_idx = None
        for r in range(4, 7):
            for c in range(1, 60):
                val = str(self._get_real_value(ws_tgt, r, c) or "").replace(" ", "").replace("\n", "")
                if "박스" in val and "SKU" in val and "수량" in val:
                    sku_col_idx = c
                    break
            if sku_col_idx:
                break
        if sku_col_idx:
            self._log(f"    📋 SKU 열 발견: {sku_col_idx}열")
        else:
            self._log(f"    ⚠️ SKU 열 못 찾음 (우측 유통기간=0 입력 생략)", "warn")

        # 컬럼 매핑 (비노출 속성은 SKIP)
        col_map = {}
        mapped_cnt = 0
        skip_cnt = 0
        for col, cell in enumerate(ws_tgt[header_row_idx], 1):
            if not cell.value:
                continue
            check_row = header_row_idx - 1
            row_4_val = str(self._get_real_value(ws_tgt, check_row, col) or "").strip().replace(" ", "")
            if ("비노출" in row_4_val) or ("속성" in row_4_val and "비노출" in row_4_val):
                col_map[col] = {'key': None, 'clean_header': 'SKIP'}
                skip_cnt += 1
                continue
            clean_h = str(cell.value).replace(" ", "").replace("\n", "")
            map_key = None
            # 긴 키부터 매칭 (부분 일치 우선순위)
            for k in sorted(MAPPING_CONFIG.keys(), key=len, reverse=True):
                if k in clean_h:
                    map_key = MAPPING_CONFIG[k]
                    break
            col_map[col] = {'key': map_key, 'clean_header': clean_h}
            if map_key:
                mapped_cnt += 1
        self._log(f"    📋 매핑된 컬럼: {mapped_cnt}개 / 비노출 SKIP: {skip_cnt}개")

        # ── 상세이미지 파일명 규칙 ──
        # 첫 번째 "상세이미지" 헤더 컬럼에만 "상세_{전체옵션명}.jpg" 를 입력한다.
        # 나머지 상세이미지 컬럼(예: 상세이미지 파일명2, 3...)은 기존 동작대로 빈칸(노란색).
        primary_detail_col = None
        for cidx in sorted(col_map.keys()):
            info_ = col_map[cidx]
            if info_.get('clean_header') == 'SKIP':
                continue
            if "상세이미지" in (info_.get('clean_header') or ""):
                primary_detail_col = cidx
                break
        if primary_detail_col:
            self._log(f"    📝 상세이미지 파일명 자동 생성: 상품명 기반 (첫 번째 상세이미지 컬럼 = {primary_detail_col}열)")
        detail_preview_count = 0  # 처음 3개 미리보기 로그 카운터

        # 데이터 입력 (헤더로부터 +4행 아래가 데이터 시작 — 원본 규약)
        start_row = header_row_idx + 4
        self._log(f"    📋 데이터 입력 중: 옵션 {len(src_data)}개, 시작행 {start_row}")
        yellow_count = 0
        for i, data in enumerate(src_data):
            r_idx = start_row + i

            # SKU 열 오른쪽 = 유통기간 0
            if sku_col_idx:
                shelf_life_col = sku_col_idx + 1
                ws_tgt.cell(row=r_idx, column=shelf_life_col, value=0)

            for col in range(1, ws_tgt.max_column + 1):
                if sku_col_idx and col == (sku_col_idx + 1):
                    continue  # 유통기간은 위에서 이미 0 넣음

                info = col_map.get(col, {'clean_header': '', 'key': None})
                if info['clean_header'] == 'SKIP':
                    continue

                clean_h = info['clean_header']
                val = None

                if "전화번호" in clean_h or "소비자상담" in clean_h:
                    val = FIXED_VALUES["FIXED_PHONE"]
                elif "추가이미지" in clean_h:
                    val = "추가1.jpg, 추가2.jpg"
                elif "상세이미지" in clean_h:
                    # 규칙: 첫 번째 상세이미지 컬럼에만 "상세_{전체옵션명}.jpg" 입력.
                    # 나머지 상세이미지 컬럼은 빈 값으로 두어 기존 노란색 표시가 적용되게 한다.
                    if col == primary_detail_col:
                        pname = str(data.get('전체옵션명', '')).strip() \
                                or str(data.get('변환상품명', '')).strip()
                        if pname:
                            val = f"상세_{pname}.jpg"
                            if detail_preview_count < 3:
                                self._log(f"       [옵션{i + 1}] {val}")
                                detail_preview_count += 1
                    # else: val 은 None 유지 → 빈칸 + 노란색
                elif "품명" in clean_h or "모델명" in clean_h:
                    val = data.get('전체옵션명')
                elif info['key']:
                    target_key = info['key']
                    if target_key == "CALC_SET_COUNT":
                        set_val = data.get('세트', '1')
                        val = f"{set_val}개"
                    elif target_key in FIXED_VALUES:
                        val = FIXED_VALUES[target_key]
                    elif target_key in data:
                        val = data[target_key]
                        hdr = str(ws_tgt.cell(header_row_idx, col).value or "")
                        if val and "이미지" in hdr and not str(val).lower().endswith('.jpg'):
                            val = f"{val}.jpg"

                if val is not None:
                    ws_tgt.cell(row=r_idx, column=col, value=val)

                # 빈 칸 노란색 (단, '선택' 키워드가 아래 안내줄에 있으면 제외)
                curr = ws_tgt.cell(row=r_idx, column=col).value
                if (curr is None or str(curr).strip() == "") and curr != 0:
                    hint = str(ws_tgt.cell(header_row_idx + 1, col).value or "")
                    if "선택" not in hint:
                        ws_tgt.cell(row=r_idx, column=col).fill = yellow_fill
                        yellow_count += 1

        self._log(f"    📋 빈칸 노란색 표시: {yellow_count}개")

        # 저장 (PermissionError 3회 재시도 — 엑셀 열려 있으면 사용자 안내)
        saved = False
        last_err = None
        for attempt in range(3):
            try:
                wb_tgt.save(qpath)
                saved = True
                break
            except PermissionError as pe:
                last_err = pe
                self._log(
                    f"    ⚠️ 엑셀이 열려 있습니다. 닫아주세요... ({attempt + 1}/3)",
                    "warn"
                )
                _time.sleep(3)
            except Exception as e:
                last_err = e
                _time.sleep(1)

        if not saved:
            self._log(f"    ❌ 견적서 저장 실패: {type(last_err).__name__}: {last_err}", "error")
            return False

        self._log(f"    💾 견적서 저장 완료: {qpath}", "success")
        return True

    # ─────────────────────────────────────────────────────────
    # 세션 유지용 Chrome 수명주기 헬퍼 (graceful shutdown + singleton 정리)
    # ─────────────────────────────────────────────────────────
    def _cleanup_singleton(self, path_prof):
        """크롬 프로필 락 잔재 파일(Singleton*, Lock*)을 제거한다.

        드라이버가 살아 있을 때 호출하면 실행 중 크롬과 충돌 → 반드시 drop 후 호출.
        """
        import os as _os
        import glob as _glob
        if self.search_driver is not None:
            return  # 안전장치: 드라이버가 있으면 건드리지 않음
        removed = 0
        try:
            for pat in ("Singleton*", "lockfile", "Lock*", "parent.lock"):
                for p in _glob.glob(_os.path.join(path_prof, pat)):
                    try:
                        _os.remove(p)
                        removed += 1
                    except Exception:
                        pass
        except Exception:
            pass
        if removed:
            self._log(f"🧹 프로필 락 정리: {removed}개 파일 삭제")

    def _kill_orphan_chrome(self, path_prof):
        """해당 프로필(path_prof) 또는 9222 포트를 점유 중인 chrome.exe 만 선택적 종료.

        stage2 의 safe-kill 패턴을 이식. 사용자의 일반 Chrome 창은 절대 건드리지 않는다.
        graceful shutdown 이 실패했거나 재기동 시 포트 충돌 시에만 최후 수단으로 호출.
        """
        import subprocess as _sub
        import os as _os
        try:
            profile_norm = _os.path.normpath(path_prof).replace("\\", "\\\\")
            ps_cmd = (
                'powershell -NoProfile -Command "'
                '$ErrorActionPreference=\'SilentlyContinue\';'
                '$p=Get-CimInstance Win32_Process -Filter \\"Name=\'chrome.exe\'\\" |'
                ' Where-Object { $_.CommandLine -like \'*--user-data-dir=*'
                + profile_norm + '*\' -or '
                '$_.CommandLine -like \'*--remote-debugging-port=9222*\' };'
                'if ($p) { $p | ForEach-Object { Stop-Process -Id $_.ProcessId -Force } }"'
            )
            _sub.run(
                ps_cmd, shell=True, capture_output=True,
                text=True, encoding='utf-8', errors='replace', timeout=15,
            )
        except Exception as e:
            self._log(f"⚠️ safe-kill 실패: {e}", "warn")

    def _shutdown_driver(self):
        """Chrome 을 graceful 하게 종료해서 쿠키/세션 flush 를 보장한다.

        순서:
          1) DevTools Protocol Browser.close  → 쿠키 저장 트리거
          2) driver.quit()                    → Selenium 정상 종료 시도
          3) 프로필/포트 기준 safe-kill        → 최후 수단 (쿠키 flush 안 됨)
          4) Singleton 락 정리
        """
        import time as _time
        if self.search_driver is None:
            return

        driver = self.search_driver
        # 다른 곳에서 재사용되지 않도록 참조 먼저 끊음
        self.search_driver = None
        self.search_wait = None

        ok_cdp = False
        try:
            driver.execute_cdp_cmd("Browser.close", {})
            _time.sleep(1)
            ok_cdp = True
        except Exception:
            pass

        try:
            driver.quit()
        except Exception as e:
            self._log(f"⚠️ driver.quit() 실패: {e}", "warn")

        # CDP 가 성공했으면 프로세스가 이미 종료 중. 짧게 대기 후 잔존분 확인.
        _time.sleep(1)

        # 최후 수단 safe-kill — path_prof 이 알려진 경우에만
        path_prof = getattr(self, "path_prof", None) or str(PROJECT_ROOT / "chrome_profile")
        if not ok_cdp:
            self._log("⚠️ graceful shutdown 실패 → safe-kill 보강 (쿠키 일부 손실 가능)", "warn")
            self._kill_orphan_chrome(path_prof)
        else:
            # graceful 성공했어도 zombie 프로세스 방지용으로 한 번 더 (safe-kill 은 매칭된 것만)
            self._kill_orphan_chrome(path_prof)

        # 프로필 락 정리
        self._cleanup_singleton(path_prof)
        self._log("🔒 쿠팡 세션 flush 완료 — 다음 실행 시 로그인 유지 예상")

    def _ensure_driver(self, config):
        """크롬(remote-debugging 9222) 준비. 이미 살아 있으면 재사용, 아니면 새로 띄운다.

        매번 taskkill 하지 않음 — graceful shutdown 이 끝난 상태 전제. 기동 실패 시만
        safe-kill 로 zombie 정리 후 1회 재시도.
        """
        # 기존 세션이 살아있는지 가볍게 확인
        if self.search_driver is not None:
            try:
                _ = self.search_driver.current_url
                return self.search_driver, self.search_wait
            except Exception:
                self.search_driver = None
                self.search_wait = None

        import os as _os, time as _time, subprocess as _sub
        from selenium import webdriver
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.chrome.service import Service
        from webdriver_manager.chrome import ChromeDriverManager

        project_root = config.get("PROJECT_ROOT", _os.getcwd())
        path_prof = _os.path.join(project_root, "chrome_profile")
        self.path_prof = path_prof  # shutdown 에서 참조
        chrome_paths = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            _os.path.join(_os.getenv('LOCALAPPDATA', ''), r'Google\Chrome\Application\chrome.exe'),
        ]
        chrome_exe = next((p for p in chrome_paths if _os.path.exists(p)), None)
        if not chrome_exe:
            raise RuntimeError("크롬을 찾을 수 없습니다.")

        # 이전 실행의 잔재 락 정리 (임시 프로필 폴백 방지)
        self._cleanup_singleton(path_prof)

        def _connect_fresh():
            _sub.Popen([chrome_exe, "--remote-debugging-port=9222",
                        f"--user-data-dir={path_prof}", "--lang=ko"])
            _time.sleep(3)
            options = webdriver.ChromeOptions()
            options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
            drv = webdriver.Chrome(
                service=Service(ChromeDriverManager().install()), options=options
            )
            w = WebDriverWait(drv, 20)
            return drv, w

        try:
            driver, wait = _connect_fresh()
        except Exception as first_err:
            # 기동 실패 — 포트 점유된 zombie chrome 가능성. safe-kill + 재정리 후 1회 재시도.
            self._log(f"⚠️ 크롬 기동 실패, safe-kill 후 재시도: {first_err}", "warn")
            self._kill_orphan_chrome(path_prof)
            _time.sleep(1)
            self._cleanup_singleton(path_prof)
            driver, wait = _connect_fresh()

        self.search_driver = driver
        self.search_wait = wait
        return driver, wait

    def search_candidates(self, config, keyword: str) -> dict:
        """쿠팡허브에서 키워드로 카테고리 후보를 스크래핑해 반환.

        반환: {"candidates": [{"path", "category_id", "selectable"}], "needs_login": bool}
        """
        from selenium.webdriver.common.by import By
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.support import expected_conditions as EC
        import time as _time

        with self.driver_lock:
            driver, wait = self._ensure_driver(config)

            if "supplier.coupang.com" not in driver.current_url:
                driver.get("https://supplier.coupang.com/qvt/registration")
                _time.sleep(3)
            current_url = (driver.current_url or "").lower()
            if ("login" in current_url) or ("xauth.coupang.com" in current_url) or ("/auth/realms/" in current_url):
                # 수동 로그인 안내 전에 stage2의 _supplier_auto_login 을 재사용해 자동 로그인 시도
                try:
                    from modules.stage2 import _supplier_auto_login
                    logged_in = _supplier_auto_login(
                        driver,
                        self._log,
                        lambda: False,
                    )
                except Exception as _e:
                    self._log(f"⚠️ 자동 로그인 호출 실패: {_e}", "warn")
                    logged_in = False
                if not logged_in:
                    return {"candidates": [], "needs_login": True}
                # 로그인 성공 → 등록 페이지로 복귀
                driver.get("https://supplier.coupang.com/qvt/registration")
                _time.sleep(3)
            if "registration" not in driver.current_url:
                driver.get("https://supplier.coupang.com/qvt/registration")
                _time.sleep(2)

            # 모달 열기
            try:
                dl_btn = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, "//button[contains(., 'Download latest excel file')]")
                ))
                driver.execute_script("arguments[0].click();", dl_btn)
                _time.sleep(1.5)
            except Exception as e:
                raise RuntimeError(f"다운로드 버튼 클릭 실패: {e}")

            # 키워드 입력
            try:
                search_input = wait.until(EC.visibility_of_element_located(
                    (By.CSS_SELECTOR, "div.modal-body input[type='text']")
                ))
                search_input.click()
                search_input.send_keys(Keys.CONTROL + "a")
                search_input.send_keys(Keys.BACKSPACE)
                search_input.send_keys(keyword)
                search_input.send_keys(Keys.ENTER)
                _time.sleep(4)
            except Exception as e:
                raise RuntimeError(f"키워드 검색 실패: {e}")

            # 후보 수집
            # TODO(쿠팡허브 DOM 확인): 아래 XPath와 category_id 속성명은 기존 등록 코드의
            # `li[1]/a` 선택 경로에서 역추적한 "합리적 추측"임. 실제 첫 등록 시도 전에 다음을
            # 개발자도구로 확인해서 보정할 것:
            #   1) 후보 리스트를 감싸는 ul 의 실제 경로
            #      (현재: //*[@id="controlled-tab-example-pane-name"]/div[1]/div[1]/div/div/div/ul/li)
            #   2) 각 <a> 에 카테고리 고유 id 가 담긴 data-* 속성 이름
            #      (현재 탐색 순서: data-id > data-category-id > data-cat-id > href > text)
            #   3) 후보의 전체 경로 텍스트가 link.text 로 충분한지 (일부 UI는 별도 span)
            candidates = []
            try:
                items = driver.find_elements(
                    By.XPATH,
                    '//*[@id="controlled-tab-example-pane-name"]/div[1]/div[1]/div/div/div/ul/li'
                )
                for item in items:
                    try:
                        link = item.find_element(By.TAG_NAME, "a")
                        text = (link.text or "").strip()
                        cid = (link.get_attribute("data-id")
                               or link.get_attribute("data-category-id")
                               or link.get_attribute("data-cat-id")
                               or link.get_attribute("href")
                               or text)
                        if text:
                            candidates.append({
                                "path": text,
                                "category_id": cid or "",
                                "selectable": True,
                            })
                    except Exception:
                        continue
            except Exception:
                pass

            # 모달 닫기 (상태 복원)
            try:
                driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                _time.sleep(0.5)
            except Exception:
                pass

            return {"candidates": candidates, "needs_login": False}

    def _upload_single_product(self, driver, wait, item):
        """Cell 9-1(modules/stage2_cells/08_cell_9_1.py) 의 단일 상품 업로드 로직을
        CategoryRunner 맥락(로그 콜백/세션 공유)으로 이식한 버전.

        - item: {"name": str, "path": str(product_root_folder),
                 "excel": str, "dest_path": Optional[str]}
        - returns: True 면 Validate 제출 & 결과 팝업 닫기까지 성공.
        """
        import os as _os
        import time as _time
        import random as _rand
        import requests as _req
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC

        def _wait_net():
            while True:
                try:
                    _req.get("https://www.google.com", timeout=3)
                    return
                except Exception:
                    self._log("🚨 [네트워크 오류] 인터넷 연결 대기 중...", "warn")
                    _time.sleep(10)

        def _hsleep(a=1.5, b=4.0):
            _time.sleep(_rand.uniform(a, b))

        def _find_file(folder, keywords, exts):
            if not _os.path.exists(folder):
                return None
            for f in _os.listdir(folder):
                if not any(f.lower().endswith(e) for e in exts):
                    continue
                if any(k in f for k in keywords):
                    return _os.path.join(folder, f)
            return None

        def _reset_page():
            _wait_net()
            try:
                target = "https://supplier.coupang.com/qvt/registration"
                if target not in driver.current_url:
                    driver.get(target)
                else:
                    driver.refresh()
                self._log("🔄 페이지 초기화 (5초 대기)...")
                _time.sleep(5)
            except Exception:
                pass

        prod_name = item["name"]
        product_root_folder = item["path"]
        excel_path = item.get("dest_path") or item.get("excel")

        self._log(f"🔹 '{prod_name}' 등록 시작")

        if not excel_path or not _os.path.exists(excel_path):
            self._log("⚠️ 견적서 파일 없음 (Pass)", "warn")
            return False

        # ── Step 1: 파일 경로 확보 ──
        size_chart_path = _find_file(
            product_root_folder, ["사이즈", "사이즈표", "size"], [".jpg", ".png"]
        )
        barcode_path = _find_file(
            product_root_folder, ["barcode", "바코드"], [".png", ".jpg"]
        )

        detail_img_paths = []
        for d_path in [
            _os.path.join(product_root_folder, "상세페이지", "상세copy이미지"),
            _os.path.join(product_root_folder, "상세페이지", "COPY"),
            _os.path.join(product_root_folder, "상세페이지", "copy"),
            _os.path.join(product_root_folder, "상세페이지"),
        ]:
            if _os.path.exists(d_path):
                jpgs = sorted(
                    f for f in _os.listdir(d_path) if f.lower().endswith(".jpg")
                )
                if jpgs:
                    detail_img_paths = [_os.path.join(d_path, f) for f in jpgs]
                    self._log(
                        f"🔍 상세페이지 폴더: .../{_os.path.basename(d_path)} "
                        f"(총 {len(detail_img_paths)}장)"
                    )
                    break
        if not detail_img_paths:
            self._log("⚠️ 상세페이지 이미지를 찾을 수 없습니다", "warn")

        additional_imgs = []
        add_folder = _os.path.join(product_root_folder, "상세페이지", "추가이미지")
        if _os.path.exists(add_folder):
            for fname in ["추가1.jpg", "추가2.jpg"]:
                p = _os.path.join(add_folder, fname)
                if _os.path.exists(p):
                    additional_imgs.append(p)
                    self._log(f"📸 추가이미지 발견: {fname}")

        # ─── [수정 20260425 파일명매칭] 시작 ───────────────────────────
        # 원본 Cell 9-1(stage2_cells/08_cell_9_1.py:161-181)은 시트의
        # '대표이미지경로'를 base_folder 로 사용해 "상품루트/대표이미지/" 서브폴더로
        # 진입한다. 카테고리 추가 등록은 Cell 10 이관 후 시점이라 시트 경로 신뢰
        # 불가이므로 product_root_folder 로 폴백했었는데, 그러면 루트엔 .jpg 가
        # 없어 representative_images 가 빈 리스트가 되어 견적서에 적힌 대표이미지
        # 파일명과 매칭이 안 된다(쿠팡허브 빨간 에러).
        # 해결: 원본과 동일하게 "상품루트/대표이미지/" 를 우선 진입한다.
        default_rep = _os.path.join(product_root_folder, "대표이미지")
        target_image_folder = default_rep if _os.path.isdir(default_rep) \
            else product_root_folder
        for cand in [
            _os.path.join(product_root_folder, "COPY"),
            _os.path.join(product_root_folder, "copy"),
            _os.path.join(default_rep, "COPY"),
            _os.path.join(default_rep, "copy"),
        ]:
            if _os.path.exists(cand) and _os.path.isdir(cand):
                if any(f.lower().endswith(".jpg") for f in _os.listdir(cand)):
                    target_image_folder = cand
                    break
        # ─── [수정 20260425 파일명매칭] 끝 ─────────────────────────────

        representative_images = []
        if _os.path.exists(target_image_folder):
            representative_images = [
                _os.path.join(target_image_folder, f)
                for f in _os.listdir(target_image_folder)
                if f.lower().endswith(".jpg") and "processed_final" not in f
            ]
        self._log(
            f"🖼️ 대표이미지 폴더: {target_image_folder} "
            f"({len(representative_images)}장)"
        )

        files_to_upload = []
        files_to_upload.extend(representative_images)
        files_to_upload.extend(detail_img_paths)
        files_to_upload.extend(additional_imgs)
        if size_chart_path:
            files_to_upload.append(size_chart_path)
            self._log("📏 사이즈표 포함됨")

        self._log(f"✅ 파일 준비 완료 ({len(files_to_upload)}장)")

        # ── Step 2: 업로드 ──
        try:
            _wait_net()
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            _hsleep(1, 2)

            # [1] 견적서
            excel_ok = False
            for _ in range(3):
                try:
                    file_inputs = wait.until(
                        EC.presence_of_all_elements_located(
                            (By.CSS_SELECTOR, "input[type='file']")
                        )
                    )
                    if file_inputs:
                        file_inputs[0].send_keys(excel_path)
                        self._log("📤 [1/3] 견적서 업로드 성공")
                        excel_ok = True
                        _hsleep(5, 7)
                        break
                except Exception:
                    _time.sleep(2)
            if not excel_ok:
                self._log("❌ 견적서 업로드 실패 → 건너뜀", "error")
                _reset_page()
                return False

            # [2] 이미지 통합
            if files_to_upload:
                try:
                    file_inputs = driver.find_elements(
                        By.CSS_SELECTOR, "input[type='file']"
                    )
                    self._log(
                        f"📤 [2/3] 이미지 일괄 업로드 ({len(files_to_upload)}장)..."
                    )
                    file_inputs[1].send_keys("\n".join(files_to_upload))
                    self._log("⏳ 이미지 서버 전송 및 썸네일 생성 대기 (20초)...")
                    _time.sleep(20)
                except Exception as e:
                    self._log(f"❌ 이미지 업로드 오류: {e}", "error")

            # [3] 바코드
            if barcode_path:
                try:
                    file_inputs = driver.find_elements(
                        By.CSS_SELECTOR, "input[type='file']"
                    )
                    self._log("📤 [3/3] 바코드 업로드...")
                    file_inputs[2].send_keys(barcode_path)
                    _hsleep(4, 6)
                except Exception:
                    pass

            # ── 하단 체크들 ──
            self._log("🖱️ 하단 필수 항목 체크 시작...")
            try:
                self._log("  👉 [1/4] 모든 'Yes(예)' 일괄 클릭...")
                for el in driver.find_elements(
                    By.XPATH,
                    "//label[normalize-space()='Yes' or normalize-space()='예']",
                ):
                    try:
                        driver.execute_script("arguments[0].click();", el)
                        _time.sleep(0.3)
                    except Exception:
                        pass
            except Exception:
                pass

            try:
                self._log("  👉 [2/4] 안전인증 팝업 Agree 처리...")
                for btn in driver.find_elements(
                    By.XPATH,
                    "//button[normalize-space()='Agree' or normalize-space()='동의']",
                ):
                    if btn.is_displayed():
                        driver.execute_script("arguments[0].click();", btn)
                        self._log("  ✅ 'Agree' 팝업 닫기")
                        _time.sleep(1)
            except Exception:
                pass
            try:
                for ca in driver.find_elements(
                    By.XPATH,
                    "//label[contains(., 'contents stated above') "
                    "or contains(., '위 내용에 동의합니다')]",
                ):
                    driver.execute_script("arguments[0].click();", ca)
                    _time.sleep(0.3)
            except Exception:
                pass

            try:
                self._log("  👉 [3/4] 법적 서류 N/A 체크...")
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                _time.sleep(0.8)
                driver.execute_script("window.scrollTo(0, 0);")
                _time.sleep(0.3)
                all_na = driver.find_elements(
                    By.XPATH,
                    "//label[contains(., 'N/A') or contains(., 'n/a')]",
                )
                self._log(f"  📋 N/A 라벨 {len(all_na)}개 발견")
                clicked = 0
                for na in all_na:
                    try:
                        driver.execute_script(
                            "arguments[0].scrollIntoView({block:'center'});", na
                        )
                        _time.sleep(0.2)
                        driver.execute_script("arguments[0].click();", na)
                        clicked += 1
                        _time.sleep(0.3)
                    except Exception:
                        pass
                self._log(f"  ✅ N/A {clicked}/{len(all_na)}개 클릭 완료")
                driver.execute_script("window.scrollTo(0, 0);")
                _time.sleep(0.5)
            except Exception as e:
                self._log(f"  ⚠️ N/A 처리 오류: {e}", "warn")

            try:
                self._log("  👉 [4/4] 로켓 설치 'No' 정밀 타격...")
                rocket_xp = (
                    "//*[contains(translate(text(), "
                    "'ABCDEFGHIJKLMNOPQRSTUVWXYZ', "
                    "'abcdefghijklmnopqrstuvwxyz'), 'rocket installation') "
                    "or contains(text(), '로켓 설치') "
                    "or contains(text(), '로켓설치')]"
                    "//following::label[normalize-space()='No' "
                    "or normalize-space()='아니오'][1]"
                )
                no_label = wait.until(
                    EC.presence_of_element_located((By.XPATH, rocket_xp))
                )
                driver.execute_script("arguments[0].click();", no_label)
                self._log("  ✅ 로켓 설치 'No' 복구 완료", "success")
                _time.sleep(0.5)
            except Exception:
                self._log("  ⚠️ 로켓 정밀 타격 실패, 폴백 시도", "warn")
                try:
                    labels = driver.find_elements(
                        By.XPATH,
                        "//label[normalize-space()='No' or normalize-space()='아니오']",
                    )
                    if labels:
                        driver.execute_script(
                            "arguments[0].click();", labels[-1]
                        )
                except Exception:
                    pass

            try:
                retail_xp = (
                    "//input[@type='checkbox' and following-sibling::*"
                    "[contains(., 'I hereby agree on Coupang')]] "
                    "| //label[contains(., 'I hereby agree on Coupang')]"
                )
                retail_el = wait.until(
                    EC.presence_of_element_located((By.XPATH, retail_xp))
                )
                driver.execute_script(
                    """
                    var el = arguments[0];
                    el.click();
                    var chk = (el.tagName === 'INPUT') ? el : el.querySelector('input[type=\"checkbox\"]');
                    if (chk) {
                        chk.checked = true;
                        chk.dispatchEvent(new Event('change', { bubbles: true }));
                        chk.dispatchEvent(new Event('input', { bubbles: true }));
                    }
                    """,
                    retail_el,
                )
                self._log("  👉 권장 소비자 가격 약관 동의 체크")
                _time.sleep(2)
            except Exception:
                pass

            # ── Step 3: Validate & 결과 팝업 ──
            try:
                validate_btn = wait.until(
                    EC.presence_of_element_located(
                        (By.XPATH, "//button[contains(., 'Validate')]")
                    )
                )
                driver.execute_script("arguments[0].click();", validate_btn)
                self._log("  👉 [Validate] 버튼 클릭 완료")

                self._log("⏳ [스마트 대기] 결과 팝업을 기다리는 중... (최대 180초)")
                try:
                    WebDriverWait(driver, 180).until(
                        EC.visibility_of_element_located(
                            (By.XPATH, "//*[contains(text(), 'Quotation has been submitted')]")
                        )
                    )
                    self._log("🚀 결과 팝업 감지 완료!", "success")
                    _time.sleep(1)
                except Exception:
                    self._log("⚠️ 180초 초과: 결과 팝업 미감지 (오류 가능성)", "warn")

                try:
                    close_btn = driver.find_element(
                        By.XPATH,
                        "//button[contains(., '닫기') or contains(., '확인') or contains(., 'Close')]",
                    )
                    driver.execute_script("arguments[0].click();", close_btn)
                    self._log("  👉 결과 팝업 '닫기' 클릭 완료", "success")
                    _time.sleep(5)
                    return True
                except Exception:
                    self._log("  ⚠️ 결과 팝업 '닫기' 버튼 미발견", "warn")
                    return False
            except Exception as e:
                self._log(f"  ❌ 등록 버튼 처리 오류: {e}", "error")
                return False

        except Exception as e:
            self._log(f"❌ 치명적 오류: {e}", "error")
            _wait_net()
            _reset_page()
            return False

    def _run(self, config, products, selection):
        try:
            import json, os, glob, shutil, time, subprocess
            import pandas as pd
            import gspread
            import openpyxl
            from google.oauth2.service_account import Credentials
            from selenium.webdriver.common.by import By
            from selenium.webdriver.common.keys import Keys
            from selenium.webdriver.support.ui import WebDriverWait
            from selenium.webdriver.support import expected_conditions as EC

            keyword = (selection or {}).get("keyword", "") or ""
            target_cat_id = (selection or {}).get("category_id", "") or ""
            target_path = (selection or {}).get("path", "") or ""
            save_as_fav = bool((selection or {}).get("save_as_favorite"))
            fav_label = (selection or {}).get("favorite_label", "") or target_path or keyword

            my_dl_dir = os.path.join(os.path.expanduser("~"), "Downloads")
            creds_file = config.get("GOOGLE_JSON_FILE", "credentials.json")

            # ── Step 1: 구글 시트 완료 탭에서 상품 데이터 로드 ──
            self._prog(1, 4, "구글 시트 연결")
            self._log("📋 구글 시트 '완료' 탭에서 상품 데이터를 가져옵니다...")

            scope = ['https://www.googleapis.com/auth/spreadsheets',
                     'https://www.googleapis.com/auth/drive']
            creds = Credentials.from_service_account_file(creds_file, scopes=scope)
            client = gspread.authorize(creds)
            doc = client.open_by_url(config["GOOGLE_SHEET_URL"])
            done_sheet = doc.worksheet("완료")
            all_values = done_sheet.get_all_values()

            if len(all_values) < 2:
                self._log("❌ '완료' 시트에 데이터가 없습니다.", "error")
                self.status = "error"
                return

            headers = all_values[0]
            done_df = pd.DataFrame(all_values[1:], columns=headers)
            self._log(f"✅ 완료 탭 데이터 {len(done_df)}개 로드 완료")

            # ── Step 2: 크롬 실행 & 견적서 다운로드 ──
            self._prog(2, 4, "견적서 다운로드")
            self._log(f"🔍 카테고리: '{target_path or keyword}' (keyword='{keyword}')")

            # 크롬 준비 — /api/category/search 에서 이미 띄워둔 세션이 있으면 재사용
            driver, wait = self._ensure_driver(config)
            self._log("✅ 브라우저 연결 성공!")

            # 즐겨찾기 저장 (등록 성공 여부와 무관하게, 사용자가 선택한 카테고리를 즉시 기억)
            if save_as_fav and (target_cat_id or target_path):
                try:
                    import uuid as _uuid
                    from datetime import datetime as _dt
                    favs = load_favorites()
                    # 동일 path 또는 동일 category_id가 이미 있으면 중복 저장하지 않음
                    dup = any(
                        (target_cat_id and f.get("category_id") == target_cat_id)
                        or (target_path and f.get("path") == target_path)
                        for f in favs
                    )
                    if not dup:
                        favs.append({
                            "id": _uuid.uuid4().hex,
                            "label": fav_label,
                            "keyword": keyword,
                            "path": target_path,
                            "category_id": target_cat_id,
                            "added_at": _dt.now().isoformat(timespec="seconds"),
                        })
                        save_favorites(favs)
                        self._log(f"⭐ 즐겨찾기에 저장: {fav_label}")
                except Exception as e:
                    self._log(f"⚠️ 즐겨찾기 저장 실패: {e}", "warn")

            # 쿠팡 서플라이허브 접속
            if "supplier.coupang.com" not in driver.current_url:
                driver.get("https://supplier.coupang.com/qvt/registration")
                time.sleep(3)

            # 로그인 확인
            if "login" in driver.current_url:
                self._log("⚠️ 쿠팡 로그인이 필요합니다. 브라우저에서 로그인 후 계속 버튼을 눌러주세요.", "warn")
                self._gate("login")

            # 각 상품별 견적서 다운로드
            downloaded = []
            not_found = []
            for prod in products:
                prod_name = prod["name"]
                prod_path = prod["path"]

                if not os.path.exists(prod_path):
                    self._log(
                        f"❌ 폴더 없음: {prod_path} — 완료된_상품 폴더 안에 해당 상품 폴더가 있는지 확인하세요",
                        "error"
                    )
                    not_found.append(prod_name)
                    continue

                # 이미 해당 카테고리 견적서 있으면 스킵 (path 의 마지막 토큰 우선, 없으면 키워드)
                # ⚠️ 여기 _tag 정규화는 저장 시 fname_tag 정규화와 반드시 동일해야 매치 성립.
                _raw_tag = (target_path.split(">")[-1].strip() if target_path else "") or keyword
                _tag = safe_tag(_raw_tag) or safe_tag(keyword) or "카테고리"
                glob_pattern = os.path.join(prod_path, f"*{_tag}*견적서*.xlsx")
                self._log(f"📁 검색 위치: {prod_path}")
                self._log(f"🔍 견적서 검색 패턴: *{_tag}*견적서*.xlsx")
                existing = glob.glob(glob_pattern)
                if existing:
                    self._log(f"📄 매치된 파일: {os.path.basename(existing[0])}")
                    self._log(f"⏭️ 이미 견적서 있음: {prod_name}")
                    downloaded.append({"name": prod_name, "path": prod_path, "excel": existing[0]})
                    continue

                self._log(f"📥 '{prod_name}' - '{_tag}' 견적서 다운로드 중...")

                # 등록 페이지로 이동
                if "registration" not in driver.current_url:
                    driver.get("https://supplier.coupang.com/qvt/registration")
                    time.sleep(2)

                # Download 버튼 클릭
                try:
                    dl_btn = wait.until(EC.element_to_be_clickable(
                        (By.XPATH, "//button[contains(., 'Download latest excel file')]")
                    ))
                    driver.execute_script("arguments[0].click();", dl_btn)
                    time.sleep(1.5)
                except:
                    self._log(f"❌ 다운로드 버튼 클릭 실패: {prod_name}", "error")
                    continue

                # 키워드 검색 후, 후보 중 사용자가 선택한 카테고리와 일치하는 항목만 클릭
                try:
                    search_input = wait.until(EC.visibility_of_element_located(
                        (By.CSS_SELECTOR, "div.modal-body input[type='text']")
                    ))
                    search_input.click()
                    search_input.send_keys(Keys.CONTROL + "a")
                    search_input.send_keys(Keys.BACKSPACE)
                    search_input.send_keys(keyword)
                    search_input.send_keys(Keys.ENTER)
                    time.sleep(4)

                    # TODO(쿠팡허브 DOM): search_candidates()와 동일한 XPath/속성명을 공유.
                    # 첫 실등록 시 개발자도구로 실제 구조 확인 후 양쪽 보정할 것.
                    items = driver.find_elements(
                        By.XPATH,
                        '//*[@id="controlled-tab-example-pane-name"]/div[1]/div[1]/div/div/div/ul/li'
                    )
                    matched_link = None
                    matched_text = ""
                    for it in items:
                        try:
                            link = it.find_element(By.TAG_NAME, "a")
                            text = (link.text or "").strip()
                            cid = (link.get_attribute("data-id")
                                   or link.get_attribute("data-category-id")
                                   or link.get_attribute("data-cat-id")
                                   or link.get_attribute("href")
                                   or text)
                            if (target_cat_id and cid == target_cat_id) \
                               or (target_path and text == target_path):
                                matched_link = link
                                matched_text = text
                                break
                        except Exception:
                            continue

                    if not matched_link:
                        self._log(
                            f"⚠️ 일치하는 카테고리를 후보에서 찾지 못함: '{target_path}' (keyword='{keyword}')",
                            "warn"
                        )
                        driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                        continue

                    driver.execute_script("arguments[0].click();", matched_link)
                    self._log(f"✅ 카테고리 선택: {matched_text}")
                except Exception as e:
                    self._log(f"⚠️ 카테고리 검색/선택 실패: {e}", "warn")
                    try:
                        driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                    except Exception:
                        pass
                    continue

                # Search & Download
                try:
                    search_btn = wait.until(EC.element_to_be_clickable(
                        (By.XPATH, "//button[contains(., 'Search')]")
                    ))
                    driver.execute_script("arguments[0].click();", search_btn)
                    time.sleep(3)

                    dn_btn = wait.until(EC.element_to_be_clickable(
                        (By.XPATH, "//div[@class='modal-footer']//button[contains(., 'Download')]")
                    ))

                    # 다운로드 시작 시각 기록 (이 시각 이후 mtime 만 인정 → 잔재 배제)
                    download_started_at = time.time()
                    driver.execute_script("arguments[0].click();", dn_btn)

                    # ── 다중 위치 검색 ── (.zip 우선: 쿠팡허브가 실제로 내려주는 형식)
                    found = self.find_downloaded_file(
                        patterns=["*.zip", "*.xlsx"],
                        timeout=60,
                        started_at=download_started_at,
                        config=config,
                    )
                    if found is None:
                        self._log(f"⚠️ 파일 감지 실패: {prod_name}", "warn")
                        continue

                    # ── 이동 경로 계산 ──
                    # ⚠️ fname_tag 정규화는 위 "이미 있는 견적서 스킵"의 _tag 와 반드시 동일해야 한다.
                    _raw_fname_tag = keyword
                    if target_path:
                        last = target_path.split(">")[-1].strip()
                        if last:
                            _raw_fname_tag = last
                    fname_tag = safe_tag(_raw_fname_tag) or safe_tag(keyword) or "카테고리"

                    is_zip = found.suffix.lower() == ".zip"

                    if is_zip:
                        # zip 에서 xlsx 추출 (bytes 로 바로 받음)
                        xlsx_bytes, xlsx_basename = self.extract_excel_from_zip(found)
                        if xlsx_bytes is None:
                            self._log(f"❌ zip 에서 xlsx 추출 실패: {found.name}", "error")
                            continue

                    if is_zip:
                        # 원본 zip 을 상품 폴더에 보관 (나중에 zip 안 다른 자료 필요할 수 있음)
                        zip_dest = Path(prod_path) / found.name
                        if zip_dest.exists():
                            try:
                                zip_dest.unlink()
                            except Exception:
                                pass
                        try:
                            shutil.move(str(found), str(zip_dest))
                            self._log(f"📦 원본 zip 보관: {zip_dest.name}")
                        except Exception as zerr:
                            self._log(f"⚠️ zip 보관 실패(계속 진행): {zerr}", "warn")

                        # 최종 xlsx 파일명
                        dest_name = f"{fname_tag}_견적서_{xlsx_basename}"
                        dest_path = Path(prod_path) / dest_name
                        self._log(f"💾 저장될 파일명: {dest_name}")

                        if dest_path.exists():
                            try:
                                dest_path.unlink()
                                self._log(f"🗑 기존 동명 파일 삭제 후 덮어쓰기: {dest_name}")
                            except Exception as rm_err:
                                self._log(f"⚠️ 기존 파일 삭제 실패(계속 진행): {rm_err}", "warn")

                        try:
                            dest_path.write_bytes(xlsx_bytes)
                            self._log(f"💾 견적서 추출 완료: {dest_path}", "success")
                        except Exception as werr:
                            self._log(
                                f"❌ xlsx 쓰기 실패: {type(werr).__name__}: {werr}",
                                "error"
                            )
                            self._log(f"   대상: {dest_path}", "error")
                            continue
                    else:
                        # 과거 호환: .xlsx 가 직접 떨어진 경우
                        dest_name = f"{fname_tag}_견적서_{found.name}"
                        dest_path = Path(prod_path) / dest_name
                        self._log(f"💾 저장될 파일명: {dest_name}")

                        if dest_path.exists():
                            try:
                                dest_path.unlink()
                                self._log(f"🗑 기존 동명 파일 삭제 후 덮어쓰기: {dest_name}")
                            except Exception as rm_err:
                                self._log(f"⚠️ 기존 파일 삭제 실패(계속 진행): {rm_err}", "warn")

                        try:
                            shutil.move(str(found), str(dest_path))
                            self._log(f"💾 견적서 이동 성공: {found.name} → {dest_path}", "success")
                        except Exception as move_err:
                            self._log(
                                f"❌ 견적서 이동 실패: {type(move_err).__name__}: {move_err}",
                                "error"
                            )
                            self._log(f"   원본: {found}", "error")
                            self._log(f"   대상: {dest_path}", "error")
                            continue

                    # 공통 검증
                    if not (dest_path.exists() and dest_path.stat().st_size > 0):
                        self._log(
                            f"❌ 이동 후 파일 검증 실패: {dest_path} (존재하지 않거나 크기 0)",
                            "error"
                        )
                        continue

                    downloaded.append({
                        "name": prod_name,
                        "path": prod_path,
                        "excel": str(dest_path),
                        "dest_path": str(dest_path),
                    })
                    self._log(f"✅ 견적서 저장 확인: {dest_path}", "success")

                    # ── 견적서 자동 작성 (원본 Cell 9/17 이식) ──
                    self._log(f"📝 견적서 자동 작성 시작: {prod_name}")
                    try:
                        filled = self.fill_quotation_excel(
                            quotation_path=str(dest_path),
                            prod_name=prod_name,
                            raw_df=done_df,
                            prod_path=prod_path,
                            brand=config.get("MY_BRAND_NAME") or "상세페이지 참조",
                            company=config.get("MY_COMPANY_NAME") or "상세페이지 참조",
                            phone=config.get("MY_PHONE_NUMBER") or "010-0000-0000",
                        )
                    except Exception as fill_err:
                        self._log(
                            f"⚠️ 견적서 자동 작성 예외: {type(fill_err).__name__}: {fill_err}",
                            "warn"
                        )
                        filled = False
                    if filled:
                        self._log("✅ 견적서 자동 작성 완료 (빈칸은 노란색 표시됨)", "success")
                    else:
                        self._log("⚠️ 견적서 자동 작성 일부 실패 - 수동 입력 필요", "warn")

                except Exception as e:
                    self._log(f"❌ 다운로드 오류: {e}", "error")

            # ── 안전장치: 다운로드된 상품이 0개면 빈 견적서 업로드 방지 ──
            if not downloaded:
                self._log(
                    "⚠️ 다운로드된 상품이 0개입니다. 작업을 중단합니다.",
                    "error"
                )
                if not_found:
                    self._log(
                        f"   ↳ 폴더 미발견: {', '.join(not_found)}",
                        "error"
                    )
                self._log(
                    "   확인사항: 완료된_상품/<변환상품명>/ 폴더가 존재하는지, "
                    "UI에서 선택한 상품명이 실제 폴더명과 일치하는지.",
                    "warn"
                )
                with self.lock:
                    self.status = "error"
                self._emit("done", ok=False, error="다운로드된 상품 0개 — 폴더를 확인하세요")
                return

            # ── Step 3: 견적서 작성 ──
            self._prog(3, 4, "견적서 작성")
            self._log(f"📝 {len(downloaded)}개 상품 견적서 작성 시작...")

            for item in downloaded:
                prod_name = item["name"]
                excel_path = item["excel"]
                prod_path = item["path"]

                # 완료 탭에서 해당 상품 데이터 가져오기
                matched = done_df[done_df['변환상품명'].str.strip().str.contains(
                    prod_name.replace("'", ""), na=False
                )]

                if matched.empty:
                    self._log(f"⚠️ '{prod_name}' 데이터를 완료 탭에서 찾지 못했어요", "warn")
                    continue

                self._log(f"✅ '{prod_name}' 데이터 {len(matched)}개 옵션 매칭!")
                # (실제 견적서 작성은 Cell 9 로직과 동일)
                # 여기서는 견적서 파일이 이미 있으므로 패스
                # → 실제 구현 시 Cell 9 로직 적용

            # 견적서 확인 게이트 — 상품별 파일 정보를 payload 로 함께 보낸다.
            gate_payload = {
                "files": [
                    {
                        "product_name": item["name"],
                        "folder_path": item["path"],
                        "excel_path": item.get("dest_path") or item["excel"],
                        "excel_filename": os.path.basename(item.get("dest_path") or item["excel"]),
                    }
                    for item in downloaded
                ],
                "count": len(downloaded),
            }
            self._log(
                "📋 견적서를 확인해주세요. 엑셀에서 카테고리/가격/옵션 등을 확인·수정한 뒤 저장(Ctrl+S)하세요.",
                "warn"
            )
            self._prog(3, 4, "사용자 검토 대기 중...")
            self._gate("quotation", payload=gate_payload)

            # ── Step 4: 쿠팡 자동 업로드 (Cell 9-1 이식) ──
            self._prog(4, 4, "쿠팡 자동 업로드")
            self._log(f"🚀 {len(downloaded)}개 상품 쿠팡 자동 업로드 시작...")

            success_count = 0
            fail_count = 0
            session_dead = False
            for item in downloaded:
                if self.stop_event.is_set():
                    self._log("🛑 사용자 중지 요청 감지 — 업로드 중단", "warn")
                    break
                # 상품마다 등록 페이지 초기화 (잔상태 오염 방지)
                try:
                    target_url = "https://supplier.coupang.com/qvt/registration"
                    if target_url not in driver.current_url:
                        driver.get(target_url)
                    else:
                        driver.refresh()
                    time.sleep(3)
                except Exception as nav_err:
                    err_text = str(nav_err).lower()
                    if ("web view not found" in err_text
                            or "no such window" in err_text
                            or "invalid session id" in err_text
                            or "disconnected" in err_text):
                        self._log(
                            f"❌ 쿠팡 세션이 끊겼습니다: {nav_err} — 자동 업로드 중단",
                            "error",
                        )
                        session_dead = True
                        break

                # 로그인으로 튕겼으면 기존 _supplier_auto_login 재사용
                try:
                    cur = (driver.current_url or "").lower()
                except Exception:
                    cur = ""
                if ("login" in cur) or ("xauth.coupang.com" in cur) \
                        or ("/auth/realms/" in cur):
                    self._log("🔐 업로드 중 로그인 페이지 감지 — 자동 로그인 재시도", "warn")
                    try:
                        from modules.stage2 import _supplier_auto_login as _autolog
                        relogged = _autolog(
                            driver, self._log, lambda: self.stop_event.is_set()
                        )
                    except Exception as _ae:
                        self._log(f"⚠️ 자동 로그인 헬퍼 예외: {_ae}", "warn")
                        relogged = False
                    if not relogged:
                        self._log("❌ 자동 로그인 실패 — 이 상품 건너뜀", "error")
                        fail_count += 1
                        continue
                    try:
                        driver.get("https://supplier.coupang.com/qvt/registration")
                        time.sleep(3)
                    except Exception:
                        pass

                try:
                    ok = self._upload_single_product(driver, wait, item)
                except Exception as up_err:
                    err_text = str(up_err).lower()
                    if ("web view not found" in err_text
                            or "no such window" in err_text
                            or "invalid session id" in err_text
                            or "disconnected" in err_text):
                        self._log(
                            f"❌ 쿠팡 세션이 업로드 중 끊겼습니다: {up_err}",
                            "error",
                        )
                        session_dead = True
                        break
                    self._log(f"⚠️ 업로드 예외: {up_err}", "warn")
                    ok = False

                if ok:
                    success_count += 1
                    self._log(
                        f"📦 '{item['name']}' 등록 완료 "
                        f"({success_count}/{len(downloaded)})",
                        "success",
                    )
                else:
                    fail_count += 1
                    self._log(f"⛔ '{item['name']}' 등록 실패", "error")

            self._log("=" * 50)
            if session_dead:
                self._log(
                    "⚠️ 쿠팡 세션 문제로 자동 업로드를 중단했습니다. "
                    "견적서는 폴더에 저장돼 있으니 쿠팡허브에서 수동 업로드해주세요.",
                    "warn",
                )
            else:
                self._log(
                    f"🎉 자동 업로드 종료! 성공 {success_count}개 / 실패 {fail_count}개",
                    "success",
                )
            self._log(
                "🌐 크롬 창은 유지됩니다. 필요 시 추가 작업 가능합니다.",
                "info",
            )

            with self.lock:
                self.status = "completed"
            self._normal_complete = True  # ⚠️ 반드시 done emit 이전에 세팅 (finally 타이밍 경쟁 방지)
            # 성공 0개이거나 세션이 죽었으면 upload_pending=True 로 프론트 수동 안내 alert 유도
            self._emit(
                "done",
                ok=True,
                count=len(downloaded),
                success_count=success_count,
                fail_count=fail_count,
                files=[
                    {
                        "product_name": item["name"],
                        "excel_path": item.get("dest_path") or item["excel"],
                    }
                    for item in downloaded
                ],
                upload_pending=(session_dead or success_count == 0),
                chrome_kept=True,
            )

        except Exception as e:
            import traceback
            with self.lock:
                self.status = "error"
            self._emit("log", level="error", msg=f"❌ 오류: {e}")
            self._emit("log", level="error", msg=traceback.format_exc())
            self._emit("done", ok=False, error=str(e))
        finally:
            # 정상 완료면 크롬 유지 (사용자 수동 업로드용),
            # 에러/중지 등 비정상 종료면 graceful shutdown 으로 쿠키 flush + 세션 정리.
            try:
                if self._normal_complete:
                    # 의도적 비종료 — 사용자가 크롬 창에서 직접 쿠팡허브 대량상품등록으로 이동.
                    # 다음 /api/category/search 호출 시 같은 세션을 재사용한다.
                    pass
                else:
                    self._shutdown_driver()
            except Exception as _sd_err:
                self._log(f"⚠️ 드라이버 종료 처리 실패: {_sd_err}", "warn")
            with self.lock:
                self.running = False



cat_runner = CategoryRunner()


app = FastAPI(title="쿠팡 로켓배송 자동화")
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


@app.get("/")
def root():
    return FileResponse(STATIC_DIR / "index.html")


@app.get("/api/health")
def health():
    return {"status": "ok", "stage": "stage3"}


# ------------------ 설정 ------------------

@app.get("/api/settings")
def get_settings():
    env = load_env()
    required = ["GOOGLE_SHEET_URL", "GEMINI_API_KEY", "SUPPLY_ID", "SUPPLY_PW"]
    is_configured = all(env.get(k) for k in required) and CREDENTIALS_FILE.exists()

    def get(k):
        return env.get(k, DEFAULTS.get(k, ""))

    return {
        "is_configured": is_configured,
        "has_credentials": CREDENTIALS_FILE.exists(),
        "google_sheet_url": get("GOOGLE_SHEET_URL"),
        "target_sheet_name": get("TARGET_SHEET_NAME"),
        "has_gemini_api_key": bool(env.get("GEMINI_API_KEY")),
        "supply_id": get("SUPPLY_ID"),
        "has_supply_pw": bool(env.get("SUPPLY_PW")),
        "my_brand_name": get("MY_BRAND_NAME"),
        "my_company_name": get("MY_COMPANY_NAME"),
        "my_phone_number": get("MY_PHONE_NUMBER"),
        "exchange_factor": get("EXCHANGE_FACTOR"),
        "add_logistics_cost": get("ADD_LOGISTICS_COST"),
        "sale_price_rate": get("SALE_PRICE_RATE"),
        "rec_price_rate": get("REC_PRICE_RATE"),
        "round_unit": get("ROUND_UNIT"),
    }


@app.post("/api/settings")
def save_settings(payload: SettingsPayload):
    env = load_env()
    new_env = {
        "GOOGLE_SHEET_URL": payload.google_sheet_url,
        "TARGET_SHEET_NAME": payload.target_sheet_name or "상품등록목록",
        "SUPPLY_ID": payload.supply_id,
        "MY_BRAND_NAME": payload.my_brand_name,
        "MY_COMPANY_NAME": payload.my_company_name,
        "MY_PHONE_NUMBER": payload.my_phone_number,
        "EXCHANGE_FACTOR": payload.exchange_factor or "350",
        "ADD_LOGISTICS_COST": payload.add_logistics_cost or "4000",
        "SALE_PRICE_RATE": payload.sale_price_rate or "1.67",
        "REC_PRICE_RATE": payload.rec_price_rate or "1.30",
        "ROUND_UNIT": payload.round_unit or "-2",
        "GEMINI_API_KEY": payload.gemini_api_key or env.get("GEMINI_API_KEY", ""),
        "SUPPLY_PW": payload.supply_pw or env.get("SUPPLY_PW", ""),
    }
    save_env(new_env)
    return {"ok": True}


@app.post("/api/credentials")
async def upload_credentials(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".json"):
        raise HTTPException(400, "JSON 파일만 업로드 가능합니다.")
    content = await file.read()
    try:
        parsed = json.loads(content)
    except json.JSONDecodeError:
        raise HTTPException(400, "유효한 JSON 파일이 아닙니다.")
    if not isinstance(parsed, dict) or "client_email" not in parsed:
        raise HTTPException(400, "구글 서비스 계정 파일이 아닙니다 (client_email 필드 없음).")
    CREDENTIALS_FILE.write_bytes(content)
    return {"ok": True, "filename": "credentials.json"}


# ------------------ 1단계 ------------------

@app.post("/api/stage1/start")
def stage1_start():
    env = load_env()
    missing = [k for k in ["GOOGLE_SHEET_URL", "GEMINI_API_KEY", "SUPPLY_ID", "SUPPLY_PW"]
               if not env.get(k)]
    if missing:
        raise HTTPException(400, f"설정이 완료되지 않았습니다. (누락: {', '.join(missing)})")
    if not CREDENTIALS_FILE.exists():
        raise HTTPException(400, "credentials.json 파일이 업로드되지 않았습니다.")

    config = dict(env)
    config["GOOGLE_JSON_FILE"] = str(CREDENTIALS_FILE.resolve())
    config["PROJECT_ROOT"] = str(PROJECT_ROOT.resolve())

    if not runner.start(config):
        raise HTTPException(409, "이미 실행 중입니다.")
    return {"ok": True}


@app.post("/api/stage1/stop")
def stage1_stop():
    if not runner.running:
        raise HTTPException(400, "실행 중이 아닙니다.")
    runner.stop()
    return {"ok": True}


@app.get("/api/stage1/status")
def stage1_status():
    return {
        "running": runner.running,
        "status": runner.status,
        "current_step": runner.current_step,
        "progress": runner.progress,
        "checklist_done": STAGE1_COMPLETED_FILE.exists(),
    }


@app.post("/api/stage1/checklist")
def stage1_checklist(payload: ChecklistPayload):
    if not (payload.item1 and payload.item2 and payload.item3):
        raise HTTPException(400, "3개 항목 모두 체크해야 합니다.")
    STAGE1_COMPLETED_FILE.write_text(str(time.time()), encoding="utf-8")
    return {"ok": True}


@app.post("/api/reset")
def reset_state():
    """UI '새로 시작' 버튼 — 1단계 체크리스트 완료 마커 제거.

    실행 중에는 거부. 로그 버퍼(runner.log_buffer, runner2.log_buffer)는
    건드리지 않아도 UI에서 DOM만 비우면 충분하지만 재접속 시 이전 로그가
    다시 흘러오지 않도록 함께 비운다.
    """
    if runner.running or runner2.running:
        raise HTTPException(409, "실행 중에는 초기화할 수 없습니다.")
    if STAGE1_COMPLETED_FILE.exists():
        try:
            STAGE1_COMPLETED_FILE.unlink()
        except Exception:
            pass
    try:
        with runner.lock:
            runner.log_buffer.clear()
            runner.status = "idle"
            runner.current_step = ""
            runner.progress = {"step": 0, "total": 7, "label": ""}
    except Exception:
        pass
    try:
        with runner2.lock:
            runner2.log_buffer.clear()
            runner2.status = "idle"
            runner2.current_step = ""
            runner2.progress = {"step": 0, "total": 9, "label": ""}
            runner2.counter = {"current": 0, "limit": 50}
            runner2.pending_gate = None
    except Exception:
        pass
    return {"ok": True}


@app.websocket("/ws/stage1")
async def stage1_ws(ws: WebSocket):
    await ws.accept()
    # 재연결 시 전체 buffer replay 방지: ?since=<ts> 이후 이벤트만 전송
    try:
        since = float(ws.query_params.get("since", "0") or "0")
    except Exception:
        since = 0.0
    while True:
        try:
            with runner.lock:
                buffer = list(runner.log_buffer)
                snapshot = {
                    "running": runner.running,
                    "status": runner.status,
                    "progress": runner.progress,
                }
            new_since = since
            for msg in buffer:
                ts = msg.get("ts", 0) or 0
                if ts > since:
                    await ws.send_json(msg)
                    if ts > new_since:
                        new_since = ts
            since = new_since
            # send periodic status snapshot so UI can sync button state
            await ws.send_json({"type": "snapshot", **snapshot})
            await asyncio.sleep(1.5)
        except WebSocketDisconnect:
            break
        except Exception as e:
            print(f"[ws/stage1] send error: {e}")
            break


# ------------------ 2단계 ------------------

@app.post("/api/stage2/start")
def stage2_start():
    # 2단계 락 해제: 1단계 체크리스트 없이도 시작 가능.
    env = load_env()
    missing = [k for k in ["GOOGLE_SHEET_URL", "GEMINI_API_KEY", "SUPPLY_ID", "SUPPLY_PW"]
               if not env.get(k)]
    if missing:
        raise HTTPException(400, f"설정이 완료되지 않았습니다. (누락: {', '.join(missing)})")
    if not CREDENTIALS_FILE.exists():
        raise HTTPException(400, "credentials.json 파일이 업로드되지 않았습니다.")
    if runner.running:
        raise HTTPException(409, "1단계가 실행 중입니다.")

    config = dict(env)
    config["GOOGLE_JSON_FILE"] = str(CREDENTIALS_FILE.resolve())
    config["PROJECT_ROOT"] = str(PROJECT_ROOT.resolve())

    if not runner2.start(config):
        raise HTTPException(409, "이미 실행 중입니다.")
    return {"ok": True}


@app.post("/api/stage2/stop")
def stage2_stop():
    if not runner2.running:
        raise HTTPException(400, "실행 중이 아닙니다.")
    runner2.stop()
    return {"ok": True}


@app.get("/api/stage2/status")
def stage2_status():
    return {
        "running": runner2.running,
        "status": runner2.status,
        "current_step": runner2.current_step,
        "progress": runner2.progress,
        "counter": runner2.counter,
        "pending_gate": runner2.pending_gate,
        "pending_gate_payload": runner2.pending_gate_payload,
        "stage1_checklist_done": STAGE1_COMPLETED_FILE.exists(),
    }


@app.post("/api/stage2/continue")
def stage2_continue(body: dict = Body(default=None)):
    """UI 게이트(예: 견적서 확인) 승인 후 호출."""
    if not runner2.running:
        raise HTTPException(400, "실행 중이 아닙니다.")
    name = (body or {}).get("gate") or runner2.pending_gate
    if not name:
        raise HTTPException(400, "대기 중인 게이트가 없습니다.")
    ok = runner2.release_gate(name)
    if not ok:
        raise HTTPException(400, f"게이트를 찾을 수 없습니다: {name}")
    return {"ok": True, "released": name}


@app.websocket("/ws/stage2")
async def stage2_ws(ws: WebSocket):
    await ws.accept()
    # 재연결 시 전체 buffer replay 방지: ?since=<ts> 이후 이벤트만 전송
    try:
        since = float(ws.query_params.get("since", "0") or "0")
    except Exception:
        since = 0.0
    while True:
        try:
            with runner2.lock:
                buffer = list(runner2.log_buffer)
                snapshot = {
                    "running": runner2.running,
                    "status": runner2.status,
                    "progress": runner2.progress,
                    "counter": runner2.counter,
                    "pending_gate": runner2.pending_gate,
                    "pending_gate_payload": runner2.pending_gate_payload,
                }
            new_since = since
            for msg in buffer:
                ts = msg.get("ts", 0) or 0
                if ts > since:
                    await ws.send_json(msg)
                    if ts > new_since:
                        new_since = ts
            since = new_since
            await ws.send_json({"type": "snapshot", **snapshot})
            await asyncio.sleep(1.5)
        except WebSocketDisconnect:
            break
        except Exception as e:
            print(f"[ws/stage2] send error: {e}")
            break




# ------------------ 카테고리 추가 등록 ------------------

@app.get("/api/category/products")
def get_completed_products():
    """완료 탭 상품 목록 반환.

    주의: 시트의 '대표이미지경로'는 Cell 10 이관 전 원작업폴더 경로라 이관 후에는 신뢰 불가.
    정답 규칙은 PROJECT_ROOT/완료된_상품/<변환상품명> — 이 경로로 재계산해 반환한다.
    옵션 행이 여러 개인 경우 변환상품명 기준으로 묶어 options 리스트에 모은다.
    """
    try:
        env = load_env()
        creds_file = str(CREDENTIALS_FILE.resolve())
        from google.oauth2.service_account import Credentials
        import gspread
        scope = ['https://www.googleapis.com/auth/spreadsheets',
                 'https://www.googleapis.com/auth/drive']
        creds = Credentials.from_service_account_file(creds_file, scopes=scope)
        client = gspread.authorize(creds)
        doc = client.open_by_url(env["GOOGLE_SHEET_URL"])
        done_sheet = doc.worksheet("완료")
        all_values = done_sheet.get_all_values()
        if len(all_values) < 2:
            return {"products": []}
        headers = all_values[0]
        name_idx = headers.index("변환상품명") if "변환상품명" in headers else -1
        opt1_idx = headers.index("옵션1") if "옵션1" in headers else -1
        opt2_idx = headers.index("옵션2") if "옵션2" in headers else -1

        # 시트 타임스탬프 컬럼 존재 여부 (현재 시트 구조엔 없지만 향후 대응 가능)
        ts_col_candidates = ["등록시각", "완료시각", "타임스탬프", "생성일"]
        ts_idx = -1
        for c in ts_col_candidates:
            if c in headers:
                ts_idx = headers.index(c)
                break

        completed_dir = PROJECT_ROOT / "완료된_상품"

        # 변환상품명 등장 순서 유지 + 옵션 누적 + 마지막 등장 행 인덱스 + 타임스탬프
        order: list = []
        per_product_options: dict = {}
        per_product_last_row: dict = {}   # 시트에서 해당 상품이 마지막에 나타난 행 번호
        per_product_ts: dict = {}         # 시트 타임스탬프 (있으면)
        for ri, row in enumerate(all_values[1:], start=1):
            name = row[name_idx].strip() if name_idx >= 0 and len(row) > name_idx else ""
            if not name:
                continue
            if name not in per_product_options:
                per_product_options[name] = []
                order.append(name)
            per_product_last_row[name] = ri  # 덮어쓰며 최종 인덱스 남김
            o1 = row[opt1_idx].strip() if opt1_idx >= 0 and len(row) > opt1_idx else ""
            o2 = row[opt2_idx].strip() if opt2_idx >= 0 and len(row) > opt2_idx else ""
            if o1 or o2:
                per_product_options[name].append({"opt1": o1, "opt2": o2})
            if ts_idx >= 0 and len(row) > ts_idx:
                raw_ts = row[ts_idx].strip()
                if raw_ts:
                    per_product_ts[name] = raw_ts

        # ── 정렬 기준 결정 ──
        # 1순위: 시트 타임스탬프 (있으면) 내림차순
        # 2순위: 완료된_상품/<변환상품명>/ 폴더 mtime 내림차순
        # 3순위: 시트 마지막 등장 행 인덱스 내림차순 (최신 입력 먼저)
        # 4순위: 변환상품명 역순 (최후 수단 안정 정렬용)
        sort_by = "sheet_row_desc"
        if per_product_ts:
            sort_by = "sheet_timestamp_desc"
        else:
            # 폴더가 하나라도 존재하면 mtime 기반으로 판정
            for name in order:
                if (completed_dir / name).exists():
                    sort_by = "folder_mtime_desc"
                    break

        def _sort_key(name: str):
            ts = per_product_ts.get(name, "")
            folder = completed_dir / name
            try:
                mtime = folder.stat().st_mtime if folder.exists() else 0.0
            except OSError:
                mtime = 0.0
            last_row = per_product_last_row.get(name, 0)
            # sorted(..., reverse=True) 에서 큰 값이 먼저.
            return (ts, mtime, last_row, name)

        sorted_order = sorted(order, key=_sort_key, reverse=True)

        products = []
        for name in sorted_order:
            folder = completed_dir / name
            try:
                mtime = folder.stat().st_mtime if folder.exists() else None
            except OSError:
                mtime = None
            products.append({
                "name": name,
                "path": str(folder),
                "path_exists": folder.exists(),
                "options": per_product_options[name],
                "mtime": mtime,
            })

        # 디버그 로그 (서버 콘솔)
        print(f"📊 완료 탭 상품 {len(products)}개 로드, 정렬: {sort_by}")
        for i, p in enumerate(products[:3], start=1):
            mt_str = ""
            if p["mtime"]:
                mt_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(p["mtime"]))
            print(f"🔝 [{i}] {p['name']}  ({mt_str or '폴더 없음'})")

        return {
            "products": products,
            "sort_by": sort_by,
            "count": len(products),
        }
    except Exception as e:
        raise HTTPException(500, str(e))


class CategoryStartPayload(BaseModel):
    products: list                      # [{"name": "...", "path": "..."}]
    keyword: str = ""                   # 쿠팡허브 검색창에 입력할 키워드 (path 매칭 실패 대비)
    category_id: str = ""               # 후보 중 매칭할 쿠팡 내부 카테고리 ID
    path: str = ""                      # "대분류 > 중분류 > ..." 전체 경로 (표시·매칭)
    save_as_favorite: bool = False
    favorite_label: str = ""


class CategorySearchPayload(BaseModel):
    keyword: str


class FavoriteAddPayload(BaseModel):
    label: str
    keyword: str = ""
    path: str
    category_id: str = ""


@app.post("/api/category/start")
def category_start(payload: CategoryStartPayload):
    if cat_runner.running:
        raise HTTPException(409, "이미 실행 중입니다.")
    if runner.running or runner2.running:
        raise HTTPException(409, "1단계 또는 2단계가 실행 중입니다.")
    if not payload.products:
        raise HTTPException(400, "등록할 상품이 비어 있습니다.")
    if not (payload.category_id or payload.path):
        raise HTTPException(400, "카테고리를 먼저 선택해주세요.")

    env = load_env()
    config = dict(env)
    config["GOOGLE_JSON_FILE"] = str(CREDENTIALS_FILE.resolve())
    config["PROJECT_ROOT"] = str(PROJECT_ROOT.resolve())

    selection = {
        "keyword": payload.keyword or (payload.path.split(">")[-1].strip() if payload.path else ""),
        "category_id": payload.category_id,
        "path": payload.path,
        "save_as_favorite": payload.save_as_favorite,
        "favorite_label": payload.favorite_label or payload.path,
    }
    if not cat_runner.start(config, payload.products, selection):
        raise HTTPException(409, "시작 실패")
    return {"ok": True}


@app.post("/api/category/search")
def category_search(payload: CategorySearchPayload):
    """키워드로 쿠팡허브 카테고리 후보를 스크래핑해 반환 (등록은 하지 않음)."""
    if cat_runner.running or runner.running or runner2.running:
        raise HTTPException(409, "다른 자동화가 실행 중입니다.")
    if not payload.keyword.strip():
        raise HTTPException(400, "키워드를 입력해주세요.")
    env = load_env()
    if not env.get("GOOGLE_SHEET_URL"):
        raise HTTPException(400, "설정이 완료되지 않았습니다.")
    config = dict(env)
    config["GOOGLE_JSON_FILE"] = str(CREDENTIALS_FILE.resolve())
    config["PROJECT_ROOT"] = str(PROJECT_ROOT.resolve())
    try:
        result = cat_runner.search_candidates(config, payload.keyword.strip())
    except Exception as e:
        raise HTTPException(500, f"검색 실패: {e}")
    return result


@app.get("/api/category/favorites")
def favorites_list():
    return {"favorites": load_favorites()}


@app.post("/api/category/favorites")
def favorites_add(payload: FavoriteAddPayload):
    import uuid as _uuid
    from datetime import datetime as _dt
    label = payload.label.strip()
    path = payload.path.strip()
    if not label or not path:
        raise HTTPException(400, "label과 path는 필수입니다.")
    favs = load_favorites()
    dup = any(
        (payload.category_id and f.get("category_id") == payload.category_id)
        or f.get("path") == path
        for f in favs
    )
    if dup:
        return {"ok": True, "favorites": favs, "duplicated": True}
    favs.append({
        "id": _uuid.uuid4().hex,
        "label": label,
        "keyword": payload.keyword.strip(),
        "path": path,
        "category_id": payload.category_id.strip(),
        "added_at": _dt.now().isoformat(timespec="seconds"),
    })
    save_favorites(favs)
    return {"ok": True, "favorites": favs}


@app.delete("/api/category/favorites/{fav_id}")
def favorites_delete(fav_id: str):
    favs = load_favorites()
    new_favs = [f for f in favs if f.get("id") != fav_id]
    if len(new_favs) == len(favs):
        raise HTTPException(404, "즐겨찾기를 찾을 수 없습니다.")
    save_favorites(new_favs)
    return {"ok": True, "favorites": new_favs}


@app.post("/api/category/stop")
def category_stop():
    cat_runner.stop()
    return {"ok": True}


@app.get("/api/category/status")
def category_status():
    return {
        "running": cat_runner.running,
        "status": cat_runner.status,
        "progress": cat_runner.progress,
        "pending_gate": cat_runner.pending_gate,
    }


@app.post("/api/category/continue")
def category_continue(body: dict = Body(default=None)):
    name = (body or {}).get("gate") or cat_runner.pending_gate
    if not name:
        raise HTTPException(400, "대기 중인 게이트가 없습니다.")
    ok = cat_runner.release_gate(name)
    if not ok:
        raise HTTPException(400, f"게이트를 찾을 수 없습니다: {name}")
    return {"ok": True}


@app.websocket("/ws/category")
async def category_ws(ws: WebSocket):
    await ws.accept()
    # 재연결 시 전체 buffer replay 방지: ?since=<ts> 이후 이벤트만 전송
    try:
        since = float(ws.query_params.get("since", "0") or "0")
    except Exception:
        since = 0.0
    while True:
        try:
            with cat_runner.lock:
                buffer = list(cat_runner.log_buffer)
                snapshot = {
                    "running": cat_runner.running,
                    "status": cat_runner.status,
                    "progress": cat_runner.progress,
                    "pending_gate": cat_runner.pending_gate,
                }
            new_since = since
            for msg in buffer:
                ts = msg.get("ts", 0) or 0
                if ts > since:
                    await ws.send_json(msg)
                    if ts > new_since:
                        new_since = ts
            since = new_since
            await ws.send_json({"type": "snapshot", **snapshot})
            await asyncio.sleep(1.5)
        except WebSocketDisconnect:
            break
        except Exception as e:
            print(f"[ws/category] send error: {e}")
            break


# ------------------ 유틸리티 (탐색기·기본앱 열기) ------------------

class PathPayload(BaseModel):
    path: str


def _resolve_within_project(path_str: str) -> Path:
    """PROJECT_ROOT 하위 경로만 허용. 나머지는 403/404."""
    if not path_str:
        raise HTTPException(400, "경로가 비어 있습니다.")
    try:
        p = Path(path_str).resolve()
    except Exception:
        raise HTTPException(400, "경로가 유효하지 않습니다.")
    root = PROJECT_ROOT.resolve()
    try:
        p.relative_to(root)
    except ValueError:
        raise HTTPException(403, "프로젝트 폴더 바깥 경로는 허용되지 않습니다.")
    if not p.exists():
        raise HTTPException(404, f"경로를 찾을 수 없습니다: {p}")
    return p


@app.post("/api/util/reveal")
def util_reveal(body: PathPayload):
    """Windows 탐색기에서 해당 경로를 연다. 파일이면 선택 상태로, 폴더면 그대로."""
    import subprocess as _sub
    import os as _os
    p = _resolve_within_project(body.path)
    try:
        if p.is_file():
            _sub.Popen(["explorer", f"/select,{str(p)}"])
        else:
            _sub.Popen(["explorer", str(p)])
        return {"ok": True, "path": str(p)}
    except Exception as e:
        raise HTTPException(500, f"탐색기 열기 실패: {e}")


@app.post("/api/util/open")
def util_open(body: PathPayload):
    """OS 기본 앱으로 파일 열기 (Windows: os.startfile)."""
    import os as _os
    p = _resolve_within_project(body.path)
    if not p.is_file():
        raise HTTPException(400, "파일이 아닙니다. (폴더를 열려면 /api/util/reveal 사용)")
    try:
        _os.startfile(str(p))
        return {"ok": True, "path": str(p)}
    except Exception as e:
        raise HTTPException(500, f"파일 열기 실패: {e}")


if __name__ == "__main__":
    print("=" * 50)
    print("  쿠팡 로켓배송 자동화 서버 시작")
    print("  브라우저 주소: http://localhost:8000")
    print("  종료: 이 창에서 Ctrl+C")
    print("=" * 50)
    uvicorn.run(app, host="127.0.0.1", port=8000, log_level="info")
