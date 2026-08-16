# [ Cell 9 ]견적서 작성 (Final_SKU50_Shelf0_Comma) SKU 50개 / 우측0 / [추가1.jpg, 추가2.jpg] 쉼표 적용
import os
import pandas as pd
import openpyxl
import gspread
import time 
from openpyxl.styles import PatternFill

print("▶️ [ Cell 9 ] 견적서 작성 (SKU=50, 우측칸=0, 추가이미지 쉼표 구분 적용)...")

# ==============================================================
# 🛠️ [헬퍼 함수] 병합된 셀 값 가져오기
# ==============================================================
def get_real_value(ws, row, col):
    cell = ws.cell(row, col)
    if cell.value is not None:
        return cell.value
    for range_ in ws.merged_cells.ranges:
        if cell.coordinate in range_:
            return ws.cell(range_.min_row, range_.min_col).value
    return None

# ==============================================================
# 1. 시트 데이터 로드
# ==============================================================
try:
    if 'doc' not in globals():
        from google.oauth2.service_account import Credentials
        scope = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
        creds = Credentials.from_service_account_file('key.json', scopes=scope)
        client = gspread.authorize(creds)
        doc = client.open_by_url(SHEET_URL)
        
    worksheet = doc.worksheet('상품등록목록')
    raw_df = pd.DataFrame(worksheet.get_all_records())
    raw_df = raw_df[raw_df['변환상품명'] != ""]
    
    detail_col_name = "상세페이지명" 
    possible_cols = ["상세이미지파일명", "상세이미지", "상세페이지"]
    for col in possible_cols:
        if col in raw_df.columns: detail_col_name = col; break
    
    folder_list_df = raw_df.drop_duplicates(subset=['변환상품명'], keep='first')
    print(f"📋 총 {len(folder_list_df)}개의 상품 견적서를 작성합니다.")

except Exception as e:
    print(f"❌ 시트 로드 실패: {e}"); folder_list_df = pd.DataFrame()

# ==============================================================
# 2. 고정값 설정 (SKU 50개 확인 필수)
# ==============================================================
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

FIXED_VALUES = {
    "FIXED_BARCODE": "바코드 없음(쿠팡 바코드 생성 요청)",
    "FIXED_BRAND": MY_BRAND_NAME if 'MY_BRAND_NAME' in globals() else "상세페이지 참조",
    "FIXED_MANUFACTURER": MY_COMPANY_NAME if 'MY_COMPANY_NAME' in globals() else "상세페이지 참조",
    "FIXED_PHONE": MY_PHONE_NUMBER if 'MY_PHONE_NUMBER' in globals() else "010-0000-0000",
    "FIXED_SIZE_CHART": "사이즈차트.jpg",
    "FIXED_TAX": "과세",
    "FIXED_DEAL_TYPE": "기타 도소매업자",
    "FIXED_IMPORT": "수입상품",
    "FIXED_SKU_COUNT": 50,  # ‼️ SKU 수량은 50으로 고정
    "FIXED_CAUTION": "해당사항없음",
    "FIXED_HANGER": "N",
    "FIXED_DIMENSION": "상세페이지참조",
    "FIXED_YEAR": 2025,
    "FIXED_SEASON": "사계절",
    "FIXED_MARK": "barcode.png",
    "FIXED_ORIGIN": "Made in China",
    "FIXED_WASH": "상세페이지 참조",
    "FIXED_WARRANTY": "제품 이상 시 공정거래위원회 고시 소비자분쟁해결 기준에 의거 보상합니다",
    "FIXED_NA": "해당사항없음",
    "FIXED_COMPOSITION": "본품",
    "FIXED_SHELF_LIFE": 0 
}

# ==== 카테고리 자동선택 20260814 ====
def _get_category_options(wb):
    import re as _re
    from openpyxl.utils import column_index_from_string as _cifs
    opts = []
    for _name in wb.defined_names:
        if 'categoryPath' in _name:
            _dest = wb.defined_names[_name].value
            _m = _re.match(r"'?([^'!]+)'?!\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)", _dest)
            if _m:
                _sh,_c1,_r1,_c2,_r2 = _m.group(1),_m.group(2),int(_m.group(3)),_m.group(4),int(_m.group(5))
                _ws = wb[_sh]; _ci1,_ci2 = _cifs(_c1),_cifs(_c2)
                for _r in range(_r1,_r2+1):
                    for _c in range(_ci1,_ci2+1):
                        _v = _ws.cell(_r,_c).value
                        if _v: opts.append(str(_v))
    return opts

def _pick_category(product_name, keyword, options):
    if not options: return None
    _text = f"{product_name} {keyword}"
    for _opt in options:
        _leaf = _opt.split('>')[-1].split('(')[0].strip()
        _core = _leaf.replace('운동기구','').replace('용품','').replace('소품','').replace('교정','').replace('/','').strip()
        if _core and _core in _text: return _opt
    for _opt in options:
        if '기타' in _opt: return _opt
    return options[0]

MAPPING_CONFIG = {
    "권장": "권장가", "소비자": "권장가", "공급가": "공급가", "판매가": "쿠팡판매가",
    "과세": "FIXED_TAX", "거래": "FIXED_DEAL_TYPE", "수입": "FIXED_IMPORT",
    "상품명": "전체옵션명", "바코드": "FIXED_BARCODE", "브랜드": "FIXED_BRAND",
    "제조사": "FIXED_MANUFACTURER", "수입자": "FIXED_MANUFACTURER",
    "검색태그": "태그", "색상": "CLEAN_COLOR_NAME", "사이즈": "사이즈", "대표이미지": "전체옵션명",
    "상세이미지": detail_col_name,          
    "수량": "CALC_SET_COUNT", 
    "차트": "FIXED_SIZE_CHART", "대체": "변환상품명",
    "구성": "FIXED_COMPOSITION", "세부사양": "메인키워드", "세부 사양": "메인키워드",
    "KC": "FIXED_NA", "인증정보": "FIXED_NA",
    "SKU": "FIXED_SKU_COUNT", "행어": "FIXED_HANGER", "포장무게": "CALC_WEIGHT",
    "단품포장": "포장사이즈", "크기": "무게", "중량": "무게", "소재": "재질", "재질": "재질", "치수": "FIXED_DIMENSION",
    "출시": "FIXED_YEAR", "계절": "FIXED_SEASON",
    "취급": "FIXED_WASH", "주의": "FIXED_WASH",
    "보증": "FIXED_WARRANTY", "AS": "FIXED_PHONE",
    "사유": "FIXED_NA", "취급주의": "FIXED_NA",
    "마크": "FIXED_NA", "타입": "FIXED_NA", "인증번호": "FIXED_NA", "EMC": "FIXED_NA",
    "안전기준": "FIXED_NA", "신고번호": "FIXED_NA", "적합확인": "FIXED_NA",
    "유통기한": "FIXED_SHELF_LIFE", "소비기한": "FIXED_SHELF_LIFE",
    "라벨": "FIXED_MARK", "도안": "FIXED_MARK", "표시사항": "FIXED_MARK",
    "제조국": "FIXED_ORIGIN", "원산지": "FIXED_ORIGIN",
    "인증/허가": "FIXED_NA", "허가사항": "FIXED_NA"
}

# ==============================================================
# 3. 메인 작업
# ==============================================================
success_count = 0
for idx, row in folder_list_df.iterrows():
    rep_img_path = str(row.get('대표이미지경로', '')).strip().rstrip("\\/")
    if not rep_img_path: continue
    rep_img_path = os.path.normpath(rep_img_path)
    prod_name = str(row.get('변환상품명', '')).strip()
    
    print(f"\n🔹 [{idx+1}] 견적서 작성: {prod_name}")

    # 견적서 파일 찾기
    quotation_path = None
    possible_paths = [rep_img_path, os.path.dirname(rep_img_path)]
    for path in possible_paths:
        if os.path.exists(path):
            files = [f for f in os.listdir(path) if "견적서" in f and f.endswith(".xlsx") and "~$" not in f]
            if files: quotation_path = os.path.join(path, files[0]); break
            
    if not quotation_path: print(f"    ⚠️ 견적서 없음 (Pass)"); continue

    src_data = raw_df[raw_df['변환상품명'].str.strip() == prod_name].to_dict('records')
    if not src_data: continue
    
    try:
        # 전처리
        for data in src_data:
            try: data['CALC_WEIGHT'] = int(''.join(filter(str.isdigit, str(data.get('무게', 0)))) or 0) + 10
            except: data['CALC_WEIGHT'] = 10
            
            opt1 = str(data.get('옵션1', '')).strip()
            full_opt = str(data.get('한글 옵션명', '')).strip()
            data['CLEAN_COLOR_NAME'] = opt1 if opt1 else full_opt

        time.sleep(0.5)
        wb_tgt = openpyxl.load_workbook(quotation_path)
        ws_tgt = None
        
        # 헤더 찾기 (상품명 있는 줄)
        header_row_idx = 5
        for sheet in wb_tgt.worksheets:
            for r in range(1, 15): 
                row_vals = [str(c.value) for c in sheet[r] if c.value]
                if any("상품명" in v for v in row_vals): 
                    ws_tgt = sheet; header_row_idx = r; break
            if ws_tgt: break
        if not ws_tgt: ws_tgt = wb_tgt.active 

        # ‼️ [위치 추적] '박스 내 SKU 수량' 열 찾기
        sku_col_idx = None
        for r in range(4, 7): # 헤더 위쪽 범위 검색
            for c in range(1, 60):
                val = str(get_real_value(ws_tgt, r, c)).replace(" ", "").replace("\n", "")
                if "박스" in val and "SKU" in val and "수량" in val:
                    sku_col_idx = c
                    break
            if sku_col_idx: break

        # 컬럼 매핑
        col_map = {}
        for col, cell in enumerate(ws_tgt[header_row_idx], 1):
            if not cell.value: continue
            
            # 비노출 속성 제외
            check_row = header_row_idx - 1
            row_4_val = str(get_real_value(ws_tgt, check_row, col)).strip().replace(" ", "")
            if "비노출" in row_4_val or "속성" in row_4_val and "비노출" in row_4_val:
                col_map[col] = {'key': None, 'clean_header': 'SKIP'} 
                continue

            clean_h = str(cell.value).replace(" ", "").replace("\n", "")
            map_key = None
            for k in sorted(MAPPING_CONFIG.keys(), key=len, reverse=True):
                if k in clean_h: map_key = MAPPING_CONFIG[k]; break
            col_map[col] = {'key': map_key, 'clean_header': clean_h}

        # 데이터 입력
        start_row = header_row_idx + 4 
        
        for i, data in enumerate(src_data):
            r_idx = start_row + i
            
            # ‼️ [핵심 로직] SKU열 오른쪽(=유통기간)에 0 넣기
            if sku_col_idx:
                shelf_life_col = sku_col_idx + 1
                ws_tgt.cell(row=r_idx, column=shelf_life_col, value=0)

            for col in range(1, ws_tgt.max_column + 1):
                # 유통기간(오른쪽 열)은 위에서 0 넣었으니 건드리지 않음
                if sku_col_idx and col == (sku_col_idx + 1): continue

                info = col_map.get(col, {'clean_header': '', 'key': None})
                if info['clean_header'] == 'SKIP': continue

                clean_h = info['clean_header']; val = None
                
                # 값 결정
                if "전화번호" in clean_h or "소비자상담" in clean_h: val = FIXED_VALUES["FIXED_PHONE"]
                
                # 사이즈차트 우선 매칭 20260516:
                # 헤더가 "사이즈차트 이미지 파일명" 인 경우, MAPPING_CONFIG 의 정렬
                # 순서상 "사이즈"(길이3) 가 "차트"(길이2) 보다 먼저 매칭되어
                # FIXED_SIZE_CHART 로 가지 않고 빈 값으로 빠지는 문제.
                # → 헤더에 "사이즈차트" 가 통째로 들어있으면 무조건 먼저 채운다.
                elif "사이즈차트" in clean_h: 
                    val = FIXED_VALUES["FIXED_SIZE_CHART"]
                
                # ‼️ [요청하신 부분 - 쉼표로 합치기 적용] ‼️
                elif "추가이미지" in clean_h:
                    # 한 칸에 쉼표로 연결해서 2장을 모두 넣습니다.
                    val = "추가1.jpg, 추가2.jpg"
                
                elif "카테고리" in clean_h:
                    try:
                        _co = _get_category_options(wb_tgt)
                        val = _pick_category(str(data.get('변환상품명','')), str(data.get('메인키워드','')), _co)
                    except Exception:
                        val = None
                elif "품명" in clean_h or "모델명" in clean_h: val = data.get('전체옵션명')
                elif info['key']:
                    target_key = info['key']
                    if target_key == "CALC_SET_COUNT":
                        set_val = data.get('세트', '1')
                        val = f"{set_val}개"
                    elif target_key in FIXED_VALUES: 
                        val = FIXED_VALUES[target_key] # 여기서 SKU 50이 들어감
                    elif target_key in data:
                        val = data[target_key]
                        if val and "이미지" in str(ws_tgt.cell(header_row_idx, col).value) and not str(val).lower().endswith('.jpg'):
                            val = f"{val}.jpg"

                if val is not None: 
                    ws_tgt.cell(row=r_idx, column=col, value=val)
                
                # 빈칸 노란색 칠하기
                curr = ws_tgt.cell(row=r_idx, column=col).value
                if (curr is None or str(curr).strip() == "") and curr != 0:
                     if "선택" not in str(ws_tgt.cell(header_row_idx+1, col).value or ""):
                        ws_tgt.cell(row=r_idx, column=col).fill = yellow_fill

        # 저장
        wait_time = 1.5 + (len(src_data) * 0.05)
        saved = False
        for attempt in range(3):
            try:
                wb_tgt.save(quotation_path)
                saved = True; break
            except PermissionError:
                print(f"    ⚠️ 엑셀 닫아주세요... ({attempt+1}/3)")
                time.sleep(3)
            except Exception as e: time.sleep(1)
        
        if saved:
            print(f"    ✅ 저장 완료"); success_count += 1
            time.sleep(wait_time)
        else: print(f"    ❌ 저장 실패")

    except Exception as e: print(f"    ❌ 실패: {e}")

print("\n" + "="*50 + f"\n🎉 견적서 완료 (성공: {success_count})\n" + "="*50)

print("\n" + "="*60)
print("✋ [최종 점검] 엑셀을 열어 확인하세요:")
print("   1. 박스 내 SKU 수량 = 50")
print("   2. 그 바로 오른쪽(유통기간) = 0")
print("   3. 추가이미지파일명 -> '추가1.jpg, 추가2.jpg' (쉼표로 구분됨)")
print("   - 확인 후 엑셀을 꼭 닫아주세요!")
print("="*60)

# [추가된 코드] 엔터 키 대기
input("\n✅ 엑셀 확인 및 닫기가 완료되었다면 [Enter] 키를 눌러 다음 단계로 넘어가세요")