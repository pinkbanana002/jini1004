"""
1단계 (Cell 0 ~ Cell 6) 통합 실행 모듈.

원본 노트북의 로직을 최대한 그대로 보존하되,
- 하드코딩된 설정값은 config dict에서 읽어옴
- print() 출력은 stdout 리다이렉션으로 log_callback에 전달
- Cell 6 말미의 input() 프롬프트는 제거 (UI 체크리스트로 대체)
"""
import os
import re
import sys
import time
from modules.detail_image_localizer import process_detail_images_in_folder


class Stopped(Exception):
    """사용자 중지 요청 시 내부적으로 발생시키는 예외."""


def _make_log_stream(log_callback):
    class LogStream:
        def __init__(self):
            self.buf = ""
        def write(self, s):
            self.buf += s
            while "\n" in self.buf:
                line, _, self.buf = self.buf.partition("\n")
                level = "info"
                if any(t in line for t in ["❌", "오류", "에러", "실패", "Error"]):
                    level = "error"
                elif any(t in line for t in ["⚠️", "경고"]):
                    level = "warn"
                elif any(t in line for t in ["✅", "🎉"]):
                    level = "success"
                log_callback(line, level=level)
        def flush(self):
            if self.buf:
                log_callback(self.buf, level="info")
                self.buf = ""
    return LogStream()


def run_stage1(config: dict, log, progress, should_stop):
    """Cell 0~6을 순차 실행. 중간에 should_stop() 참이면 Stopped 예외 발생."""
    original_stdout = sys.stdout
    original_cwd = os.getcwd()
    sys.stdout = _make_log_stream(log)

    # BASE_DRIVE를 프로젝트 루트로 고정 (원본 노트북 cwd와 동일 효과)
    project_root = config["PROJECT_ROOT"]
    os.chdir(project_root)

    driver = None
    try:
        # ==========================================================
        # [ Cell 0 ] 통합 환경 설정 — config에서 주입
        # ==========================================================
        progress(1, 7, "Cell 0: 환경 설정 로드")

        GOOGLE_JSON_FILE = config["GOOGLE_JSON_FILE"]
        GOOGLE_SHEET_URL = config["GOOGLE_SHEET_URL"]
        TARGET_SHEET_NAME = config.get("TARGET_SHEET_NAME") or "상품등록목록"
        GEMINI_API_KEY = config["GEMINI_API_KEY"]
        SUPPLY_ID = config.get("SUPPLY_ID", "")
        SUPPLY_PW = config.get("SUPPLY_PW", "")
        MY_BRAND_NAME = config.get("MY_BRAND_NAME") or "브랜드"
        MY_COMPANY_NAME = config.get("MY_COMPANY_NAME") or ""
        MY_PHONE_NUMBER = config.get("MY_PHONE_NUMBER") or ""

        EXCHANGE_FACTOR = int(float(config.get("EXCHANGE_FACTOR") or 350))
        ADD_LOGISTICS_COST = int(float(config.get("ADD_LOGISTICS_COST") or 4000))
        SALE_PRICE_RATE = float(config.get("SALE_PRICE_RATE") or 1.67)
        REC_PRICE_RATE = float(config.get("REC_PRICE_RATE") or 1.30)
        ROUND_UNIT = int(float(config.get("ROUND_UNIT") or -2))

        BASE_DRIVE = project_root
        PATH_TEMPLATES = os.path.join(BASE_DRIVE, "templates")
        PATH_FONTS = os.path.join(BASE_DRIVE, "fonts")
        FONT_BOLD = "강원교육모두 Bold.ttf"
        FONT_REGULAR = "강원교육모두 Light.ttf"
        MY_DOWNLOAD_DIR = os.path.join(os.path.expanduser("~"), "Downloads")
        PATH_PROFILE = os.path.join(BASE_DRIVE, "chrome_profile")
        PATH_DOWNLOAD = os.path.join(BASE_DRIVE, "temp_downloads")
        PATH_COMPLETED = os.path.join(BASE_DRIVE, "완료된_상품")

        for p in [PATH_TEMPLATES, PATH_FONTS, PATH_DOWNLOAD, PATH_PROFILE, PATH_COMPLETED]:
            if not os.path.exists(p):
                os.makedirs(p)

        print("-" * 50)
        print(f"📂 작업 기준 폴더: {BASE_DRIVE}")
        print(f"🚀 [환경 설정 완료]")

        # ============= 측정 로그 도구 (2026-05-15 추가) =============
        # 각 Cell의 시작/끝 시각을 기록해서 마지막에 한 번에 출력.
        # 로직 변경 없이 print 만 추가하는 진단 도구.
        _cell_timings = {}
        _stage_start_time = time.perf_counter()
        def _t_start(name):
            _cell_timings[name] = {'start': time.perf_counter()}
        def _t_end(name):
            if name in _cell_timings:
                _cell_timings[name]['elapsed'] = time.perf_counter() - _cell_timings[name]['start']
        def _t_report():
            print("\n" + "=" * 50)
            print("⏱️  1단계 구간별 소요 시간")
            print("=" * 50)
            for cell_name, data in _cell_timings.items():
                elapsed = data.get('elapsed', 0)
                bar_len = int(elapsed / 2)
                bar = "█" * min(bar_len, 30)
                print(f"  {cell_name:<28} {elapsed:>7.2f}초  {bar}")
            total = time.perf_counter() - _stage_start_time
            print("-" * 50)
            print(f"  {'총 소요 시간':<28} {total:>7.2f}초  ({total/60:.1f}분)")
            print("=" * 50)
        # ===============================================================

        # ============= 진단 로그 스위치 (2026-05-15 추가) =============
        # 평소엔 False (빠름). 문제 생겨서 진단 필요할 때만 True로 바꿈.
        # 영향: FORCE-DEBUG, CLASSIFY, GROUP-DIAG, DEBUG-DIAG, DEBUG-PRICE,
        #       DEBUG-IMG, wait_for_page 내부 DEBUG 등 진단용 print만 토글.
        # 메인 흐름 print (▶️/🎉/⚡/📌/💰/📸/✅) 는 항상 출력.
        DIAG_LOG = True

        # 진단 로그 필터: stdout 의 write 를 가로채서 진단 키워드 들어간 줄은 버림.
        # (print 자체를 가리는 방식은 Python 스코프 규칙 때문에 안전하지 않아서 stdout 레벨에서 처리)
        _DIAG_MARKERS = (
            "FORCE-DEBUG", "CLASSIFY", "GROUP-DIAG",
            "DEBUG-DIAG", "DEBUG-PRICE", "DEBUG-IMG",
            "[DEBUG]",  # wait_for_page 내부 진단 5종 + body_text 미리보기 등
        )
        if not DIAG_LOG:
            class _DiagFilter:
                def __init__(self, target):
                    self._target = target
                    self._buf = ""
                def write(self, s):
                    if not s:
                        return
                    self._buf += s
                    while "\n" in self._buf:
                        line, _, self._buf = self._buf.partition("\n")
                        # 진단 키워드 포함된 줄은 통째 스킵
                        if any(m in line for m in _DIAG_MARKERS):
                            continue
                        self._target.write(line + "\n")
                def flush(self):
                    if self._buf:
                        # 마지막 미완 줄도 필터 적용
                        if not any(m in self._buf for m in _DIAG_MARKERS):
                            self._target.write(self._buf)
                        self._buf = ""
                    if hasattr(self._target, 'flush'):
                        self._target.flush()
            sys.stdout = _DiagFilter(sys.stdout)
        # ===============================================================

        if should_stop(): raise Stopped()

        # ==========================================================
        # [ Cell 1 ] 라이브러리 import (설치는 startup에서 처리됨)
        # ==========================================================
        progress(2, 7, "Cell 1: 라이브러리 로드")
        _t_start("Cell 1: 라이브러리 로드")
        print("▶️ [ Cell 1 ] 필수 라이브러리를 로드합니다...")

        import subprocess
        import re
        import shutil
        import json
        import glob
        import random
        from io import BytesIO
        from PIL import Image, ImageDraw, ImageFont
        import textwrap

        import gspread
        from oauth2client.service_account import ServiceAccountCredentials
        import pandas as pd
        import requests

        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from webdriver_manager.chrome import ChromeDriverManager

        from gspread_dataframe import set_with_dataframe
        import openpyxl
        from openpyxl.styles import PatternFill

        import google.generativeai as genai
        from deep_translator import GoogleTranslator

        print("🚀 [성공] 모든 도구(라이브러리)가 준비되었습니다.")
        _t_end("Cell 1: 라이브러리 로드")
        if should_stop(): raise Stopped()

        # ==========================================================
        # [ Cell 2 ] 구글 시트 연결
        # ==========================================================
        progress(3, 7, "Cell 2: 구글 시트 연결")
        _t_start("Cell 2: 구글 시트 연결")
        print("▶️ [ Cell 2 ] 구글 시트와 연결을 시도합니다...")

        if not os.path.exists(GOOGLE_JSON_FILE):
            raise FileNotFoundError(f"credentials.json 파일이 없습니다: {GOOGLE_JSON_FILE}")

        gc = gspread.service_account(filename=GOOGLE_JSON_FILE)
        doc = gc.open_by_url(GOOGLE_SHEET_URL)
        # 입력 탭은 '준비시트'에서만 읽음. '상품등록목록' 탭은 stage1 Cell 5/6 + stage2 워크스페이스로만 사용.
        worksheet = doc.worksheet('준비시트')
        all_rows = worksheet.get_all_values()

        if len(all_rows) < 1:
            print("⚠️ 시트가 비어있습니다. 데이터를 채워주세요.")
            df_original = pd.DataFrame()
        else:
            headers = all_rows[0]
            data = all_rows[1:]
            df_original = pd.DataFrame(data, columns=headers)

        print(f"\n✅ '준비시트' 탭 확인 완료!")
        print(f"📦 총 {len(df_original)}개의 상품 데이터를 가져왔습니다.")
        if not df_original.empty:
            print("\n--- 📋 불러온 데이터 미리보기 (상위 3개) ---")
            print(df_original.head(3))
        print(f"🎉 구글 시트와 연결이 잘 되었습니다!")
        _t_end("Cell 2: 구글 시트 연결")
        if should_stop(): raise Stopped()

        # ==========================================================
        # [ Cell 3 ] 브라우저 실행
        # ==========================================================
        progress(4, 7, "Cell 3: 크롬 브라우저 실행")
        _t_start("Cell 3: 크롬 브라우저 실행")
        print("▶️ [ Cell 3 ] 크롬 브라우저를 설정하고 실행합니다...")

        # ==========================================================
        # [ATTACH 방식] 개인 크롬을 디버그 포트로 띄우고 붙는다 (봇 감지 우회)
        #   2단계 쿠팡(cell_8)과 동일 방식. 로그인 세션을 그대로 재사용.
        # ==========================================================
        import subprocess as _sp
        _DEBUG_PORT = 9223
        _PROFILE_DIR = os.path.join(BASE_DRIVE, "chrome_auto_profile")

        # 이 프로필+포트로 떠 있던 기존 크롬만 정리
        try:
            _cmd = ('wmic process where "name=\'chrome.exe\' and commandline like '
                    '\'%%remote-debugging-port=' + str(_DEBUG_PORT) + '%%\'" get processid')
            _out = _sp.check_output(_cmd, shell=True, encoding='utf-8', errors='replace')
            for _p in _out.split():
                if _p.isdigit():
                    os.system('taskkill /F /PID ' + _p + ' > nul 2>&1')
            time.sleep(1)
        except Exception:
            pass

        _chrome_paths = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            os.path.join(os.getenv('LOCALAPPDATA', ''), r'Google\Chrome\Application\chrome.exe'),
        ]
        _chrome_exe = next((p for p in _chrome_paths if os.path.exists(p)), None)
        if not _chrome_exe:
            raise RuntimeError("크롬 실행파일을 찾을 수 없습니다.")

        _target_url = "https://sellochomes.co.kr/sourcinglife/"
        _launch = [
            _chrome_exe,
            "--remote-debugging-port=" + str(_DEBUG_PORT),
            "--user-data-dir=" + _PROFILE_DIR,
            "--lang=ko-KR",
            "--disable-features=Translate",
            "--no-first-run",
            "--no-default-browser-check",
            "--start-maximized",
            _target_url,
        ]
        _sp.Popen(_launch)
        print("▶️ [ Cell 3 ] 크롬 실행(attach 모드) 완료. 연결 대기...")
        time.sleep(5)

        # 디버그 포트로 attach
        options = webdriver.ChromeOptions()
        options.add_experimental_option("debuggerAddress", "127.0.0.1:" + str(_DEBUG_PORT))
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
        driver.implicitly_wait(10)
        print("✅ 브라우저 연결 성공!")

        # 다운로드 경로 지정 (attach 모드에서는 prefs 대신 CDP 사용)
        try:
            driver.execute_cdp_cmd("Page.setDownloadBehavior",
                                   {"behavior": "allow", "downloadPath": MY_DOWNLOAD_DIR})
        except Exception:
            pass

        # 로그인 페이지로 튕기면 사용자 로그인 대기 (최대 3분, 자동 감지)
        try:
            _cur = driver.current_url
        except Exception:
            _cur = ""
        if "/auth/login" in _cur or "로그인" in (driver.title or ""):
            print("=" * 58)
            print("⛔ 셀록홈즈 로그인이 필요합니다.")
            print("👉 방금 열린 크롬 창에서 '구글로 시작하기' 등으로 로그인해주세요.")
            print("   로그인되면 자동으로 감지하여 진행합니다 (최대 3분 대기)...")
            print("=" * 58)
            _deadline = time.time() + 180
            while time.time() < _deadline:
                if should_stop():
                    raise Stopped()
                time.sleep(3)
                try:
                    _u = driver.current_url
                    _t = driver.title or ""
                except Exception:
                    continue
                # 로그인 페이지를 벗어나면 통과
                if "/auth/login" not in _u and "로그인" not in _t:
                    print("✅ 로그인 감지 → 자동 진행합니다.")
                    break
            else:
                print("⏰ 3분 경과 — 그대로 진행을 시도합니다.")

        # 상품 크롤링을 위해 소싱라이프 메인 확보
        try:
            if "sourcinglife" not in (driver.current_url or ""):
                driver.get(_target_url)
                time.sleep(2)
        except Exception:
            pass
        print(f"✅ 접속 성공! 현재 페이지: {driver.title}")
        _t_end("Cell 3: 크롬 브라우저 실행")
        if should_stop(): raise Stopped()

        # ==========================================================
        # [ Cell 4-1 ] 1688 상세 크롤링 (직접 URL)
        # ==========================================================
        progress(5, 7, "Cell 4-1: 1688 상품 크롤링")
        _t_start("Cell 4-1: 1688 상품 크롤링")
        print("\n▶️ [ Cell 4-1 ] 데이터 채굴 엔진 가동 🚀...")

        try:
            translator = GoogleTranslator(source='auto', target='ko')
        except Exception:
            translator = None

        crawled_buffer = []
        wait = WebDriverWait(driver, 15)

        def clean_text_for_system(text):
            if not text: return ""
            text = str(text)
            text = text.replace('*', 'x').replace('＊', 'x')
            text = re.sub(r'[\\/:*?"<>|]', '', text).strip()
            return text

        def parse_price_robust(text):
            if not text: return 0.0
            clean = "".join(re.findall(r'[\d\.]+', str(text)))
            try: return float(clean) if clean else 0.0
            except: return 0.0

        def extract_itid(link):
            m = re.search(r'itId=(\d+)', link)
            if m: return m.group(1)
            m = re.search(r'/offer/(\d+)', link)
            if m: return m.group(1)
            return None

        def wait_for_page(drv, timeout=10):
            # 진단 20260430 셀러라이프UI: 디버그 5종 + 실패 사유 (로직 변경 없음, print 만 추가)
            try:
                try:
                    print(f"    🔍 [DEBUG] current_url: {drv.current_url}")
                except Exception: pass
                try:
                    print(f"    🔍 [DEBUG] title: {drv.title}")
                except Exception: pass
                try:
                    body_preview = drv.find_element(By.TAG_NAME, 'body').text[:500]
                    print(f"    🔍 [DEBUG] body 첫500자:\n{body_preview}")
                except Exception: pass
                try:
                    oc_cnt = len(drv.find_elements(By.CSS_SELECTOR, 'img.object-contain'))
                    print(f"    🔍 [DEBUG] img.object-contain 개수: {oc_cnt}")
                except Exception: pass
                try:
                    all_imgs = drv.find_elements(By.TAG_NAME, 'img')
                    ali_cnt = sum(1 for img in all_imgs if 'alicdn' in (img.get_attribute('src') or ''))
                    print(f"    🔍 [DEBUG] alicdn img 개수: {ali_cnt} / 전체 img: {len(all_imgs)}")
                except Exception: pass
                # lazy fix 20260429 셀러라이프UI: 셀렉터 교체 — alicdn img 1개 이상 등장 대기
                # (스크롤 블록은 관련상품 가격이 body 텍스트에 섞여 parse_body_text 부작용 → 제거)
                WebDriverWait(drv, timeout).until(
                    lambda d: any(
                        'alicdn' in (img.get_attribute('src') or '')
                        for img in d.find_elements(By.TAG_NAME, 'img')
                    )
                )
                # 안정화 마진 랜덤화 20260512: 0.3 고정 → 0.15~0.25 랜덤
                time.sleep(random.uniform(0.15, 0.25))
                return True
            except Exception as e:
                try:
                    print(f"    🔍 [DEBUG] wait_for_page 실패: {type(e).__name__}: {str(e)[:200]}")
                except Exception: pass
                return False

        def parse_body_text(body_text):
            title = ''
            price_krw = 0.0
            price_cny = 0.0
            lines = [l.strip() for l in body_text.split('\n') if l.strip()]
            for i, line in enumerate(lines):
                if line == '원문보기' and i > 0:
                    title = lines[i - 1]
                if '가격 (₩' in line and i + 1 < len(lines):
                    price_krw = parse_price_robust(lines[i + 1])
                if '가격 (¥' in line and i + 1 < len(lines):
                    price_cny = parse_price_robust(lines[i + 1])
            return title, price_krw, price_cny

        # collect_opts_from_dropdown() 함수 정리 20260512:
        # 셀로크홈즈 UI 변경 이후 get_all_options() 가 컨테이너 직접 스캔 방식으로
        # 대체됐고, 이 함수는 어디서도 호출되지 않음. 죽은 코드라 통째 제거.
        # (이 함수 안에 있던 time.sleep(0.5)/0.2/0.2 도 같이 사라짐)

        def get_all_options(drv):
            # 셀러라이프 UI 변경 대응 20260429 (5차) - 옵션 정공법:
            # 새 UI: 색상 옵션이 div.group.cursor-pointer 로 마크업.
            # 페이지에 두 벌(데스크톱+모바일 등)이 렌더되어 N×2개로 매칭됨 →
            # (옵션명, 이미지URL) 조합 set 으로 중복 제거하여 unique N개만 남김.
            # 옵션 그룹 라벨은 새 UI 위치 미확인이라 '옵션' 으로 fallback.
            # 옵션별 가격은 박스 텍스트의 '¥숫자' 를 정규식으로 추출 (20260502 추가).
            # 못 찾으면 0.0 으로 두고 외부의
            # 'if o["price"] == 0.0: o["price"] = price_cny' 보정으로 본문 가격 fallback.

            # ============================================================
            # 속도 개선 20260515 (implicit_wait 임시 OFF):
            # Cell 3 에서 driver.implicitly_wait(10) 설정됨. 빈 옵션 박스 (#9~#16) 의
            # find_elements / .text 호출 시 매번 10초 타임아웃까지 기다려서
            # 상품당 약 160초 추가 소요. 옵션 박스 자체는 이미 DOM 에 다 렌더된
            # 상태이므로 implicit wait 가 필요 없음. 이 함수 안에서만 임시로 0으로
            # 낮추고 finally 에서 반드시 10으로 복원.
            # 안전망: 옵션 인식 결과는 그대로 — '✅ 색상(N개) / 사이즈(N개) 원천 데이터
            # 확보' 가 정상 출력되어야 함. 만약 0개로 떨어지면 백업으로 복원할 것.
            # ============================================================
            _orig_implicit = 10
            try:
                drv.implicitly_wait(0)
            except Exception:
                pass

            try:
                return _get_all_options_body(drv)
            finally:
                try:
                    drv.implicitly_wait(_orig_implicit)
                except Exception:
                    pass

        def _get_all_options_body(drv):

            # ============================================================
            # 임시 진단 코드 20260514 (DOM 그룹 구조 파악용, 진단 끝나면 제거 예정)
            # 색상 × 크기 같은 다중 옵션 그룹이 한 키로 합쳐지는 문제 — 그룹
            # 컨테이너/헤더 셀렉터를 찾기 위해 박스의 부모 체인과 헤더 후보를 dump.
            # 정식 분류 로직(아래)에는 영향 없음.
            # 속도 개선 20260515: 이 진단 블록이 1단계 시간의 80%(약 200초) 차지 확인.
            # 빈 옵션 박스 (#9~#16) 의 부모체인/text 추출에 execute_script + XPath
            # //*[contains(text(),'X')] 가 페이지 전체 스캔하느라 느림.
            # 진단 끝났으므로 DIAG_LOG=True 일 때만 실행.
            # ============================================================
            if DIAG_LOG:
                try:
                    _diag_boxes = drv.find_elements(
                        By.CSS_SELECTOR,
                        "div.group.flex.cursor-default.items-start, div.group.flex.items-center"
                    )
                    print(f"\n🔬 [GROUP-DIAG] 옵션 박스 {len(_diag_boxes)}개 발견")

                    # 1) 첫 박스의 부모 체인 (위로 12단계) — 공통 부모 컨테이너 추적
                    if _diag_boxes:
                        _chain = drv.execute_script("""
                            var el = arguments[0], r = [];
                            for (var i = 0; i < 12 && el; i++) {
                                r.push({
                                    tag: el.tagName,
                                    cls: (el.className || '').toString().slice(0, 180),
                                    text: (el.innerText || '').slice(0, 100).replace(/\\n/g, ' | ')
                                });
                                el = el.parentElement;
                            }
                            return r;
                        """, _diag_boxes[0])
                        print("🔬 [GROUP-DIAG] 첫 박스의 부모 체인 (위로 12단계):")
                        for _i, _n in enumerate(_chain):
                            print(f"  [{_i:02d}] <{_n['tag']} class='{_n['cls']}'>")
                            if _n['text']:
                                print(f"        text: '{_n['text']}'")

                    # 2) 헤더 후보 텍스트("색상"/"크기"/"사이즈"/"규격"/"컬러") 가진 요소 위치
                    for _kw in ['색상', '크기', '사이즈', '규격', '컬러']:
                        try:
                            _els = drv.find_elements(By.XPATH, f"//*[contains(text(),'{_kw}')]")
                            if _els:
                                print(f"🔬 [GROUP-DIAG] '{_kw}' 텍스트 가진 요소: {len(_els)}개")
                                for _el in _els[:3]:
                                    try:
                                        _cls = (_el.get_attribute('class') or '')[:140]
                                        print(f"    <{_el.tag_name} class='{_cls}'>")
                                    except Exception:
                                        pass
                        except Exception:
                            pass

                    # 3) 박스별 부모 1단계 클래스 — 같은 그룹끼리 같은 부모를 공유하는지 검사
                    if _diag_boxes:
                        print("🔬 [GROUP-DIAG] 박스별 부모 1단계 (그룹별 묶음 단서):")
                        for _bi, _b in enumerate(_diag_boxes):
                            try:
                                _pcls = drv.execute_script(
                                    "return arguments[0].parentElement ? "
                                    "arguments[0].parentElement.className : ''",
                                    _b
                                )
                                _name = (_b.text or '').strip()[:40].replace('\n', ' | ')
                                print(f"  #{_bi+1:02d} 박스='{_name}' / 부모 class='{(_pcls or '')[:120]}'")
                            except Exception:
                                pass
                except Exception as _gde:
                    try:
                        print(f"🔬 [GROUP-DIAG] 진단 실패: {_gde}")
                    except Exception:
                        pass
            # ============== 임시 진단 코드 끝 ==============

            all_opts = {}
            try:
                # 셀러라이프 UI 변경 대응 20260429 (5차-v2): 셀렉터 보강
                # 색상 옵션: cursor-pointer / 규격 모델 옵션: cursor-default 둘 다 매칭하기 위해
                # 공통 부모 클래스인 div.group.flex.items-center 사용.
                # 이미지 없는 옵션(규격 모델)은 (name,) 만으로 중복 제거.
                # 진단 디버그 20260430: F12 Console 매칭 vs Selenium 매칭 차이 원인 파악용
                # 속도 개선 20260515: DIAG_LOG=True 일 때만 진단 출력 (page_source 길이 측정이 무거움)
                if DIAG_LOG:
                    try:
                        print(f"      🔍 [DEBUG-DIAG] page_source 길이: {len(drv.page_source)}자")
                    except Exception: pass
                    try:
                        print(f"      🔍 [DEBUG-DIAG] '히비스커스' 포함 여부: {'히비스커스' in drv.page_source}")
                    except Exception: pass
                    try:
                        print(f"      🔍 [DEBUG-DIAG] '규격 모델' 포함 여부: {'규격 모델' in drv.page_source}")
                    except Exception: pass
                    try:
                        print(f"      🔍 [DEBUG-DIAG] iframe 개수: {len(drv.find_elements(By.TAG_NAME, 'iframe'))}")
                    except Exception: pass
                    try:
                        selectors_to_test = [
                            'div.group.flex.items-center',
                            'div.group.cursor-pointer',
                            'div.group.cursor-default',
                            'div.group',
                            'tr',
                            'td',
                        ]
                        for sel in selectors_to_test:
                            try:
                                count = len(drv.find_elements(By.CSS_SELECTOR, sel))
                                print(f"      🔍 [DEBUG-DIAG] {sel}: {count}개")
                            except Exception:
                                pass
                    except Exception: pass
                # 셀러라이프 UI 변경 대응 20260505:
                # 신규 UI는 옵션 컨테이너가 'div.group.flex.cursor-default.items-start' 로 바뀜.
                # 구 UI(items-center)도 함께 매칭하여 회귀 방지.
                elements = drv.find_elements(
                    By.CSS_SELECTOR,
                    "div.group.flex.cursor-default.items-start, div.group.flex.items-center"
                )
                seen = set()
                unique_options = []
                for i, el in enumerate(elements):
                    try:
                        full_text = (el.text or '').strip()
                    except Exception:
                        full_text = ''
                    # 진단 디버그 20260502: 옵션별 ¥ 가격 추출이 0.0으로 떨어지는 원인 파악용.
                    try:
                        print(f"      🔍 [DEBUG-PRICE] raw text (옵션 #{i}): {repr(full_text)[:200]}")
                    except Exception:
                        pass
                    # 옵션별 가격 추출 20260502: 옵션 박스 텍스트에서 ¥ 다음 숫자.
                    # 못 찾으면 0.0 → 외부 'if price==0.0: price=price_cny' 보정.
                    option_price = 0.0
                    try:
                        price_match = re.search(r'¥\s*([\d.]+)', full_text)
                        if price_match:
                            option_price = float(price_match.group(1))
                    except (ValueError, TypeError):
                        option_price = 0.0
                    # 셀러라이프 UI 변경 대응 20260505: 옵션명 추출 우선순위
                    #   1) <img alt="..."> — 신규 UI에서 가장 안정적
                    #   2) div.min-w-0 div.font-medium — 신규 UI fallback
                    #   3) 컨테이너 텍스트의 ¥ 앞부분 — 구 UI fallback
                    name = ''
                    try:
                        imgs = el.find_elements(By.TAG_NAME, 'img')
                        for img_el in imgs:
                            alt = (img_el.get_attribute('alt') or '').strip()
                            if alt and re.search(r'[가-힣一-鿿]', alt):
                                name = alt
                                break
                    except Exception:
                        pass
                    if not name:
                        try:
                            name_divs = el.find_elements(
                                By.CSS_SELECTOR, "div.min-w-0 div.font-medium"
                            )
                            for nd in name_divs:
                                try:
                                    text = (nd.text or '').strip()
                                except Exception:
                                    text = ''
                                if text and re.search(r'[가-힣一-鿿]', text):
                                    name = text
                                    break
                        except Exception:
                            pass
                    if not name and full_text:
                        tmp = full_text
                        if '¥' in tmp:
                            tmp = tmp.split('¥')[0].strip()
                        tmp = ' '.join(tmp.split())
                        if tmp and (re.search(r'[가-힣一-鿿]', tmp) or re.fullmatch(r'\d{1,3}', tmp)):
                            name = tmp
                    # 셀러라이프 UI 변경 대응 20260505 (필터 완화):
                    # img alt에서 직접 가져오므로 글자 수 제한을 1자+로 완화.
                    # 사이즈 옵션 누락 수정 20260515:
                    # 기존 starts_with_digit 필터가 "2XL【67.5-75kg에 적합】",
                    # "3XL【75-82.5kg에 적합】" 같은 정상 사이즈 옵션을
                    # 첫 글자가 '2'/'3'이라는 이유로 거르고 있었음.
                    # → 가격박스("¥10.3", "10") 같은 순수 숫자/통화만 거르는
                    # is_pure_number 필터로 교체.
                    text_chars_kr_cn = re.findall(r'[가-힣一-鿿]', name)
                    is_pure_number = bool(name) and re.fullmatch(r'[\d.,\s¥₩원]+', name) is not None
                    # 색상명이 없고 번호(1,2,3...)로만 구분되는 상품 대응 20260707:
                    # 1~3자리 순수 숫자는 색상 인덱스로 보고 통과, 그 외 순수 숫자/통화는 계속 제외
                    is_index_label = bool(name) and re.fullmatch(r'\d{1,3}', name) is not None
                    if (len(text_chars_kr_cn) < 1 and not is_index_label) or (is_pure_number and not is_index_label):
                        continue
                    img_src = ''
                    try:
                        imgs = el.find_elements(By.TAG_NAME, 'img')
                        if imgs:
                            img_src = imgs[0].get_attribute('src') or imgs[0].get_attribute('data-src') or ''
                    except Exception:
                        pass
                    if not name:
                        continue
                    # 박스의 부모 1단계 클래스로 그룹 라벨 결정 20260514:
                    # 'flex-wrap' 포함 → '색상', 'flex-col' 포함 → '사이즈', 둘 다 아님 → '옵션'.
                    # 1차원 list 는 그대로 유지 (단일 키 '옵션'), 각 dict 에 'group' 필드만 추가.
                    try:
                        parent_cls = drv.execute_script(
                            "return arguments[0].parentElement ? "
                            "(arguments[0].parentElement.className || '').toString() : ''",
                            el
                        ) or ''
                    except Exception:
                        parent_cls = ''
                    if 'flex-wrap' in parent_cls:
                        group_label = '색상'
                    elif 'flex-col' in parent_cls:
                        group_label = '사이즈'
                    else:
                        group_label = '옵션'
                    key = (group_label, name, img_src) if img_src else (group_label, name)
                    if key in seen:
                        continue
                    seen.add(key)
                    unique_options.append({
                        'name': name,
                        'price': option_price,
                        'img': img_src,
                        'group': group_label,
                    })
                try:
                    # 그룹 라벨 분포를 로그에 같이 표시
                    from collections import Counter as _C
                    _gcnt = _C(o.get('group', '옵션') for o in unique_options)
                    _gpv = ", ".join(f"{k}:{v}" for k, v in _gcnt.items())
                    preview = [o['name'][:20] for o in unique_options[:5]]
                    print(f"      🔍 [DEBUG] get_all_options 결과: {len(unique_options)}개 ({_gpv}) {preview}...")
                except Exception:
                    pass
                if unique_options:
                    all_opts['옵션'] = unique_options
            except Exception:
                pass
            return all_opts

        def classify_images(drv):
            # 원본 로직과 동일 (렌더크기 기준 890/790/62 분류) 유지하되
            # 웹앱 컨테이너가 좁아 rendered width가 축소되는 문제 → naturalWidth 사용
            # alicdn 이미지가 lazy load 상태일 때 src 대신 data-src를 봄
            #
            # 상세이미지 과다수집 대응 20260509:
            # ① 컨테이너 스코핑 — 페이지 전체 <img> 가 아니라 본문 영역
            #    (div.overflow-y-scroll) 안의 <img> 만 수집.
            #    관련상품/추천상품 캐러셀의 alicdn 정사각 타일이 detail 로
            #    오분류되어 detail_001~026 까지 26장 받히던 문제 차단.
            # ② 세로 비율 필터 — detail 분류 시 naturalHeight ≥ naturalWidth × 1.2
            #    인 세로형만 통과. 1688 진짜 상세 이미지는 세로 긴 스트립이고
            #    추천상품 타일은 정사각(800×800)이라 비율로 명확히 구분됨.
            #    main_imgs / thumb_imgs 분류는 비율 필터 미적용 (메인은 정사각이 정상).
            #
            # 스마트 fallback 20260509:
            # 일부 페이지(품절 등)에서 div.overflow-y-scroll 컨테이너는 존재하지만
            # 그 안에 갤러리/상세 이미지가 없는 경우(텍스트 패널만 감싸는 경우)
            # 이미지 0장 회귀 발생. 컨테이너 안 alicdn 0건이면 페이지 전체로 fallback.
            # ================================================================
            # FORCE-DEBUG 20260509 (임시 — 진단 후 제거 예정)
            # 8장에서 변화 없는 원인이 분류 로직인지 페이지 자체인지 가르기 위해
            # 컨테이너/fallback 와 무관하게 페이지 전체 alicdn img 를 raw 출력.
            # 분류 결과(detail_imgs / main_imgs / thumb_imgs)에는 영향 없음 — 본 함수의
            # 정식 분류는 이 블록 아래에서 그대로 진행.
            # ================================================================
            try:
                _fd_all = drv.find_elements(By.TAG_NAME, "img")
                print(f"    🔬 [FORCE-DEBUG] 페이지 전체 <img>: {len(_fd_all)}개 — alicdn 만 추출하여 진단")
                _fd_idx = 0
                for _fd in _fd_all:
                    try:
                        _fd_src = _fd.get_attribute('src') or ''
                    except Exception:
                        _fd_src = ''
                    try:
                        _fd_dsrc = _fd.get_attribute('data-src') or ''
                    except Exception:
                        _fd_dsrc = ''
                    if 'alicdn' not in _fd_src and 'alicdn' not in _fd_dsrc:
                        continue
                    _fd_idx += 1
                    _fd_use = _fd_src if 'alicdn' in _fd_src else _fd_dsrc
                    try:
                        _fd_w = int(drv.execute_script("return arguments[0].naturalWidth || 0;", _fd))
                    except Exception:
                        _fd_w = 0
                    try:
                        _fd_h = int(drv.execute_script("return arguments[0].naturalHeight || 0;", _fd))
                    except Exception:
                        _fd_h = 0
                    try:
                        _fd_rw = int(_fd.size.get('width', 0))
                    except Exception:
                        _fd_rw = 0
                    _fd_eff_w = _fd_w if _fd_w > 0 else _fd_rw
                    _fd_ratio = (_fd_h / _fd_w) if (_fd_w > 0 and _fd_h > 0) else 0
                    _fd_ratio_s = f"{_fd_ratio:.2f}" if _fd_ratio else "N/A"
                    # 본 분류와 동일한 로직으로 시뮬레이션 (20260509 비율 명시 분기 동기화)
                    if _fd_eff_w >= 600:
                        if _fd_h <= 0:
                            _fd_cat = "→제외(차원미상)"
                        elif _fd_w == _fd_h:
                            if _fd_w >= 1000:
                                _fd_cat = "→DETAIL"
                            else:
                                _fd_cat = "→제외(정사각)"
                        else:
                            _fd_r = _fd_h / _fd_w
                            if _fd_r >= 0.40:
                                _fd_cat = "→DETAIL"
                            else:
                                _fd_cat = "→제외(과도한가로)"
                    elif _fd_eff_w >= 500:
                        _fd_cat = "→MAIN"
                    elif _fd_eff_w == 62:
                        _fd_cat = "→THUMBNAIL"
                    else:
                        _fd_cat = "→제외(너비부족)"
                    _fd_tail = _fd_use[-30:] if len(_fd_use) > 30 else _fd_use
                    print(f"    🔬 [FORCE-DEBUG] {_fd_idx:02d}: nat={_fd_w}x{_fd_h} ratio={_fd_ratio_s} rendered_w={_fd_rw} eff={_fd_eff_w} {_fd_cat} ...{_fd_tail}")
                print(f"    🔬 [FORCE-DEBUG] 페이지 전체 alicdn 이미지 총 {_fd_idx}장")
            except Exception as _fd_e:
                try:
                    print(f"    🔬 [FORCE-DEBUG] 진단 실패: {type(_fd_e).__name__}: {str(_fd_e)[:200]}")
                except Exception: pass
            # ================ FORCE-DEBUG 끝 (이하 정식 분류 로직) ================
            main_imgs = []
            thumb_imgs = []
            detail_imgs = []
            container = None
            try:
                container = drv.find_element(By.CSS_SELECTOR, 'div.overflow-y-scroll')
            except Exception:
                container = None
            # fallback 임계값 20260509:
            # 정상 상품은 메인+썸네일+상세 합쳐 alicdn 15장+ 이 정상.
            # 컨테이너 안 alicdn 이 5장 미만이면 컨테이너가 너무 좁게 잡힌 것으로 보고
            # 페이지 전체로 fallback. 0만 체크하면 1~4장만 보이는 케이스에서
            # '메인 1 / 썸네일 0 / 상세 0' 회귀가 발생함.
            #
            # 비율 기반 fallback 추가 20260509:
            # 컨테이너에 alicdn 5장 이상 있어도 페이지 전체 alicdn 의 1/3 미만이면
            # 진짜 상세이미지가 컨테이너 밖에 있는 경우 (예: 정사각 갤러리만 컨테이너
            # 안에 있고 가로형 상세 스트립은 별도 패널에 있는 페이지). 이때도 fallback.
            MIN_SCOPED_ALICDN = 5
            FALLBACK_PAGE_TO_SCOPED_RATIO = 3  # page_alicdn > scoped_alicdn × 3 면 fallback
            fallback_fired = False
            fallback_reason = ""
            page_imgs = drv.find_elements(By.TAG_NAME, "img")
            page_alicdn = sum(
                1 for im in page_imgs
                if 'alicdn' in (im.get_attribute('src') or '')
                or 'alicdn' in (im.get_attribute('data-src') or '')
            )
            if container is not None:
                scoped_imgs = container.find_elements(By.TAG_NAME, "img")
                scoped_alicdn = sum(
                    1 for im in scoped_imgs
                    if 'alicdn' in (im.get_attribute('src') or '')
                    or 'alicdn' in (im.get_attribute('data-src') or '')
                )
                try:
                    print(f"    🔍 [DEBUG-IMG] 컨테이너 안 img: {len(scoped_imgs)}장 (alicdn: {scoped_alicdn}장) / 페이지 전체 alicdn: {page_alicdn}장")
                except Exception: pass
                if scoped_alicdn < MIN_SCOPED_ALICDN:
                    fallback_fired = True
                    fallback_reason = f"컨테이너 안 alicdn {scoped_alicdn}장 < 임계 {MIN_SCOPED_ALICDN}장 (개수 부족)"
                    all_imgs = page_imgs
                elif page_alicdn > scoped_alicdn * FALLBACK_PAGE_TO_SCOPED_RATIO:
                    fallback_fired = True
                    fallback_reason = (
                        f"페이지 전체 alicdn {page_alicdn}장 > 컨테이너 {scoped_alicdn}장 × {FALLBACK_PAGE_TO_SCOPED_RATIO} "
                        f"(컨테이너 밖에 진짜 상세 의심)"
                    )
                    all_imgs = page_imgs
                else:
                    all_imgs = scoped_imgs
            else:
                try:
                    print(f"    🔍 [DEBUG-IMG] 컨테이너 div.overflow-y-scroll 없음 / 페이지 전체 alicdn: {page_alicdn}장")
                except Exception: pass
                fallback_fired = True
                fallback_reason = "컨테이너 div.overflow-y-scroll 미발견"
                all_imgs = page_imgs
            try:
                if fallback_fired:
                    print(f"    🔍 [DEBUG-IMG] fallback 발동: True — 사유: {fallback_reason}")
                else:
                    print(f"    🔍 [DEBUG-IMG] fallback 발동: False")
            except Exception: pass
            # 크기+비율 조합 필터 20260509 (방향 C):
            # DETAIL 후보 진입 조건: nat_w >= 600 (작은 추천/아이콘 차단)
            # DETAIL 통과 조건: 정사각(nat_w == nat_h) 아님 그리고
            #   (a) 세로형: nat_h >= nat_w * 1.05  또는
            #   (b) 적당한 가로형: nat_w <= nat_h * 1.6 (= 가로/세로 ≤ 1.6)
            # 색상 비교샷 같은 가로형 detail (예: 715x495, ratio 1.44) 을 살리고
            # 광폭 배너 / 정사각 추천 타일은 차단.
            img_idx = 0
            for img in all_imgs:
                src = img.get_attribute('src') or ''
                data_src = img.get_attribute('data-src') or ''
                if 'alicdn' in src:
                    use_src = src
                elif 'alicdn' in data_src:
                    use_src = data_src
                else:
                    continue
                img_idx += 1
                try:
                    nat_w = int(drv.execute_script("return arguments[0].naturalWidth || 0;", img))
                except Exception:
                    nat_w = 0
                try:
                    nat_h = int(drv.execute_script("return arguments[0].naturalHeight || 0;", img))
                except Exception:
                    nat_h = 0
                w = nat_w if nat_w > 0 else img.size['width']
                classification = "제외(기타)"
                if nat_w >= 600:
                    # 비율 명시 분기 20260509:
                    # ratio = nat_h / nat_w 로 명시 계산하여 분류.
                    # - 세로형 (ratio >= 1.05): DETAIL
                    # - 가로형 (0.40 <= ratio < 1.05): DETAIL (납작한 띠 제외)
                    # - 정사각 (ratio == 1.0): nat_w >= 1000 만 DETAIL 통과
                    # - 너무 납작 (ratio < 0.40): 제외(과도한가로)
                    if nat_h <= 0:
                        classification = "제외(차원미상)"
                    elif nat_w == nat_h:
                        if nat_w >= 1000:
                            classification = "DETAIL"
                            detail_imgs.append(use_src)
                        else:
                            classification = "제외(정사각)"
                    else:
                        ratio = nat_h / nat_w
                        if ratio >= 0.40:
                            classification = "DETAIL"
                            detail_imgs.append(use_src)
                        else:
                            classification = "제외(과도한가로)"
                elif w >= 500:
                    classification = "MAIN"
                    main_imgs.append(use_src)
                elif w == 62:
                    classification = "THUMBNAIL"
                    thumb_imgs.append(use_src)
                else:
                    classification = "제외(너비부족)"
                try:
                    if nat_w > 0 and nat_h > 0:
                        ratio_str = f"{nat_h / nat_w:.2f}"
                    else:
                        ratio_str = "N/A"
                    print(f"      🔍 [CLASSIFY] {img_idx:02d}: {nat_w}x{nat_h} ratio={ratio_str} → {classification}")
                except Exception: pass
            if not main_imgs and thumb_imgs:
                main_imgs = [thumb_imgs[0]]
            try:
                print(f"    🔍 [DEBUG-IMG] 분류 결과 — 메인:{len(main_imgs)} / 썸네일:{len(thumb_imgs)} / 상세:{len(detail_imgs)}")
            except Exception: pass
            return main_imgs, thumb_imgs, detail_imgs

        for index, row in df_original.iterrows():
            if should_stop(): raise Stopped()
            original_row_dict = row.to_dict()
            link_to_search = str(original_row_dict.get('상품링크', '')).strip()
            if not link_to_search or str(original_row_dict.get('다운로드여부', '')).upper() == 'TRUE':
                crawled_buffer.append({'type': 'skip', 'data': original_row_dict})
                continue
            print(f"\n⚡ [{index+2}] 수집 시작: {link_to_search[:50]}...")

            # ===== 세부 측정 시작 (2026-05-15 추가) =====
            _sub_t = {}
            _sub_total_start = time.perf_counter()
            # ============================================

            try:
                item_id = extract_itid(link_to_search)
                if not item_id:
                    raise ValueError(f"itId 추출 실패: {link_to_search}")
                direct_url = f"https://sellochomes.co.kr/sourcinglife/alibaba/item?itId={item_id}"

                # [1] 페이지 이동
                _t1 = time.perf_counter()
                driver.get(direct_url)
                _sub_t['1.driver.get(페이지이동)'] = time.perf_counter() - _t1

                # [2] 페이지 렌더 대기
                _t2 = time.perf_counter()
                rendered = wait_for_page(driver, timeout=10)
                _sub_t['2.wait_for_page(렌더대기)'] = time.perf_counter() - _t2
                if not rendered:
                    raise ValueError("페이지 렌더링 타임아웃")

                # [3] body_text 추출 + 가격 파싱
                _t3 = time.perf_counter()
                # 셀러라이프 UI 변경 대응 20260429:
                # body.innerText 는 '관련상품' 영역의 가격이 앞쪽에 위치하여
                # parse_body_text 의 가격 마커가 오작동함.
                # F12 로 확인한 본문 영역(상품명/원문보기/가격/옵션 모두 포함)은
                # div.overflow-y-scroll 임. 해당 영역의 innerText 만 사용.
                #
                # 셀러라이프 UI 변경 대응 20260429 (2차):
                # 가격 영역에 할인 전(취소선) + 할인 후 가격이 동시 노출됨.
                # 취소선 span(class="line-through")을 DOM 에서 제거하면
                # parse_body_text 가 깨끗한 가격 1개만 보게 됨.
                try:
                    driver.execute_script("""
                        document.querySelectorAll('span.line-through').forEach(
                            function(el) { el.remove(); }
                        );
                    """)
                except Exception:
                    pass

                try:
                    body_text = driver.find_element(
                        By.CSS_SELECTOR, 'div.overflow-y-scroll'
                    ).text
                except Exception:
                    body_text = driver.find_element(By.TAG_NAME, 'body').text
                print(f"      🔍 [DEBUG] body_text 첫 300자: {body_text[:300]}")
                # 셀러라이프 UI 변경 대응 20260429 (3차):
                # 가격 범위 형식 "₩A ~ ₩B" / "¥A ~ ¥B" → 첫 가격만 남김.
                # 이유: parse_body_text 가 두 숫자를 join 하면서
                #   - ₩: 정수 두개 합쳐져 8자리 폭주 (3437+3619 → 34373619)
                #   - ¥: 소수 두개 합쳐져 점 2개 → float 변환 실패 → 0.0
                # 함수의 어제 정상 입력(숫자 묶음 1개 한 줄)과 동일하게 맞춰줌.
                body_text = re.sub(
                    r'(₩\s*[\d,]+(?:\.\d+)?)\s*~\s*₩\s*[\d,]+(?:\.\d+)?',
                    r'\1',
                    body_text
                )
                body_text = re.sub(
                    r'(¥\s*[\d,]+(?:\.\d+)?)\s*~\s*¥\s*[\d,]+(?:\.\d+)?',
                    r'\1',
                    body_text
                )
                print(f"      🔍 [DEBUG] 가격범위 정제 후 body_text 첫 200자: {body_text[:200]}")
                title, price_krw, price_cny = parse_body_text(body_text)
                print(f"    📌 상품명: {title[:40]}")
                print(f"    💰 가격: ₩{price_krw:,.0f} / ¥{price_cny}")
                _sub_t['3.body_text+가격파싱'] = time.perf_counter() - _t3

                # [4] 스크롤 루프 (lazy load)
                _t4 = time.perf_counter()
                # 레이지 로딩 대응: 상세 이미지가 DOM에 로드되도록 페이지 끝까지 스크롤
                # 속도 개선 20260515: sleep 0.7→0.25~0.35 / 안전 위해 2회 연속 안정 시 종료 / 상한 15회
                last_h = driver.execute_script("return document.body.scrollHeight")
                stable_count = 0
                _scroll_iters = 0
                for _ in range(15):
                    _scroll_iters += 1
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(random.uniform(0.25, 0.35))
                    new_h = driver.execute_script("return document.body.scrollHeight")
                    if new_h == last_h:
                        stable_count += 1
                        if stable_count >= 2:
                            break
                    else:
                        stable_count = 0
                        last_h = new_h
                driver.execute_script("window.scrollTo(0, 0);")
                # 페이지 상단 복귀 안정화 랜덤화 20260512: 0.4 고정 → 0.2~0.3 랜덤
                time.sleep(random.uniform(0.2, 0.3))
                _sub_t[f'4.스크롤루프({_scroll_iters}회)'] = time.perf_counter() - _t4

                # [5] classify_images (이미지 분류)
                _t5 = time.perf_counter()
                main_imgs, thumb_imgs, detail_imgs = classify_images(driver)
                _sub_t['5.classify_images(이미지분류)'] = time.perf_counter() - _t5

                # [6] get_all_options (옵션 박스 추출)
                _t6 = time.perf_counter()
                options_dict = get_all_options(driver)
                _sub_t['6.get_all_options(옵션추출)'] = time.perf_counter() - _t6
                # 그룹 메타 분리 20260514: get_all_options 는 1차원 list 를 단일 키 '옵션'으로 반환.
                # 각 옵션 dict 의 'group' 메타로 색상/사이즈/옵션 그룹 분리.
                # 그룹 1개만 등장하면 자동으로 opt2_final 빈 list → 기존 동작 그대로.
                flat_options = options_dict.get('옵션') or next(iter(options_dict.values()), [])
                _groups = {}
                for _o in flat_options:
                    _g = _o.get('group', '옵션')
                    _groups.setdefault(_g, []).append(_o)
                _gkeys = list(_groups.keys())
                opt1_key = _gkeys[0] if _gkeys else '옵션1'
                opt2_key = _gkeys[1] if len(_gkeys) > 1 else None
                opt1_final = _groups.get(opt1_key, [])
                opt2_final = _groups.get(opt2_key, []) if opt2_key else []
                for o in opt1_final + opt2_final:
                    if o['price'] == 0.0:
                        o['price'] = price_cny
                opt1_has_img = any(o['img'] for o in opt1_final)
                opt2_has_img = any(o['img'] for o in opt2_final)
                is_switched = (
                    bool(opt1_final) and bool(opt2_final)
                    and not opt1_has_img
                    and opt2_has_img
                )
                first_p = opt2_final[0]['price'] if opt2_final else (opt1_final[0]['price'] if opt1_final else 0.0)
                print(f"    💰 가격분석 완료 첫번째 가격: {first_p}")
                if is_switched:
                    print(f"    📸 옵션이미지({opt2_key}): {len([o for o in opt2_final if o['img']])}장")
                    print(f"    🔄 [스위칭 적용] {opt2_key}가 메인 이미지로 설정됩니다.")
                else:
                    print(f"    📸 옵션이미지({opt1_key}): {len([o for o in opt1_final if o['img']])}장")
                if opt2_final:
                    print(f"    ✅ {opt1_key}({len(opt1_final)}개) / {opt2_key}({len(opt2_final)}개) 원천 데이터 확보")
                elif opt1_final:
                    print(f"    ✅ {opt1_key}({len(opt1_final)}개) 단독 옵션 확보")
                else:
                    print(f"    ✅ 옵션 없음 - 단일 상품")
                print(f"    🖼️ 메인:{len(main_imgs)}장 / 썸네일:{len(thumb_imgs)}장 / 상세:{len(detail_imgs)}장")
                crawled_buffer.append({
                    'type': 'new',
                    'original_row': original_row_dict,
                    'common_info': {
                        'orig_name': clean_text_for_system(title),
                        'main_imgs': main_imgs,
                        'thumb_imgs': thumb_imgs,
                        'detail_urls': detail_imgs,
                    },
                    'opt1': opt1_final,
                    'opt2': opt2_final,
                    'is_switched': is_switched,
                    'price_krw': price_krw,
                    'price_cny': price_cny,
                })
            except Exception as e:
                print(f"❌ 오류 발생: {e}")
                crawled_buffer.append({'type': 'error', 'data': original_row_dict})

            # ===== 상품별 세부 측정 결과 출력 (2026-05-15 추가) =====
            try:
                _sub_total = time.perf_counter() - _sub_total_start
                print(f"\n  ⏱️ [상품 #{index+2} 세부 측정] 총 {_sub_total:.2f}초")
                for _stg_name, _stg_elapsed in _sub_t.items():
                    _bar_len = int(_stg_elapsed * 2)
                    _bar = "▓" * min(_bar_len, 30)
                    _pct = (_stg_elapsed / _sub_total * 100) if _sub_total > 0 else 0
                    print(f"     {_stg_name:<30} {_stg_elapsed:>6.2f}초  ({_pct:>4.1f}%)  {_bar}")
            except Exception:
                pass
            # =====================================================

            # 상품 간 대기 완화 20260512: 2.0~4.0 → 1.3~2.6 (평균 약 1초 단축)
            # 셀로크홈즈는 봇 감지가 약해서 더 짧게도 안전. 랜덤 폭은 유지해 인간 패턴.
            _t7 = time.perf_counter()
            time.sleep(random.uniform(1.3, 2.6))
            try:
                print(f"     {'7.상품간 대기 sleep':<30} {time.perf_counter()-_t7:>6.2f}초")
            except Exception:
                pass

        print(f"\n🎉 [ Cell 4-1 ] 데이터 원천 수집 완료! 총 {len(crawled_buffer)}건 처리")
        _t_end("Cell 4-1: 1688 상품 크롤링")
        if should_stop(): raise Stopped()

        # ==========================================================
        # [ Cell 4-2 ] 데이터 가공 + 원가 + 이미지 다운로드
        # ==========================================================
        progress(6, 7, "Cell 4-2: 데이터 가공 + 이미지 다운로드")
        _t_start("Cell 4-2: 데이터 가공 + 다운로드")
        print("\n▶️ [ Cell 4-2 ] 데이터 가공 및 이미지 다운로드 시작...")

        from urllib.parse import urlparse
        from concurrent.futures import ThreadPoolExecutor

        TRANS_CACHE = {}
        try:
            ko_to_zh = GoogleTranslator(source='ko', target='zh-CN')
        except Exception:
            ko_to_zh = None

        stats = {'products': 0, 'options': 0}
        download_tasks = []
        det_dirs = []

        def purify_option_name(name):
            if not name: return ""
            pure = re.split(r'(재고|가격|개|元|위안|:)', str(name))[0].strip()
            return re.sub(r'[\\/*?:"<>|]', "", pure).strip()

        def get_cached_translation(text):
            if not text: return ""
            if not ko_to_zh: return text
            if text in TRANS_CACHE: return TRANS_CACHE[text]
            try:
                translated = ko_to_zh.translate(text)
                TRANS_CACHE[text] = translated
                return translated
            except Exception:
                return text

        def execute_download_safe_debug(task):
            url, save_path, info_msg = task
            if os.path.exists(save_path): return (True, url, "이미 존재", info_msg)
            if not url: return (False, "URL 없음", "", info_msg)
            if url.startswith("//"): url = "https:" + url
            if not url.startswith("http"): return (False, f"잘못된 URL: {url}", "", info_msg)
            headers = {"User-Agent": "Mozilla/5.0"}
            for attempt in range(3):
                try:
                    r = requests.get(url, headers=headers, timeout=10)
                    if r.status_code == 200:
                        with open(save_path, 'wb') as f: f.write(r.content)
                        return (True, url, "성공", info_msg)
                except Exception:
                    time.sleep(random.uniform(0.5, 1.5))
            return (False, url, "타임아웃/연결실패", info_msg)

        results_list = []
        if not crawled_buffer:
            print("❌ [오류] 수집된 데이터가 없습니다.")
        else:
            final_sheet_rows = []
            valid_items = [i for i in crawled_buffer if i['type'] not in ['skip', 'error']]
            print(f"⚡ 총 {len(valid_items)}개 상품 처리 및 다운로드 예약 중...")
            for item in valid_items:
                stats['products'] += 1
                orig = item['original_row']
                info = item.get('common_info', {})
                opt1_list = item.get('opt1', [])
                opt2_list = item.get('opt2', [])
                is_switched = item.get('is_switched', False)
                # 셀러라이프 UI 변경 대응 20260429 (4차) - 옵션 추출 fallback:
                # 새 UI 에서 get_all_options 가 옵션 0개 반환하는 경우 발생.
                # 옵션 추출 정공법 수정 전까지 임시 fallback —
                # opt1_list 가 비어있으면 단일 fake 옵션 1개 생성해서
                # 최소 1행이라도 시트에 올라가도록 함.
                # 메인 가격/이미지 정보는 살아있음. 색상 옵션만 손실.
                if not opt1_list:
                    fb_imgs = info.get('main_imgs') or info.get('thumb_imgs') or info.get('detail_urls') or []
                    opt1_list = [{
                        'name': '',
                        'price': item.get('price_cny', 0.0),
                        'img': fb_imgs[0] if fb_imgs else '',
                    }]
                    print("      ⚠️ [DEBUG] 옵션 0개 → fallback 단일행 생성 (UI 변경 대응)")
                prod_name = f"{MY_BRAND_NAME} {info.get('orig_name', '상품명없음')}"
                base_dir = os.path.abspath(purify_option_name(prod_name))
                img_dir = os.path.join(base_dir, "대표이미지")
                det_dir = os.path.join(base_dir, "상세페이지")
                add_img_dir = os.path.join(det_dir, "추가이미지")
                for p in [img_dir, det_dir, add_img_dir]: os.makedirs(p, exist_ok=True)
                det_dirs.append(det_dir)
                for i, url in enumerate(info.get('detail_urls', [])):
                    save_path = os.path.join(det_dir, f"detail_{i+1:03d}.jpg")
                    task_info = f"[{prod_name}] 상세이미지 {i+1}번"
                    download_tasks.append((url, save_path, task_info))
                try: val_set = int(str(orig.get('세트', '1')).strip())
                except: val_set = 1
                target_opt2 = opt2_list if opt2_list else [{'name': '', 'price': 0, 'img': ''}]
                for o1 in opt1_list:
                    for o2 in target_opt2:
                        stats['options'] += 1
                        clean_n1 = purify_option_name(o1['name'])
                        clean_n2 = purify_option_name(o2['name'])
                        if val_set > 1: clean_n1 = f"{clean_n1} {val_set}개"
                        main_opt_img = o2.get('img', '') if is_switched else o1.get('img', '')
                        final_ko1 = clean_n2 if is_switched else clean_n1
                        final_ko2 = clean_n1 if is_switched else clean_n2
                        china_opt_name = get_cached_translation(final_ko1)
                        final_price = o2['price'] if o2['price'] > 0 else (o1['price'] if o1['price'] > 0 else item.get('base_price', 0))
                        temp_fname = f"TEMP_{final_ko1}.jpg"
                        target_img_path = os.path.join(img_dir, temp_fname)
                        if main_opt_img:
                            task_info = f"[{prod_name}] 옵션: {final_ko1}"
                            download_tasks.append((main_opt_img, target_img_path, task_info))
                        new_row = orig.copy()
                        new_row.update({
                            '변환상품명': prod_name,
                            '원본상품명': info.get('orig_name'),
                            '대표이미지경로': img_dir,
                            '중국 옵션명': china_opt_name,
                            '옵션1': final_ko1,
                            '옵션2': final_ko2,
                            '대표이미지링크': main_opt_img,
                            '1688': final_price,
                            '추가이미지명': '추가1.jpg',
                            '추가이미지명2': '추가2.jpg',
                            '임시파일명': temp_fname,
                            '다운로드여부': 'TRUE',
                        })
                        final_sheet_rows.append(new_row)

            print(f"    👉 데이터 매칭 완료. {len(download_tasks)}개 이미지 안전 다운로드 시작...")
            start_time = time.time()
            failed_list = []
            with ThreadPoolExecutor(max_workers=20) as executor:
                results = list(executor.map(execute_download_safe_debug, download_tasks))
            end_time = time.time()
            success_cnt = 0
            for res in results:
                is_ok, url, msg, info = res
                if is_ok: success_cnt += 1
                else: failed_list.append(f"{info} -> {url}")
            print(f"⏱️ 다운로드 완료! 소요: {end_time - start_time:.2f}초")
            print(f"   ✅ 성공: {success_cnt}장")
            print(f"   ❌ 실패: {len(failed_list)}장")
            detail_text_mode = os.getenv("DETAIL_IMAGE_TEXT_MODE", "none")
            if detail_text_mode != "none":
                print(f"\n🈶 [상세페이지 중국어 텍스트 처리] mode={detail_text_mode}")
                for dd in det_dirs:
                    try:
                        process_detail_images_in_folder(dd, mode=detail_text_mode)
                    except Exception as e:
                        print(f"    ⚠️ 처리 실패 ({dd}): {e}")
            for i, row in enumerate(final_sheet_rows, start=2):
                row['원가'] = f'=S{i}*E{i}*{EXCHANGE_FACTOR}'
                row['공급가'] = f'=ROUND(T{i}+{ADD_LOGISTICS_COST}, {ROUND_UNIT})'
                row['쿠팡판매가'] = f'=ROUND(U{i}*{SALE_PRICE_RATE}, {ROUND_UNIT})'
                row['권장가'] = f'=ROUND(V{i}*{REC_PRICE_RATE}, {ROUND_UNIT})'
                row['내마진'] = f'=ROUND(U{i}-T{i}, {ROUND_UNIT})'
                row['쿠팡마진'] = f'=ROUND(V{i}-U{i}, {ROUND_UNIT})'
                row['한글 옵션명'] = f'=TRIM(J{i} & " " & K{i})'
                row['전체옵션명'] = f'=TRIM(G{i} & " " & L{i})'
            results_list = final_sheet_rows
            print(f"\n🎉 [ Cell 4-2 ] 처리 완료! (세트 원가 반영됨)")
        _t_end("Cell 4-2: 데이터 가공 + 다운로드")
        if should_stop(): raise Stopped()

        # ==========================================================
        # [ Cell 5 ] 구글시트 업로드
        # ==========================================================
        progress(7, 7, "Cell 5: 구글시트 업로드")
        _t_start("Cell 5: 구글시트 업로드")
        print(f"\n▶️ [ Cell 5 ] 최종 데이터를 구글 시트에 업로드합니다...")
        if results_list:
            worksheet = doc.worksheet(TARGET_SHEET_NAME)
            df_final = pd.DataFrame(results_list).fillna("")
            standard_headers = [
                '재질', '상품링크', '포장사이즈', '무게', '세트',
                '원본상품명', '변환상품명', '메인키워드', '대표이미지경로',
                '옵션1', '옵션2', '한글 옵션명', '중국 옵션명',
                'COPY폴더 경로', '사이즈', '태그', '대표이미지링크',
                '전체옵션명',
                '1688', '원가', '공급가', '쿠팡판매가', '권장가', '내마진', '쿠팡마진',
                '임시파일명', '상세이미지파일명',
                '추가이미지명', '추가이미지명2',
            ]
            for col in standard_headers:
                if col not in df_final.columns: df_final[col] = ""
            df_final = df_final[standard_headers]
            # 보호 ① 20260514: 메인 시트 와이프 방지 안전망.
            # worksheet 가 메인 탭(TARGET_SHEET_NAME) 가리키는지, df_final 이 비어있지 않은지
            # 둘 다 확인하고 clear/업로드 진행. 둘 중 하나라도 어긋나면 통째 차단.
            if worksheet.title != TARGET_SHEET_NAME:
                raise RuntimeError(
                    f"⚠️ worksheet 변수가 메인 탭이 아닙니다: '{worksheet.title}' "
                    f"(예상: '{TARGET_SHEET_NAME}'). 시트 와이프 방지 차원에서 업로드 중단."
                )
            if df_final.empty:
                print("⚠️ [보호 ①] df_final 비어있음 — clear/업로드 통째 스킵 (기존 시트 보존)")
            else:
                print(f"    🧹 시트('{worksheet.title}')를 초기화하고 {len(df_final)}행 데이터를 올립니다...")
                worksheet.clear()
                set_with_dataframe(worksheet, df_final, include_column_header=True)
                try: worksheet.format("1:1", {"textFormat": {"bold": True}})
                except Exception: pass
                print(f"🎉 [업로드 성공]")
        else:
            print("⚠️ [경고] 업로드할 데이터가 없습니다.")
        _t_end("Cell 5: 구글시트 업로드")
        if should_stop(): raise Stopped()

        # ==========================================================
        # [ Cell 6 ] AI 분석 + 중국어 옵션 번역 + 백업
        # ==========================================================
        progress(7, 7, "Cell 6: AI 분석 + 번역")
        _t_start("Cell 6: AI 분석 + 번역")
        print("\n▶️ [ Cell 6 ] AI 분석 및 중국어 옵션 번역을 시작합니다...")

        if not GEMINI_API_KEY:
            raise ValueError("Gemini API Key가 설정되지 않았습니다.")

        model = None
        try:
            genai.configure(api_key=GEMINI_API_KEY)
            try:
                model = genai.GenerativeModel('gemini-2.0-flash')
                _ = model.generate_content("test")
                print("✅ Gemini AI 연결 성공! (gemini-2.0-flash)")
            except Exception:
                model = genai.GenerativeModel('gemini-flash-latest')
                print("✅ Gemini AI 연결 성공! (gemini-flash-latest)")
        except Exception as e:
            print(f"❌ Gemini AI 연결 실패: {e}")
            model = None

        translator_ai = GoogleTranslator(source='auto', target='ko')

        # \ubc88\uc5ed \ub2e8\uc704 \ubcf4\uc874 20260514:
        # deep_translator(GoogleTranslator) \ubb34\ub8cc \ubc88\uc5ed\uae30\uac00 "\u9002\u540880-110\u78c5" \uac19\uc740
        # \ud14d\uc2a4\ud2b8\ub97c "10\ub300\uc5d0 \uc801\ud569" \ucc98\ub7fc \ub2e8\uc704 \uc790\uccb4\ub97c \uc758\uc5ed\ud574\ubc84\ub9ac\ub294 \ubb38\uc81c \ud68c\ud53c\uc6a9.
        # \uc6d0\ubcf8\uc5d0\uc11c \uc22b\uc790+\ub2e8\uc704 \ud1a0\ud070\uc744 \ubbf8\ub9ac \ucd94\ucd9c \u2192 \ubc88\uc5ed \ud6c4 \ud55c\uad6d\uc5b4 \ub2e8\uc704 \ud0a4\uc6cc\ub4dc\uac00
        # \uc0b4\uc544\uc788\ub294\uc9c0 \uac80\uc0ac \u2192 \ub204\ub77d\ub41c \ud1a0\ud070\ub9cc \ud55c\uad6d\uc5b4\ub85c \uce58\ud658\ud574 \uacb0\uacfc \ub4a4\uc5d0 \ucca8\ubd80.
        # \ud328\ud134: \uc815\uc218/\uc18c\uc218 + (\uc120\ud0dd) ~ \ub610\ub294 - \ubc94\uc704 + \ub2e8\uc704 \ud0a4\uc6cc\ub4dc.
        _unit_pattern = re.compile(
            r'\d+(?:\.\d+)?\s*[~\-]\s*\d+(?:\.\d+)?\s*(?:cm|mm|kg|lb|inch|\u82f1\u5bf8|\u5bf8|\u65a4|\u78c5)'
            r'|\d+(?:\.\d+)?\s*(?:cm|mm|kg|lb|inch|\u82f1\u5bf8|\u5bf8|\u65a4|\u78c5)',
            re.IGNORECASE
        )
        # \uc911\uad6d\uc5b4 \ub2e8\uc704 \u2192 \ud55c\uad6d\uc5b4 \ub2e8\uc704 \ub9e4\ud551. cm/mm/kg/lb/inch \ub294 \ud55c\uad6d\uc5b4 \ud14d\uc2a4\ud2b8\uc5d0\uc11c\ub3c4
        # \uadf8\ub300\ub85c \uc4f0\uc774\ubbc0\ub85c \ub9e4\ud551 \ubd88\ud544\uc694(\ubc88\uc5ed\uae30\uac00 \uc0b4\ub824\ub450\uba74 \uadf8\ub300\ub85c \ud1b5\uacfc).
        _unit_ko_map = {
            '\u65a4': '\uadfc',
            '\u78c5': '\ud30c\uc6b4\ub4dc',
            '\u82f1\u5bf8': '\uc778\uce58',
            '\u5bf8': '\uc778\uce58',
        }

        def _normalize_unit_token(token):
            """\uc6d0\ubcf8 \ud1a0\ud070\uc758 \uc911\uad6d\uc5b4 \ub2e8\uc704 \ubd80\ubd84\uc744 \ud55c\uad6d\uc5b4\ub85c \uce58\ud658. \uc601\ubb38 \ub2e8\uc704\ub294 \uadf8\ub300\ub85c."""
            for cn, ko in _unit_ko_map.items():
                token = token.replace(cn, ko)
            return token

        def translate_if_chinese(text):
            if not text:
                return text
            if not re.search(r'[\u4e00-\u9fff]', str(text)):
                return text
            # 1) \uc6d0\ubcf8\uc5d0\uc11c \uc22b\uc790+\ub2e8\uc704 \ud1a0\ud070 \ucd94\ucd9c (\ud55c\uad6d\uc5b4 \ub2e8\uc704\ub85c \uce58\ud658\ud55c \ud615\ud0dc\ub85c \ubcf4\uad00)
            raw_tokens = []
            for m in _unit_pattern.finditer(str(text)):
                token = m.group()
                unit_m = re.search(r'(cm|mm|kg|lb|inch|\u82f1\u5bf8|\u5bf8|\u65a4|\u78c5)', token, re.IGNORECASE)
                if unit_m:
                    ko_token = _normalize_unit_token(token)
                    ko_unit = _normalize_unit_token(unit_m.group())
                    raw_tokens.append((ko_token, ko_unit))
            # 2) \ubc88\uc5ed
            try:
                time.sleep(0.2)
                translated = translator_ai.translate(text)
            except Exception:
                return text
            if not translated:
                return text
            if not raw_tokens:
                return translated
            # 3) \ud55c\uad6d\uc5b4 \ub2e8\uc704 \ud0a4\uc6cc\ub4dc\uac00 \uacb0\uacfc\uc5d0 \uc0b4\uc544\uc788\ub294\uc9c0 \uac80\uc0ac (\ub300\uc18c\ubb38\uc790 \ubb34\uc2dc)
            translated_lower = translated.lower()
            missing = []
            for ko_token, ko_unit in raw_tokens:
                if ko_unit.lower() not in translated_lower:
                    missing.append(ko_token)
            # 4) \ub204\ub77d\ub41c \ud1a0\ud070\ub9cc \uacb0\uacfc \ub4a4\uc5d0 \uacf5\ubc31 \uad6c\ubd84\uc73c\ub85c \ucca8\ubd80
            if missing:
                return f"{translated} {' '.join(missing)}".strip()
            return translated

        def analyze_seo_only(image_url, brand_name):
            if not model or not image_url: return None, None
            try:
                headers = {'User-Agent': 'Mozilla/5.0'}
                res = requests.get(image_url, headers=headers, timeout=10)
                if res.status_code != 200: return None, None
                img_data = Image.open(BytesIO(res.content))
                prompt = f"""
당신은 쿠팡 상위노출 전문 SEO 카피라이터입니다.
제공된 상품 이미지를 보고, 쿠팡에서 검색량이 가장 많은 핵심 키워드만 골라
**한 줄짜리 짧은 상품명**과 **연관 태그 10개**를 생성하세요.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
[상품명 작성 절대 규칙 — 위반 시 사용 불가]
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

▣ 형식
   - 반드시: '{brand_name} ' 으로 시작 (브랜드 뒤 공백 1칸)
   - 그 다음: **공백으로만 구분된 명사형 키워드 3~5개**
   - 전체 길이: **공백 포함 25~40자 이내** (절대 50자 초과 금지)

▣ 절대 금지
   ❌ 쉼표(,) 마침표(.) 슬래시(/) 같은 구두점 일절 사용 금지
   ❌ 조사 사용 금지: "을/를/이/가/은/는/의/로/으로/에서/위한/에/와/과"
   ❌ 동사·형용사·서술형 금지: "닦는" "사용하는" "편리한" "게으른" "위한"
   ❌ 색상 금지: 블랙/화이트/그레이/베이지/아이보리/색상/랜덤
   ❌ 수량/세트 금지: 1+1, 2개입, 세트, 묶음
   ❌ 사이즈 금지: 대형/중형/소형/대/중/소
   ❌ 어린이/아이/유아 관련 키워드 절대 금지
   ❌ 설명문장 금지 — "~를 위한 ~도구" 같은 표현 일절 사용 금지

▣ 권장
   ✅ 명사 키워드만 공백으로 나열 (쿠팡 검색 알고리즘 최적화)
   ✅ 핵심 카테고리 단어 1개 + 용도/기능/형태 키워드 2~3개 + 대표 카테고리 1개
   ✅ 검색량 높은 일반명사 우선

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
[좋은 예시 — 이 패턴을 그대로 따를 것]
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
   ✅ {brand_name} 유리창 물기제거 스퀴지 청소도구
   ✅ {brand_name} 데스크탑 미니걸레 청소용품
   ✅ {brand_name} 욕실 곰팡이제거 청소솔
   ✅ {brand_name} 주방 다용도 수납 정리함
   ✅ {brand_name} 접이식 빨래건조대 베란다

[나쁜 예시 — 절대 이렇게 만들지 말 것]
   ❌ {brand_name} 걸레 대신 미니걸레, 데스크탑 청소용품, 게으른 사람들을 위한 욕실 유리 닦는 도구
      → 이유: 너무 길고, 쉼표 있고, 조사 "위한" 있고, 서술형 "닦는" 있음
   ❌ {brand_name} 편리하게 사용하는 청소용 도구
      → 이유: 형용사 "편리하게", 서술형 "사용하는"

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
[연관 태그 작성 규칙]
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
   - 검색량 많은 명사 키워드 10개
   - 쉼표(,)로 구분
   - 활용 공간(주방/욕실/거실 등), 용도, 카테고리, 동의어 위주

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
[출력 형식 — JSON 외 일절 출력 금지]
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{{
    "name": "{brand_name} 명사1 명사2 명사3 명사4",
    "tags": "태그1,태그2,태그3,태그4,태그5,태그6,태그7,태그8,태그9,태그10"
}}
"""
                ai_res = model.generate_content([prompt, img_data])
                text = ai_res.text.strip()
                # 코드펜스(```json ... ```) 제거
                text = re.sub(r'^```(?:json)?\s*', '', text)
                text = re.sub(r'\s*```$', '', text)
                if "{" in text:
                    m = re.search(r'\{.*\}', text, re.DOTALL)
                    if m: text = m.group()
                try:
                    data = json.loads(text)
                except Exception as je:
                    print(f"    ⚠️ AI JSON 파싱 실패: {je} | raw 앞200자: {text[:200]}")
                    return None, None
                name = data.get("name", "") or ""
                tags = data.get("tags", "") or ""
                # tags 가 리스트로 올 경우 쉼표로 조인
                if isinstance(tags, list):
                    tags = ",".join(str(t).strip() for t in tags if str(t).strip())
                tags = str(tags).strip()
                # name 정리: 금지어 + 구두점 제거 + 공백 정리
                for c in ["블랙", "화이트", "그레이", "베이지", "아이보리", "색상", "랜덤"]:
                    name = name.replace(c, "")
                name = re.sub(r'[,./]', ' ', name)
                name = re.sub(r'\s+', ' ', name).strip()
                return name, tags
            except Exception as e:
                print(f"    ⚠️ AI 분석 예외: {e}")
                return None, None

        try:
            worksheet = doc.worksheet(TARGET_SHEET_NAME)
            all_values = worksheet.get_all_values()
            if len(all_values) < 2: raise ValueError("시트 데이터 없음")
            header = [str(h).strip() for h in all_values[0]]
            df = pd.DataFrame(all_values[1:], columns=header)

            def find_col(keywords):
                for i, h in enumerate(header):
                    for k in keywords:
                        if k in h.replace(" ", ""): return i
                return -1

            idx_prod_link = find_col(['상품링크'])
            idx_img_url = find_col(['대표이미지링크', '이미지링크'])
            idx_trans_name = find_col(['변환상품명'])
            idx_full_name = find_col(['전체옵션명'])
            idx_tags = find_col(['태그', '검색태그'])
            idx_kor_opt = find_col(['한글옵션명'])
            idx_opt1 = find_col(['옵션1'])
            idx_opt2 = find_col(['옵션2'])
            idx_path = find_col(['대표이미지경로', '이미지경로'])
            idx_copy_path = find_col(['COPY폴더경로', 'COPY'])
            idx_temp_file = find_col(['임시파일명'])

            if idx_trans_name == -1: raise Exception("'변환상품명' 열이 없습니다.")
            if idx_tags == -1:
                print("    ⚠️ 시트에 '태그' 열이 없어 AI 태그가 기록되지 않습니다. 시트에 '태그' 또는 '검색태그' 열을 추가하세요.")

            updates = []
            success_count = 0
            fail_count = 0
            trans_count = 0
            unchanged_list = []
            forbidden_char_list = []
            product_cache = {}

            print(f"📦 총 {len(df)}개 행 데이터를 처리합니다.")

            for idx, row in df.iterrows():
                if should_stop(): raise Stopped()
                row_num = idx + 2
                prod_link = str(row.iloc[idx_prod_link]).strip()
                img_url = str(row.iloc[idx_img_url]).strip()
                current_name = str(row.iloc[idx_trans_name]).strip()
                current_opt1 = str(row.iloc[idx_opt1]) if idx_opt1 != -1 else ""
                current_opt2 = str(row.iloc[idx_opt2]) if idx_opt2 != -1 else ""
                new_opt1 = translate_if_chinese(current_opt1)
                if new_opt1 != current_opt1:
                    col_letter = openpyxl.utils.get_column_letter(idx_opt1 + 1)
                    updates.append({'range': f"{col_letter}{row_num}", 'values': [[new_opt1]]})
                    print(f"    🇨🇳 [옵션1 번역] {current_opt1} -> {new_opt1}")
                    trans_count += 1
                    current_opt1 = new_opt1
                new_opt2 = translate_if_chinese(current_opt2)
                if new_opt2 != current_opt2:
                    col_letter = openpyxl.utils.get_column_letter(idx_opt2 + 1)
                    updates.append({'range': f"{col_letter}{row_num}", 'values': [[new_opt2]]})
                    print(f"    🇨🇳 [옵션2 번역] {current_opt2} -> {new_opt2}")
                    trans_count += 1
                    current_opt2 = new_opt2
                pattern = r'[\\/*?:"<>|]'
                if re.search(pattern, current_opt1):
                    forbidden_char_list.append(f"📍 {row_num}행 옵션1(J열): {current_opt1}")
                if re.search(pattern, current_opt2):
                    forbidden_char_list.append(f"📍 {row_num}행 옵션2(K열): {current_opt2}")
                if not img_url:
                    fail_count += 1
                    continue
                ai_name, ai_tags = None, None
                if prod_link in product_cache:
                    cached_data = product_cache[prod_link]
                    if cached_data:
                        ai_name = cached_data['name']
                        ai_tags = cached_data['tags']
                else:
                    ai_name, ai_tags = analyze_seo_only(img_url, MY_BRAND_NAME)
                    if ai_name:
                        product_cache[prod_link] = {'name': ai_name, 'tags': ai_tags}
                        print(f"    🤖 [API 호출] {ai_name}")
                    else:
                        product_cache[prod_link] = None
                if ai_name:
                    if re.search(pattern, ai_name):
                        forbidden_char_list.append(f"📍 {row_num}행 변환상품명(AI생성): {ai_name}")
                    if ai_name == current_name or len(ai_name) < 5:
                        unchanged_list.append(f"📍 {row_num}행: 상품명이 변경되지 않았거나 너무 짧음 ({ai_name})")
                    col_letter = openpyxl.utils.get_column_letter(idx_trans_name + 1)
                    updates.append({'range': f"{col_letter}{row_num}", 'values': [[ai_name]]})
                    if idx_tags != -1 and ai_tags:
                        col_letter = openpyxl.utils.get_column_letter(idx_tags + 1)
                        updates.append({'range': f"{col_letter}{row_num}", 'values': [[ai_tags]]})
                    if idx_full_name != -1:
                        full_name = f"{ai_name} {current_opt1} {current_opt2}".strip()
                        col_letter = openpyxl.utils.get_column_letter(idx_full_name + 1)
                        updates.append({'range': f"{col_letter}{row_num}", 'values': [[full_name]]})
                    success_count += 1
                else:
                    fail_count += 1
                    unchanged_list.append(f"📍 {row_num}행: AI 생성 실패 (기존 이름 유지)")
                    if re.search(pattern, current_name):
                        forbidden_char_list.append(f"📍 {row_num}행 변환상품명(기존값): {current_name}")

            if updates:
                # 보호 ③ 20260514: batch_update range 형식 검증.
                # 'G2' 같은 단일 셀 range 만 허용. 빈 문자열 / 컬럼 통째 range / 잘못된 형식이
                # 끼어 있으면 시트 통째 wipe 위험 — 통째 스킵.
                bad = [u for u in updates if not re.match(r'^[A-Z]+\d+$', str(u.get('range', '')))]
                if bad:
                    print(f"⚠️ [보호 ③] batch_update 에 잘못된 range {len(bad)}개 발견 — "
                          f"통째 스킵 (앞 3개: {bad[:3]})")
                else:
                    print("\n    ☁️ 구글 시트 업데이트 중 (번역 및 AI 적용)...")
                    worksheet.batch_update(updates)

            # 스마트 백업
            print("-" * 30)
            backup_sheet_name = "[백업]링크데이터"
            try:
                try: backup_sheet = doc.worksheet(backup_sheet_name)
                except Exception: backup_sheet = doc.add_worksheet(title=backup_sheet_name, rows=len(df)+50, cols=5)
                # 보호 ② 20260514: backup_sheet 가 진짜 백업 탭인지 확인 후 진행.
                # 어떤 이유로든 메인 탭(TARGET_SHEET_NAME) 을 가리키면 백업 단계 통째 스킵.
                if backup_sheet.title != backup_sheet_name:
                    print(f"⚠️ [보호 ②] backup_sheet 가 백업 탭 아님 "
                          f"(실제: '{backup_sheet.title}', 예상: '{backup_sheet_name}') — 백업 단계 스킵")
                else:
                    backup_sheet.clear()
                    backup_data = [["이미지링크", "대표이미지경로", "COPY폴더 경로", "변환상품명", "임시파일명"]]
                    for idx, row in df.iterrows():
                        img_link = str(row.iloc[idx_img_url]) if idx_img_url != -1 else ""
                        h_path = str(row.iloc[idx_path]) if idx_path != -1 else ""
                        i_copy = str(row.iloc[idx_copy_path]) if idx_copy_path != -1 else ""
                        c_name = str(row.iloc[idx_trans_name]) if idx_trans_name != -1 else ""
                        t_file = str(row.iloc[idx_temp_file]) if idx_temp_file != -1 else ""
                        backup_data.append([img_link, h_path, i_copy, c_name, t_file])
                    backup_sheet.update(range_name="A1", values=backup_data)
                    print(f"📦 스마트 백업 완료 ('{backup_sheet_name}' 시트)")
            except Exception as e:
                print(f"⚠️ 백업 중 경미한 오류: {e}")

            print("\n" + "="*60)
            if forbidden_char_list:
                print("⚠️ [경고] 아래 항목에 '특수문자'가 포함되어 있습니다.")
                for err in sorted(list(set(forbidden_char_list))):
                    print(err)
                print("-" * 30)
            print(f"✅ 번역 완료: {trans_count}개 옵션")
            print(f"✅ AI 적용 완료: {success_count}건")
            print("="*60)

        except Stopped:
            raise
        except Exception as e:
            print(f"❌ Cell 6 실행 중 오류: {e}")

        _t_end("Cell 6: AI 분석 + 번역")
        print("\n🎉 [1단계 전체 완료] 구글시트를 검수 후 체크리스트 3개를 확인하세요.")

        # 측정 결과 출력 (2026-05-15 추가)
        try:
            _t_report()
        except Exception:
            pass

    finally:
        sys.stdout = original_stdout
        try:
            if driver is not None:
                driver.quit()
        except Exception:
            pass
        os.chdir(original_cwd)
