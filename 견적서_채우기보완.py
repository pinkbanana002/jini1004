# -*- coding: utf-8 -*-
# 견적서_채우기보완.py
#   이미 만들어진 로켓 견적서(xlsx)에서 [비어있는 칸]을 자동으로 채웁니다.
#   - 이미 채워진 칸은 절대 건드리지 않습니다 (빈칸만 채움)
#   - 카테고리(B) / KC인증정보 / 크기·중량 / 제품구성 / 세부사양 / 수량·배터리·세트·설치 등
#   - 사용법:  이 파일을 견적서와 같은 폴더에 두고
#            .venv312\Scripts\python.exe 견적서_채우기보완.py "견적서파일.xlsx"
#            (파일명 없이 실행하면 폴더 안 '견적서' xlsx 를 자동으로 찾음)
import sys, os, glob
try: sys.stdout.reconfigure(encoding="utf-8")
except: pass
import openpyxl

def clean(s):
    return str(s or "").replace(" ", "").replace("\n", "").strip()

# ── 채움 규칙 (헤더에 포함된 키워드 → 값). 위에서부터 먼저 매칭 ──
#   값이 None 인 규칙은 '건드리지 않고 비워둠' (식별자/HTML 등)
RULES = [
    ("globaltrade", None), ("gtin", None),
    ("parentmanufacturer", None), ("manufacturerpart", None),
    ("html", None), ("소싱채널", None),
    ("카테고리", "__CATEGORY__"),
    ("kc인증정보", "해당사항없음"), ("인증정보", "해당사항없음"),
    ("크기,중량", "상세페이지 참조"), ("크기중량", "상세페이지 참조"),
    ("제품구성", "본품"), ("포함구성", "본품"), ("포함구성요소", "본품"),
    ("상품별세부사양", "상세페이지 참조"), ("세부사양", "상세페이지 참조"),
    ("운동부위", "상세페이지 참조"), ("기본재료", "상세페이지 참조"),
    ("배터리", "배터리 불필요"),
    ("세트여부", "세트 아님"), ("세트", "세트 아님"),
    ("설치지원방식", "__CLEAR__"), ("설치지원", "__CLEAR__"),   # 설치지원방식 열은 전부 비움(지우기)
    ("수량", "1개"),
]

def get_category_value(wb):
    # HQF 시트 등에서 'A>B>C' 형태 경로 목록을 모아 '기타' 항목을 우선 반환
    cands = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str) and v.count(">") >= 2:
                    cands.append(v.strip())
    # QF 시트명에서 leaf 추출 (예: '... _기타 운동기구' → '기타운동기구')
    leaf = ""
    for s in wb.sheetnames:
        if s.startswith("QF_"):
            leaf = clean(s.split("_")[-1])
    # 1) leaf 일치 우선
    for c in cands:
        last = clean(c.split(">")[-1].split("(")[0])
        if leaf and leaf in last:
            return c
    # 2) '기타' 포함 폴백
    for c in cands:
        if "기타" in c:
            return c
    return cands[0] if cands else None

def main():
    # 대상 파일 결정
    if len(sys.argv) > 1:
        target = sys.argv[1]
    else:
        found = [f for f in glob.glob("*.xlsx") if "견적서" in f and "~$" not in f]
        if not found:
            print("[중단] 폴더에서 '견적서' xlsx 를 못 찾았습니다. 파일명을 뒤에 붙여 실행하세요.")
            return
        target = found[0]
    print("대상 파일:", target)

    wb = openpyxl.load_workbook(target)

    # 데이터 시트 찾기 (상품명 있는 헤더행)
    ws = None; header_row = 5
    for sheet in wb.worksheets:
        for r in range(1, 15):
            vals = [str(c.value) for c in sheet[r] if c.value]
            if any("상품명" in v for v in vals):
                ws = sheet; header_row = r; break
        if ws: break
    if ws is None:
        print("[중단] 상품명 헤더를 못 찾음"); return
    print(f"데이터 시트: {ws.title!r}  헤더행: {header_row}")

    cat_value = get_category_value(wb)
    print("카테고리 값:", cat_value)

    # 상품명 열 찾기 → 데이터 행 범위
    name_col = None
    for c in range(1, ws.max_column+1):
        if "상품명" in str(ws.cell(header_row, c).value or ""):
            name_col = c; break
    start_row = header_row + 4
    last_row = start_row
    for r in range(start_row, ws.max_row+1):
        if ws.cell(r, name_col).value not in (None, ""):
            last_row = r
    print(f"데이터 행: {start_row} ~ {last_row}")

    # 헤더별 채움값 미리 계산
    col_fill = {}
    for c in range(1, ws.max_column+1):
        h = clean(ws.cell(header_row, c).value)
        if not h: continue
        for key, val in RULES:
            if key in h.lower() or key in h:
                col_fill[c] = None if val is None else (cat_value if val == "__CATEGORY__" else val)
                break

    # 채우기 / 지우기
    filled = {}; cleared = {}
    for r in range(start_row, last_row+1):
        for c, val in col_fill.items():
            if val is None: continue
            cur = ws.cell(r, c).value
            hh = ws.cell(header_row, c).value
            if val == "__CLEAR__":
                if cur not in (None, ""):
                    ws.cell(r, c).value = None
                    cleared[hh] = cleared.get(hh, 0) + 1
                continue
            if cur in (None, ""):
                ws.cell(r, c).value = val
                filled[hh] = filled.get(hh, 0) + 1

    # 저장
    out = target
    wb.save(out)
    print("\n[완료] 채운 열 요약:")
    for h, n in filled.items():
        print(f"  ✅ {h} : {n}칸")
    if not filled:
        print("  (채울 빈칸이 없었습니다)")
    if cleared:
        print("\n[비움] 지운 열:")
        for h, n in cleared.items():
            print(f"  🧹 {h} : {n}칸")
    print("\n저장:", out)

if __name__ == "__main__":
    main()
